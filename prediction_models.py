# -*- coding: utf-8 -*-
"""
多模型预测对比分析
- 用户退款预测（二分类）：XGBoost、RandomForest、LogisticRegression
- 销量预测（回归）：XGBoost、RandomForest、LinearRegression
- 可视化图表保存
- 模型指标对比保存
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
import os
import joblib
import json
import warnings
warnings.filterwarnings('ignore')

# 尝试导入必要的库
try:
    import xgboost as xgb
    from xgboost import XGBClassifier, XGBRegressor
except ImportError:
    print("请先安装 xgboost: pip install xgboost")
    exit(1)

try:
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    matplotlib.rcParams['axes.unicode_minus'] = False
except ImportError:
    print("请先安装 matplotlib: pip install matplotlib")
    exit(1)

from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.linear_model import LogisticRegression, LinearRegression, Ridge
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor, GradientBoostingClassifier
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    classification_report, confusion_matrix, roc_auc_score, roc_curve,
    mean_squared_error, mean_absolute_error, r2_score
)


# ============================================================
# 配置路径
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_DIR = os.path.join(BASE_DIR, 'models')
METRICS_DIR = os.path.join(BASE_DIR, 'model_metrics')
CHARTS_DIR = os.path.join(BASE_DIR, 'model_charts')

# 创建目录
for dir_path in [MODEL_DIR, METRICS_DIR, CHARTS_DIR]:
    os.makedirs(dir_path, exist_ok=True)


# ============================================================
# 数据加载函数
# ============================================================

def safe_float(v):
    """安全转换为浮点数"""
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        s = str(v).replace('元','').replace('¥','').replace(',','').replace(' ','').strip()
        return float(s) if s else 0.0
    except:
        return 0.0


def load_tmall_data(filepath):
    """加载天猫订单数据"""
    print("正在加载 tmall_order_report.xlsx...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            '订单编号': row[0],
            '总金额': safe_float(row[1]),
            '买家实际支付金额': safe_float(row[2]),
            '收货地址': str(row[3]).strip() if row[3] else '',
            '订单创建时间': row[4],
            '订单付款时间': row[5],
            '退款金额': safe_float(row[6]),
        })
    df = pd.DataFrame(rows)
    wb.close()
    print(f"  加载完成: {len(df)} 条记录")
    return df


def load_rihua_data(filepath):
    """加载日化销售数据"""
    print("正在加载 日化.xlsx...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # 销售订单表
    ws_sales = wb[wb.sheetnames[0]]
    sales_rows = []
    for row in ws_sales.iter_rows(min_row=2, values_only=True):
        sales_rows.append({
            '订单编码': row[0],
            '订单日期': row[1],
            '客户编码': row[2],
            '所在区域': row[3],
            '所在省份': row[4],
            '所在地市': row[5],
            '商品编号': row[6],
            '订购数量': safe_float(row[7]),
            '订购单价': safe_float(row[8]),
            '金额': safe_float(row[9]),
        })
    df_sales = pd.DataFrame(sales_rows)
    
    # 商品信息表
    ws_prod = wb[wb.sheetnames[1]]
    prod_rows = []
    for row in ws_prod.iter_rows(min_row=2, values_only=True):
        prod_rows.append({
            '商品编号': row[0],
            '商品名称': row[1],
            '商品小类': row[2],
            '商品大类': row[3],
            '销售单价': safe_float(row[4]),
        })
    df_prod = pd.DataFrame(prod_rows)
    wb.close()
    
    print(f"  销售订单: {len(df_sales)} 条")
    print(f"  商品信息: {len(df_prod)} 条")
    return df_sales, df_prod


# ============================================================
# 数据准备函数
# ============================================================

def prepare_refund_data(df_tmall):
    """准备退款预测数据"""
    print("\n准备退款预测数据...")
    
    df = df_tmall.copy()
    df['订单创建时间'] = pd.to_datetime(df['订单创建时间'], errors='coerce')
    df['订单付款时间'] = pd.to_datetime(df['订单付款时间'], errors='coerce')
    
    # 只保留已付款的订单
    df = df[df['买家实际支付金额'] > 0].copy()
    
    # 创建目标变量
    df['是否退款'] = (df['退款金额'] > 0).astype(int)
    
    # 特征工程
    df['下单小时'] = df['订单创建时间'].dt.hour
    df['下单星期'] = df['订单创建时间'].dt.weekday
    df['是否周末'] = df['下单星期'].apply(lambda x: 1 if x >= 5 else 0)
    df['下单月份'] = df['订单创建时间'].dt.month
    
    df['付款延迟(小时)'] = df.apply(
        lambda r: (r['订单付款时间'] - r['订单创建时间']).total_seconds() / 3600
        if pd.notna(r['订单付款时间']) and pd.notna(r['订单创建时间']) else 0,
        axis=1
    )
    
    df['金额段'] = pd.cut(df['总金额'], bins=[0, 50, 100, 200, 500, 1000, float('inf')],
                        labels=[0, 1, 2, 3, 4, 5]).astype(float).fillna(0)
    
    le_region = LabelEncoder()
    df['地区编码'] = le_region.fit_transform(df['收货地址'].fillna('未知'))
    
    region_refund_rate = df.groupby('收货地址')['是否退款'].mean()
    df['地区退款率'] = df['收货地址'].map(region_refund_rate).fillna(0.2)
    
    feature_cols = ['总金额', '买家实际支付金额', '下单小时', '是否周末', 
                   '付款延迟(小时)', '金额段', '地区编码', '地区退款率']
    
    X = df[feature_cols].fillna(0)
    y = df['是否退款']
    
    print(f"  特征数量: {len(feature_cols)}")
    print(f"  样本数量: {len(X)}")
    print(f"  退款比例: {y.mean()*100:.1f}%")
    
    return X, y, feature_cols, le_region


def prepare_sales_data(df_sales, df_prod):
    """准备销量预测数据（订单级别）"""
    print("\n准备销量预测数据...")
    
    # 合并商品信息
    df = df_sales.merge(
        df_prod[['商品编号', '商品小类', '商品大类', '销售单价']],
        on='商品编号', how='left'
    )
    
    # 类型转换
    df['订单日期'] = pd.to_datetime(df['订单日期'], errors='coerce')
    df['订购数量'] = pd.to_numeric(df['订购数量'], errors='coerce').fillna(0)
    df['订购单价'] = pd.to_numeric(df['订购单价'], errors='coerce').fillna(0)
    df['金额'] = pd.to_numeric(df['金额'], errors='coerce').fillna(0)
    
    # 过滤异常数据
    df = df[df['订单日期'].notna()]
    df = df[df['订单日期'].apply(lambda x: x.year < 2030 if pd.notna(x) else False)]
    df = df[df['订购数量'] > 0]  # 只保留有销量的订单
    
    # 特征工程
    # 时间特征
    df['年份'] = df['订单日期'].dt.year
    df['月'] = df['订单日期'].dt.month
    df['季度'] = df['月'].apply(lambda x: (x-1)//3 + 1)
    df['日'] = df['订单日期'].dt.day
    df['星期'] = df['订单日期'].dt.weekday
    df['是否周末'] = df['星期'].apply(lambda x: 1 if x >= 5 else 0)
    
    # 编码分类特征
    le_category = LabelEncoder()
    le_subcategory = LabelEncoder()
    le_region = LabelEncoder()
    le_province = LabelEncoder()
    
    df['大类编码'] = le_category.fit_transform(df['商品大类'].fillna('未知'))
    df['小类编码'] = le_subcategory.fit_transform(df['商品小类'].fillna('未知'))
    df['区域编码'] = le_region.fit_transform(df['所在区域'].fillna('未知'))
    df['省份编码'] = le_province.fit_transform(df['所在省份'].fillna('未知'))
    
    # 价格相关特征
    df['单价差异'] = df['订购单价'] - df['销售单价']
    df['折扣率'] = df['金额'] / (df['订购数量'] * df['销售单价'] + 1)
    
    # 选择特征
    feature_cols = ['年份', '月', '季度', '日', '是否周末', 
                   '大类编码', '小类编码', '区域编码', '省份编码',
                   '订购单价', '单价差异', '折扣率']
    
    X = df[feature_cols].fillna(0)
    y = df['订购数量']  # 目标：预测订单的订购数量
    
    print(f"  特征数量: {len(feature_cols)}")
    print(f"  样本数量: {len(X)}")
    print(f"  平均订购量: {y.mean():.1f}")
    print(f"  订购量范围: {y.min():.0f} ~ {y.max():.0f}")
    
    return X, y, feature_cols, le_category, le_subcategory


# ============================================================
# 多模型训练 - 退款预测
# ============================================================

def train_refund_models(X, y):
    """训练多个退款预测模型并对比"""
    print("\n" + "="*60)
    print("退款预测模型训练与对比")
    print("="*60)
    
    # 划分数据集
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42, stratify=y
    )
    
    print(f"\n数据划分:")
    print(f"  训练集: {len(X_train)} 条 ({len(X_train)/len(X)*100:.0f}%)")
    print(f"  测试集: {len(X_test)} 条 ({len(X_test)/len(X)*100:.0f}%)")
    
    # 定义多个模型
    models = {
        'XGBoost': XGBClassifier(
            n_estimators=100, max_depth=6, learning_rate=0.1,
            objective='binary:logistic', eval_metric='auc',
            use_label_encoder=False, random_state=42,
            scale_pos_weight=(len(y_train) - y_train.sum()) / y_train.sum()
        ),
        'RandomForest': RandomForestClassifier(
            n_estimators=100, max_depth=10, random_state=42,
            class_weight='balanced'
        ),
        'LogisticRegression': LogisticRegression(
            max_iter=1000, class_weight='balanced', random_state=42
        ),
        'GradientBoosting': GradientBoostingClassifier(
            n_estimators=100, max_depth=6, learning_rate=0.1, random_state=42
        )
    }
    
    # 存储结果
    results = {}
    predictions = {}
    
    for name, model in models.items():
        print(f"\n训练 {name} 模型...")
        
        # 训练
        model.fit(X_train, y_train)
        
        # 预测
        y_pred = model.predict(X_test)
        y_pred_proba = model.predict_proba(X_test)[:, 1] if hasattr(model, 'predict_proba') else None
        
        # 评估
        metrics = {
            'accuracy': accuracy_score(y_test, y_pred),
            'precision': precision_score(y_test, y_pred),
            'recall': recall_score(y_test, y_pred),
            'f1': f1_score(y_test, y_pred),
            'auc': roc_auc_score(y_test, y_pred_proba) if y_pred_proba is not None else None
        }
        
        results[name] = {
            'model': model,
            'metrics': metrics,
            'y_test': y_test,
            'y_pred': y_pred,
            'y_pred_proba': y_pred_proba
        }
        
        predictions[name] = y_pred_proba if y_pred_proba is not None else y_pred
        
        print(f"  准确率: {metrics['accuracy']:.4f}")
        print(f"  精确率: {metrics['precision']:.4f}")
        print(f"  召回率: {metrics['recall']:.4f}")
        print(f"  F1分数: {metrics['f1']:.4f}")
        if metrics['auc']:
            print(f"  AUC: {metrics['auc']:.4f}")
    
    return results, X_test, y_test


# ============================================================
# 多模型训练 - 销量预测
# ============================================================

def train_sales_models(X, y):
    """训练多个销量预测模型并对比"""
    print("\n" + "="*60)
    print("销量预测模型训练与对比")
    print("="*60)
    
    # 划分数据集
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42
    )
    
    print(f"\n数据划分:")
    print(f"  训练集: {len(X_train)} 条 ({len(X_train)/len(X)*100:.0f}%)")
    print(f"  测试集: {len(X_test)} 条 ({len(X_test)/len(X)*100:.0f}%)")
    
    # 定义多个模型
    models = {
        'XGBoost': XGBRegressor(
            n_estimators=100, max_depth=6, learning_rate=0.1,
            objective='reg:squarederror', random_state=42
        ),
        'RandomForest': RandomForestRegressor(
            n_estimators=100, max_depth=10, random_state=42
        ),
        'LinearRegression': LinearRegression(),
        'Ridge': Ridge(alpha=1.0)
    }
    
    # 存储结果
    results = {}
    predictions = {}
    
    for name, model in models.items():
        print(f"\n训练 {name} 模型...")
        
        # 训练
        model.fit(X_train, y_train)
        
        # 预测
        y_pred = model.predict(X_test)
        
        # 评估
        metrics = {
            'mse': mean_squared_error(y_test, y_pred),
            'rmse': np.sqrt(mean_squared_error(y_test, y_pred)),
            'mae': mean_absolute_error(y_test, y_pred),
            'r2': r2_score(y_test, y_pred)
        }
        
        results[name] = {
            'model': model,
            'metrics': metrics,
            'y_test': y_test,
            'y_pred': y_pred
        }
        
        predictions[name] = y_pred
        
        print(f"  MSE: {metrics['mse']:.2f}")
        print(f"  RMSE: {metrics['rmse']:.2f}")
        print(f"  MAE: {metrics['mae']:.2f}")
        print(f"  R²: {metrics['r2']:.4f}")
    
    return results, X_test, y_test


# ============================================================
# 可视化函数
# ============================================================

def plot_refund_comparison(results, y_test, save_dir):
    """绘制退款预测模型对比图表"""
    print("\n生成退款预测可视化图表...")
    
    # 1. ROC曲线对比图
    fig, ax = plt.subplots(figsize=(10, 8))
    
    for name, data in results.items():
        if data['y_pred_proba'] is not None:
            fpr, tpr, _ = roc_curve(y_test, data['y_pred_proba'])
            auc = data['metrics']['auc']
            ax.plot(fpr, tpr, label=f"{name} (AUC={auc:.3f})", linewidth=2)
    
    ax.plot([0, 1], [0, 1], 'k--', label='随机猜测', linewidth=1)
    ax.set_xlabel('假正率 (FPR)', fontsize=12)
    ax.set_ylabel('真正率 (TPR)', fontsize=12)
    ax.set_title('退款预测模型 ROC 曲线对比', fontsize=14, fontweight='bold')
    ax.legend(loc='lower right', fontsize=10)
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'refund_roc_comparison.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  ✓ refund_roc_comparison.png")
    
    # 2. 模型指标对比柱状图
    fig, ax = plt.subplots(figsize=(12, 6))
    
    model_names = list(results.keys())
    metrics_names = ['accuracy', 'precision', 'recall', 'f1', 'auc']
    metrics_labels = ['准确率', '精确率', '召回率', 'F1分数', 'AUC']
    
    x = np.arange(len(model_names))
    width = 0.15
    
    for i, (metric, label) in enumerate(zip(metrics_names, metrics_labels)):
        values = [results[name]['metrics'].get(metric, 0) for name in model_names]
        ax.bar(x + i*width, values, width, label=label)
    
    ax.set_xlabel('模型', fontsize=12)
    ax.set_ylabel('分数', fontsize=12)
    ax.set_title('退款预测模型指标对比', fontsize=14, fontweight='bold')
    ax.set_xticks(x + width*2)
    ax.set_xticklabels(model_names, fontsize=11)
    ax.legend(loc='upper right', fontsize=10)
    ax.set_ylim(0, 1)
    ax.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'refund_metrics_comparison.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  ✓ refund_metrics_comparison.png")
    
    # 3. 混淆矩阵（每个模型）
    for name, data in results.items():
        fig, ax = plt.subplots(figsize=(6, 5))
        cm = confusion_matrix(y_test, data['y_pred'])
        
        im = ax.imshow(cm, cmap='Blues')
        ax.set_xlabel('预测结果', fontsize=12)
        ax.set_ylabel('实际结果', fontsize=12)
        ax.set_title(f'{name} 混淆矩阵', fontsize=14, fontweight='bold')
        ax.set_xticks([0, 1])
        ax.set_xticklabels(['未退款', '退款'])
        ax.set_yticks([0, 1])
        ax.set_yticklabels(['未退款', '退款'])
        
        # 显示数值
        for i in range(2):
            for j in range(2):
                ax.text(j, i, str(cm[i, j]), ha='center', va='center', fontsize=14, color='white' if cm[i,j] > cm.max()/2 else 'black')
        
        plt.colorbar(im, ax=ax)
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, f'refund_confusion_matrix_{name}.png'), dpi=150, bbox_inches='tight')
        plt.close()
    print(f"  ✓ 混淆矩阵图表已生成")


def plot_sales_comparison(results, y_test, save_dir):
    """绘制销量预测模型对比图表"""
    print("\n生成销量预测可视化图表...")
    
    # 1. 预测值 vs 实际值散点图
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    axes = axes.flatten()
    
    for i, (name, data) in enumerate(results.items()):
        ax = axes[i]
        y_pred = data['y_pred']
        
        ax.scatter(y_test, y_pred, alpha=0.5, s=20)
        ax.plot([y_test.min(), y_test.max()], [y_test.min(), y_test.max()], 'r--', linewidth=2)
        ax.set_xlabel('实际销量', fontsize=10)
        ax.set_ylabel('预测销量', fontsize=10)
        ax.set_title(f'{name} (R²={data["metrics"]["r2"]:.3f})', fontsize=12, fontweight='bold')
        ax.grid(True, alpha=0.3)
    
    plt.suptitle('销量预测模型对比：预测值 vs 实际值', fontsize=14, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'sales_prediction_comparison.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  ✓ sales_prediction_comparison.png")
    
    # 2. 模型指标对比柱状图
    fig, ax = plt.subplots(figsize=(10, 6))
    
    model_names = list(results.keys())
    metrics_names = ['rmse', 'mae', 'r2']
    metrics_labels = ['RMSE', 'MAE', 'R²']
    
    x = np.arange(len(model_names))
    width = 0.25
    
    colors = ['#e74c3c', '#f39c12', '#2ecc71']
    
    for i, (metric, label, color) in enumerate(zip(metrics_names, metrics_labels, colors)):
        values = [results[name]['metrics'][metric] for name in model_names]
        ax.bar(x + i*width, values, width, label=label, color=color)
    
    ax.set_xlabel('模型', fontsize=12)
    ax.set_ylabel('指标值', fontsize=12)
    ax.set_title('销量预测模型指标对比', fontsize=14, fontweight='bold')
    ax.set_xticks(x + width)
    ax.set_xticklabels(model_names, fontsize=11)
    ax.legend(loc='upper right', fontsize=10)
    ax.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'sales_metrics_comparison.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  ✓ sales_metrics_comparison.png")
    
    # 3. 残差分布图
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    axes = axes.flatten()
    
    for i, (name, data) in enumerate(results.items()):
        ax = axes[i]
        residuals = y_test - data['y_pred']
        
        ax.hist(residuals, bins=30, edgecolor='black', alpha=0.7)
        ax.axvline(x=0, color='red', linestyle='--', linewidth=2)
        ax.set_xlabel('残差 (实际 - 预测)', fontsize=10)
        ax.set_ylabel('频数', fontsize=10)
        ax.set_title(f'{name} 残差分布', fontsize=12, fontweight='bold')
        ax.grid(True, alpha=0.3)
    
    plt.suptitle('销量预测模型残差分布', fontsize=14, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'sales_residual_distribution.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  ✓ sales_residual_distribution.png")


# ============================================================
# 保存模型和指标
# ============================================================

def save_models_and_metrics(refund_results, sales_results, encoders):
    """保存所有模型和指标"""
    print("\n保存模型和指标...")
    
    # 保存退款预测模型
    for name, data in refund_results.items():
        model_path = os.path.join(MODEL_DIR, f'refund_{name.lower()}.joblib')
        joblib.dump(data['model'], model_path)
        print(f"  ✓ {model_path}")
    
    # 保存销量预测模型
    for name, data in sales_results.items():
        model_path = os.path.join(MODEL_DIR, f'sales_{name.lower()}.joblib')
        joblib.dump(data['model'], model_path)
        print(f"  ✓ {model_path}")
    
    # 保存编码器
    joblib.dump(encoders, os.path.join(MODEL_DIR, 'encoders.joblib'))
    print(f"  ✓ encoders.joblib")
    
    # 保存退款预测指标
    refund_metrics_df = pd.DataFrame({
        name: data['metrics'] for name, data in refund_results.items()
    }).T
    refund_metrics_df.to_csv(os.path.join(METRICS_DIR, 'refund_metrics.csv'), encoding='utf-8-sig')
    refund_metrics_df.to_json(os.path.join(METRICS_DIR, 'refund_metrics.json'), orient='index')
    print(f"  ✓ refund_metrics.csv/json")
    
    # 保存销量预测指标
    sales_metrics_df = pd.DataFrame({
        name: data['metrics'] for name, data in sales_results.items()
    }).T
    sales_metrics_df.to_csv(os.path.join(METRICS_DIR, 'sales_metrics.csv'), encoding='utf-8-sig')
    sales_metrics_df.to_json(os.path.join(METRICS_DIR, 'sales_metrics.json'), orient='index')
    print(f"  ✓ sales_metrics.csv/json")
    
    # 保存模型对比总结
    summary = {
        '退款预测': {
            name: data['metrics'] for name, data in refund_results.items()
        },
        '销量预测': {
            name: data['metrics'] for name, data in sales_results.items()
        }
    }
    
    with open(os.path.join(METRICS_DIR, 'model_summary.json'), 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"  ✓ model_summary.json")


# ============================================================
# 生成报告
# ============================================================

def generate_report(refund_results, sales_results):
    """生成模型对比报告"""
    print("\n" + "="*60)
    print("模型对比报告")
    print("="*60)
    
    # 退款预测最佳模型
    print("\n【退款预测】最佳模型:")
    best_refund = max(refund_results.items(), key=lambda x: x[1]['metrics']['f1'])
    print(f"  模型: {best_refund[0]}")
    print(f"  F1分数: {best_refund[1]['metrics']['f1']:.4f}")
    print(f"  AUC: {best_refund[1]['metrics']['auc']:.4f}")
    
    # 销量预测最佳模型
    print("\n【销量预测】最佳模型:")
    best_sales = max(sales_results.items(), key=lambda x: x[1]['metrics']['r2'])
    print(f"  模型: {best_sales[0]}")
    print(f"  R²: {best_sales[1]['metrics']['r2']:.4f}")
    print(f"  RMSE: {best_sales[1]['metrics']['rmse']:.2f}")
    
    # 详细对比表格
    print("\n【退款预测模型详细对比】")
    refund_df = pd.DataFrame({
        name: {
            '准确率': data['metrics']['accuracy'],
            '精确率': data['metrics']['precision'],
            '召回率': data['metrics']['recall'],
            'F1分数': data['metrics']['f1'],
            'AUC': data['metrics']['auc']
        } for name, data in refund_results.items()
    }).T.round(4)
    print(refund_df.to_string())
    
    print("\n【销量预测模型详细对比】")
    sales_df = pd.DataFrame({
        name: {
            'MSE': data['metrics']['mse'],
            'RMSE': data['metrics']['rmse'],
            'MAE': data['metrics']['mae'],
            'R²': data['metrics']['r2']
        } for name, data in sales_results.items()
    }).T.round(4)
    print(sales_df.to_string())


# ============================================================
# 主函数
# ============================================================

def main():
    """主流程"""
    print("=" * 60)
    print("多模型预测对比分析")
    print("=" * 60)
    
    # ========== 1. 数据加载 ==========
    print("\n【步骤1】数据加载")
    print("-" * 40)
    
    df_tmall = load_tmall_data(os.path.join(BASE_DIR, 'tmall_order_report.xlsx'))
    df_sales, df_prod = load_rihua_data(os.path.join(BASE_DIR, '日化.xlsx'))
    
    # ========== 2. 数据准备 ==========
    print("\n【步骤2】数据准备")
    print("-" * 40)
    
    X_refund, y_refund, refund_features, le_region = prepare_refund_data(df_tmall)
    X_sales, y_sales, sales_features, le_category, le_subcategory = prepare_sales_data(df_sales, df_prod)
    
    # ========== 3. 模型训练 ==========
    print("\n【步骤3】多模型训练")
    print("-" * 40)
    
    refund_results, X_refund_test, y_refund_test = train_refund_models(X_refund, y_refund)
    sales_results, X_sales_test, y_sales_test = train_sales_models(X_sales, y_sales)
    
    # ========== 4. 可视化 ==========
    print("\n【步骤4】生成可视化图表")
    print("-" * 40)
    
    plot_refund_comparison(refund_results, y_refund_test, CHARTS_DIR)
    plot_sales_comparison(sales_results, y_sales_test, CHARTS_DIR)
    
    # ========== 5. 保存模型和指标 ==========
    print("\n【步骤5】保存模型和指标")
    print("-" * 40)
    
    encoders = {
        'refund_region': le_region,
        'category': le_category,
        'subcategory': le_subcategory,
        'refund_features': refund_features,
        'sales_features': sales_features
    }
    
    save_models_and_metrics(refund_results, sales_results, encoders)
    
    # ========== 6. 生成报告 ==========
    generate_report(refund_results, sales_results)
    
    # ========== 7. 总结 ==========
    print("\n" + "=" * 60)
    print("完成！输出文件位置:")
    print("=" * 60)
    print(f"\n模型文件 ({MODEL_DIR}):")
    print("  - refund_xgboost.joblib, refund_randomforest.joblib, ...")
    print("  - sales_xgboost.joblib, sales_randomforest.joblib, ...")
    print(f"\n指标文件 ({METRICS_DIR}):")
    print("  - refund_metrics.csv/json")
    print("  - sales_metrics.csv/json")
    print("  - model_summary.json")
    print(f"\n图表文件 ({CHARTS_DIR}):")
    print("  - refund_roc_comparison.png")
    print("  - refund_metrics_comparison.png")
    print("  - refund_confusion_matrix_*.png")
    print("  - sales_prediction_comparison.png")
    print("  - sales_metrics_comparison.png")
    print("  - sales_residual_distribution.png")
    
    return refund_results, sales_results, encoders


if __name__ == "__main__":
    refund_results, sales_results, encoders = main()