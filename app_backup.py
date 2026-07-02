# -*- coding: utf-8 -*-
"""日化美妆产业链分析平台 - L4系统化答辩项目"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from collections import Counter
from datetime import datetime
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="日化美妆产业链分析平台",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# PROVINCE NAME MAPPING (for GeoJSON matching)
# ============================================================
def normalize_province(name):
    """Normalize province names to match DataV GeoJSON format"""
    if not isinstance(name, str): return name
    name = name.strip()
    mapping = {
        '北京': '北京市', '上海': '上海市', '天津': '天津市', '重庆': '重庆市',
        '广东省': '广东省', '浙江省': '浙江省', '江苏省': '江苏省',
        '山东省': '山东省', '四川省': '四川省', '福建省': '福建省',
        '安徽省': '安徽省', '湖南省': '湖南省', '辽宁省': '辽宁省',
        '河北省': '河北省', '河南省': '河南省', '江西省': '江西省',
        '云南省': '云南省', '贵州省': '贵州省', '湖北省': '湖北省',
        '陕西省': '陕西省', '山西省': '山西省', '甘肃省': '甘肃省',
        '吉林省': '吉林省', '黑龙江省': '黑龙江省', '海南省': '海南省',
        '青海省': '青海省', '台湾省': '台湾省',
        '内蒙古自治区': '内蒙古自治区',
        '广西壮族自治区': '广西壮族自治区',
        '新疆维吾尔自治区': '新疆维吾尔自治区',
        '宁夏回族自治区': '宁夏回族自治区',
        '西藏自治区': '西藏自治区',
        '香港特别行政区': '香港特别行政区',
        '澳门特别行政区': '澳门特别行政区',
    }
    return mapping.get(name, name)

# ============================================================
# UTILITIES
# ============================================================
def safe_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        s = str(v).replace('元','').replace('¥','').replace(',','').replace(' ','').strip()
        return float(s) if s else 0.0
    except:
        return 0.0

# Province coordinates for map viz (approximate centroids)
PROVINCE_COORDS = {
    '北京': (116.4, 39.9), '上海': (121.5, 31.2), '天津': (117.2, 39.1),
    '重庆': (106.5, 29.5), '广东省': (113.3, 23.1), '浙江省': (120.2, 30.3),
    '江苏省': (118.8, 32.1), '山东省': (117.0, 36.7), '四川省': (104.1, 30.6),
    '福建省': (119.3, 26.1), '安徽省': (117.3, 31.9), '湖南省': (113.0, 28.2),
    '辽宁省': (123.4, 41.8), '河北省': (114.5, 38.0), '河南省': (113.7, 34.8),
    '江西省': (115.9, 28.7), '云南省': (102.7, 25.0), '贵州省': (106.7, 26.6),
    '广西壮族自治区': (108.3, 22.8), '湖北省': (114.4, 30.6), '陕西省': (108.9, 34.3),
    '山西省': (112.6, 37.9), '甘肃省': (103.8, 36.1), '吉林省': (125.3, 43.9),
    '黑龙江省': (126.6, 45.8), '新疆维吾尔自治区': (87.6, 43.8), '海南省': (110.3, 20.0),
    '内蒙古自治区': (111.7, 40.8), '宁夏回族自治区': (106.3, 38.5),
    '青海省': (101.8, 36.6), '西藏自治区': (91.1, 29.6),
    '香港特别行政区': (114.2, 22.3), '台湾省': (121.0, 23.5),
}

# ============================================================
# DATA LOADING (cached)
# ============================================================
@st.cache_data
def load_tmall_data(filepath):
    """Load 天猫订单报告"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        create_dt = row[4] if isinstance(row[4], datetime) else None
        pay_dt = row[5] if isinstance(row[5], datetime) else None
        rows.append({
            '订单编号': str(row[0]) if row[0] else '',
            '总金额': safe_float(row[1]),
            '实付金额': safe_float(row[2]),
            '收货地址': str(row[3]).strip() if row[3] else '',
            '创建时间': create_dt,
            '付款时间': pay_dt,
            '退款金额': safe_float(row[6]),
        })
    return pd.DataFrame(rows)

@st.cache_data
def load_beauty_data(filepath):
    """Load 双十一淘宝美妆数据"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        brand = str(row[6]).strip() if row[6] else 'Unknown'
        price = safe_float(row[3])
        sale = int(row[4]) if row[4] else 0
        comment = int(row[5]) if row[5] else 0
        rows.append({
            'update_time': row[0],
            'id': str(row[1]) if row[1] else '',
            'title': str(row[2]) if row[2] else '',
            'price': price,
            'sale_count': sale,
            'comment_count': comment,
            '品牌': brand,
            'gmv': price * sale,
        })
    return pd.DataFrame(rows)

@st.cache_data
def load_rihua_data(filepath):
    """Load 日化数据 (sales detail + product info)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # Sheet 0: 销售明细
    ws_sales = wb[wb.sheetnames[0]]
    sales_rows = []
    for row in ws_sales.iter_rows(min_row=2, values_only=True):
        dt_val = row[1] if isinstance(row[1], datetime) else None
        sales_rows.append({
            '订单编码': str(row[0]) if row[0] else '',
            '订单日期': dt_val,
            '客户编码': str(row[2]).strip() if row[2] else '',
            '所在区域': str(row[3]).strip() if row[3] else '',
            '所在省份': str(row[4]).strip() if row[4] else '',
            '所在地市': str(row[5]).strip() if row[5] else '',
            '商品编号': str(row[6]).strip() if row[6] else '',
            '订购数量': safe_float(row[7]),
            '订购单价': safe_float(row[8]),
            '金额': safe_float(row[9]),
        })
    df_sales = pd.DataFrame(sales_rows)

    # Remove the 2050 outlier
    df_sales = df_sales[df_sales['订单日期'].notna()]
    df_sales = df_sales[df_sales['订单日期'].apply(lambda x: x.year < 2030 if isinstance(x, datetime) else True)]

    # Sheet 1: 商品信息表
    ws_prod = wb[wb.sheetnames[1]]
    prod_rows = []
    for row in ws_prod.iter_rows(min_row=2, values_only=True):
        prod_rows.append({
            '商品编号': str(row[0]).strip() if row[0] else '',
            '商品名称': str(row[1]).strip() if row[1] else '',
            '商品小类': str(row[2]).strip() if row[2] else '',
            '商品大类': str(row[3]).strip() if row[3] else '',
            '销售单价': safe_float(row[4]),
        })
    df_prod = pd.DataFrame(prod_rows)

    return df_sales, df_prod

# ============================================================
# ANALYSIS FUNCTIONS
# ============================================================
def compute_brand_scores(df):
    """Compute brand competitiveness scores (6 dimensions)"""
    brands = df['品牌'].unique()
    scores = []
    all_avg_price = df['price'].mean()
    all_avg_sales = df['sale_count'].mean()

    for brand in brands:
        bd = df[df['品牌'] == brand]
        skus = len(bd)
        avg_price = bd['price'].mean()
        total_gmv = bd['gmv'].sum() / 1e8  # in 100M
        total_sales = bd['sale_count'].sum()
        comment_rate = bd['comment_count'].sum() / max(total_sales, 1)
        hit_rate = len(bd[bd['sale_count'] > 10000]) / max(skus, 1)

        # Price coverage: count of price bands covered
        bands = {'0-50':0, '50-100':0, '100-200':0, '200-500':0, '500-1000':0, '1000+':0}
        for p in bd['price']:
            if p <= 50: bands['0-50'] += 1
            elif p <= 100: bands['50-100'] += 1
            elif p <= 200: bands['100-200'] += 1
            elif p <= 500: bands['200-500'] += 1
            elif p <= 1000: bands['500-1000'] += 1
            else: bands['1000+'] += 1
        coverage = sum(1 for v in bands.values() if v > 0)

        scores.append({
            '品牌': brand,
            'SKU数': skus,
            '均售价(¥)': round(avg_price, 0),
            'GMV(亿)': round(total_gmv, 2),
            '总销量(万)': round(total_sales/1e4, 1),
            '评论率(%)': round(comment_rate*100, 1),
            '爆款率(%)': round(hit_rate*100, 1),
            '价格带覆盖': coverage,
            '品牌溢价指数': round(avg_price / max(all_avg_price, 1) * 100, 0),
            '销量效率指数': round((total_sales/skus) / max(all_avg_sales, 1) * 100, 0),
        })

    return pd.DataFrame(scores)

def compute_rfm(df):
    """Compute RFM for each customer"""
    now = df['订单日期'].max()
    if not isinstance(now, datetime):
        now = datetime(2025, 10, 1)

    rfm_list = []
    for cust, grp in df.groupby('客户编码'):
        recency = (now - grp['订单日期'].max()).days
        frequency = len(grp['订单编码'].unique())
        monetary = grp['金额'].sum()
        rfm_list.append({
            '客户编码': cust,
            'Recency(天)': recency,
            'Frequency': frequency,
            'Monetary(¥)': monetary,
        })

    rfm_df = pd.DataFrame(rfm_list)
    if len(rfm_df) >= 3:
        rfm_df['R_Score'] = pd.qcut(rfm_df['Recency(天)'], 4, labels=[4,3,2,1]).astype(int)
        rfm_df['F_Score'] = pd.qcut(rfm_df['Frequency'].rank(method='first'), 4, labels=[1,2,3,4]).astype(int)
        rfm_df['M_Score'] = pd.qcut(rfm_df['Monetary(¥)'].rank(method='first'), 4, labels=[1,2,3,4]).astype(int)
        rfm_df['RFM总分'] = rfm_df['R_Score'] + rfm_df['F_Score'] + rfm_df['M_Score']

        def tier(r, f, m):
            total = r + f + m
            if total >= 10: return '核心客户'
            elif total >= 7: return '重要客户'
            elif total >= 5: return '一般客户'
            else: return '流失风险'

        rfm_df['客户层级'] = rfm_df.apply(lambda x: tier(x['R_Score'], x['F_Score'], x['M_Score']), axis=1)
    else:
        rfm_df['R_Score'] = 0; rfm_df['F_Score'] = 0; rfm_df['M_Score'] = 0
        rfm_df['RFM总分'] = 0; rfm_df['客户层级'] = '数据不足'

    return rfm_df

def compute_monthly_trend(df):
    """Compute monthly sales trend with MoM growth"""
    df = df.copy()
    df['月份'] = df['订单日期'].apply(lambda x: f'{x.year}-{x.month:02d}' if isinstance(x, datetime) else None)
    monthly = df.groupby('月份').agg(
        销售额=('金额', 'sum'),
        订单数=('订单编码', 'nunique'),
        销售数量=('订购数量', 'sum'),
    ).reset_index().dropna()
    monthly['环比增长(%)'] = monthly['销售额'].pct_change() * 100
    return monthly

def compute_province_stats(df):
    """Compute province-level statistics with coordinates for mapping"""
    province_stats = df.groupby('所在省份').agg(
        销售额=('金额', 'sum'),
        订单数=('订单编码', 'nunique'),
        客户数=('客户编码', 'nunique'),
        销售数量=('订购数量', 'sum'),
    ).reset_index()
    province_stats['客单价'] = province_stats['销售额'] / province_stats['客户数']
    province_stats['lat'] = province_stats['所在省份'].map(lambda x: PROVINCE_COORDS.get(x, (None, None))[1])
    province_stats['lon'] = province_stats['所在省份'].map(lambda x: PROVINCE_COORDS.get(x, (None, None))[0])
    province_stats = province_stats.dropna(subset=['lat', 'lon'])
    return province_stats

def compute_region_tree(df):
    """Compute region → province hierarchy for treemap"""
    # Clean region names: merge near-duplicates
    df_clean = df.copy()
    region_map = {}
    for r in df_clean['所在区域'].unique():
        cleaned = str(r).strip().replace(' ', '')
        region_map[r] = cleaned
    df_clean['区域_clean'] = df_clean['所在区域'].map(region_map)

    data = []
    for (reg, prov), grp in df_clean.groupby(['区域_clean', '所在省份']):
        amt = round(grp['金额'].sum() / 1e8, 2)
        if amt <= 0:
            continue
        data.append({
            '区域': reg,
            '省份': prov,
            '销售额(亿)': amt,
            '客户数': grp['客户编码'].nunique(),
        })
    result = pd.DataFrame(data)
    # Drop rows that may cause treemap issues
    result = result[result['销售额(亿)'] > 0]
    return result

# ============================================================
# VISUALIZATION FUNCTIONS
# ============================================================
def plot_brand_gmv_bar(scores_df, n=15, sort_by='GMV(亿)'):
    """Horizontal bar chart of brand ranking, sortable"""
    top = scores_df.nlargest(n, sort_by).sort_values(sort_by)
    fig = px.bar(
        top, x=sort_by, y='品牌', orientation='h',
        text=sort_by, color=sort_by,
        color_continuous_scale='Blues',
        title=f'Top {n} 品牌排名（按{sort_by}）'
    )
    if 'GMV' in sort_by or '金额' in sort_by:
        fig.update_traces(texttemplate='¥%{text:.1f}亿', textposition='outside')
    elif '率' in sort_by:
        fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
    elif '价' in sort_by:
        fig.update_traces(texttemplate='¥%{text:.0f}', textposition='outside')
    else:
        fig.update_traces(texttemplate='%{text}', textposition='outside')
    fig.update_layout(height=500, yaxis={'categoryorder': 'total ascending'})
    return fig

def plot_brand_radar(scores_df, selected_brands):
    """Radar chart for selected brands"""
    dimensions = ['均售价(¥)', 'GMV(亿)', 'SKU数', '爆款率(%)', '评论率(%)', '价格带覆盖']

    # Normalize each dimension to 0-100
    norm_df = scores_df.copy()
    for dim in dimensions:
        max_v = max(norm_df[dim].max(), 1)
        norm_df[f'{dim}_norm'] = norm_df[dim] / max_v * 100

    fig = go.Figure()
    for brand in selected_brands:
        bd = norm_df[norm_df['品牌'] == brand]
        if len(bd) == 0: continue
        values = [bd[f'{dim}_norm'].values[0] for dim in dimensions]
        values_raw = [bd[dim].values[0] for dim in dimensions]

        fig.add_trace(go.Scatterpolar(
            r=values,
            theta=dimensions,
            fill='toself',
            name=f'{brand}',
            hovertemplate='%{theta}: %{customdata}<extra></extra>',
            customdata=[f'{v:.1f}' for v in values_raw],
        ))

    fig.update_layout(
        polar=dict(radialaxis=dict(range=[0, 105], showticklabels=False)),
        height=500,
        title='品牌竞争力雷达图（归一化对比）'
    )
    return fig

def plot_price_vs_sales_scatter(df, selected_brands):
    """Scatter plot: price vs sales, bubble size = GMV"""
    plot_df = df[df['品牌'].isin(selected_brands)].copy()
    plot_df['gmv_m'] = plot_df['gmv'] / 1e6

    fig = px.scatter(
        plot_df, x='price', y='sale_count', size='gmv_m',
        color='品牌', hover_name='title',
        log_y=True,
        title='价格-销量散点图（气泡大小=GMV）',
        labels={'price': '价格(¥)', 'sale_count': '销量(件)', 'gmv_m': 'GMV(百万)'},
        height=550,
    )
    fig.update_layout(legend=dict(orientation='h', y=-0.2))
    return fig

def plot_province_map(prov_stats):
    """China province bubble map"""
    china_provs = prov_stats[
        (prov_stats['lat'] >= 18) & (prov_stats['lat'] <= 54) &
        (prov_stats['lon'] >= 73) & (prov_stats['lon'] <= 135)
    ].copy()

    if len(china_provs) == 0:
        fig = go.Figure()
        fig.update_layout(title='省域销售分布地图（无数据）', height=400)
        return fig

    max_sales = china_provs['销售额'].max()
    sizes = [18 + (s / max_sales) * 42 for s in china_provs['销售额']]

    fig = go.Figure()
    fig.add_trace(go.Scattergeo(
        lat=china_provs['lat'], lon=china_provs['lon'],
        marker=dict(
            size=sizes, color=china_provs['销售额'],
            colorscale='Reds', showscale=True,
            colorbar=dict(title='销售额(¥)', x=0.85, len=0.7),
            line=dict(width=1.5, color='#c0392b'),
        ),
        text=china_provs.apply(
            lambda r: f"<b>{r['所在省份']}</b><br>销售额: ¥{r['销售额']/1e8:.2f}亿<br>客户数: {int(r['客户数'])}",
            axis=1
        ),
        hoverinfo='text', mode='markers',
    ))
    fig.update_geos(
        scope='asia', center=dict(lat=35, lon=104),
        projection_scale=3.5, showcountries=True,
        countrycolor='#ddd', showcoastlines=True,
        coastlinecolor='#999', showland=True, landcolor='#fafafa',
        showocean=True, oceancolor='#eef5fb', fitbounds=False,
    )
    fig.update_layout(
        title='省域销售分布地图',
        height=600, margin=dict(l=10, r=10, t=40, b=10),
        geo=dict(lataxis=dict(range=[15,55], showgrid=False),
                 lonaxis=dict(range=[70,140], showgrid=False)),
    )
    return fig

def plot_province_bar(prov_stats):
    """Province bar chart sorted by sales"""
    sorted_df = prov_stats.nlargest(15, '销售额').sort_values('销售额')
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=sorted_df['所在省份'], x=sorted_df['销售额'],
        orientation='h',
        marker=dict(color=sorted_df['销售额'], colorscale='Reds', showscale=True,
                     colorbar=dict(title='销售额(¥)')),
        text=sorted_df['销售额'].apply(lambda x: f'¥{x/1e8:.1f}亿'),
        textposition='outside',
    ))
    fig.update_layout(
        title='省份销售额排名（Top 15）',
        height=500,
        xaxis_title='销售额(¥)',
        yaxis=dict(categoryorder='total ascending'),
    )
    return fig

def plot_region_treemap(region_df):
    """Treemap showing region → province hierarchy"""
    # Clean data: remove rows with zero/NaN values and empty paths
    clean_df = region_df.copy()
    clean_df = clean_df[clean_df['销售额(亿)'] > 0]
    clean_df = clean_df[clean_df['区域'].notna() & (clean_df['区域'] != '')]
    clean_df = clean_df[clean_df['省份'].notna() & (clean_df['省份'] != '')]

    if len(clean_df) == 0:
        fig = go.Figure()
        fig.update_layout(title='区域-省份销售额构成（无数据）', height=400)
        return fig

    try:
        fig = px.treemap(
            clean_df,
            path=['区域', '省份'],
            values='销售额(亿)',
            color='销售额(亿)',
            color_continuous_scale='RdYlGn',
            title='区域-省份销售额构成（亿元）',
            height=550,
        )
        fig.update_traces(texttemplate='%{label}<br>¥%{value:.1f}亿')
        return fig
    except Exception:
        # Fallback: simple bar chart
        fig = px.bar(
            clean_df.groupby('区域')['销售额(亿)'].sum().reset_index(),
            x='区域', y='销售额(亿)',
            title='区域销售额构成（亿元）',
            color='销售额(亿)',
            color_continuous_scale='RdYlGn',
        )
        return fig

def plot_monthly_trend(monthly_df):
    """Monthly trend with bar + line for MoM growth"""
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Bar(x=monthly_df['月份'], y=monthly_df['销售额'], name='销售额',
               marker_color='steelblue', hovertemplate='%{x}<br>¥%{y:,.0f}'),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(x=monthly_df['月份'], y=monthly_df['环比增长(%)'], name='环比增长%',
                   mode='lines+markers', line=dict(color='red', width=2, dash='dash'),
                   marker=dict(size=8)),
        secondary_y=True,
    )
    fig.update_layout(
        title='月度销售额趋势 & 环比增长',
        height=450,
        hovermode='x unified',
        legend=dict(orientation='h', y=1.1),
    )
    fig.update_yaxes(title_text='销售额(¥)', secondary_y=False)
    fig.update_yaxes(title_text='环比增长(%)', secondary_y=True)
    return fig

def plot_rfm_scatter(rfm_df):
    """RFM 3D scatter plot"""
    color_map = {'核心客户': '#2ecc71', '重要客户': '#3498db', '一般客户': '#f39c12', '流失风险': '#e74c3c'}

    fig = go.Figure()
    for tier in ['核心客户', '重要客户', '一般客户', '流失风险']:
        td = rfm_df[rfm_df['客户层级'] == tier]
        if len(td) == 0: continue
        fig.add_trace(go.Scatter3d(
            x=td['Recency(天)'], y=td['Frequency'], z=td['Monetary(¥)']/1e4,
            mode='markers', name=f'{tier}({len(td)})',
            marker=dict(size=6, color=color_map.get(tier, '#999'), opacity=0.7),
            hovertemplate='%{text}<extra></extra>',
            text=[f'{c}<br>R:{r}天 F:{f} M:¥{m/1e4:.1f}万' for c, r, f, m in
                  zip(td['客户编码'], td['Recency(天)'], td['Frequency'], td['Monetary(¥)'])],
        ))
    fig.update_layout(
        title='客户RFM三维分布',
        height=550,
        scene=dict(
            xaxis_title='Recency(距最后购买天数)',
            yaxis_title='Frequency(购买频次)',
            zaxis_title='Monetary(万元)',
        ),
        legend=dict(orientation='h', y=1.1),
    )
    return fig

def plot_rfm_pie(rfm_df):
    """Pie chart of customer tiers"""
    tier_counts = rfm_df['客户层级'].value_counts()
    colors = {'核心客户': '#2ecc71', '重要客户': '#3498db', '一般客户': '#f39c12', '流失风险': '#e74c3c'}
    fig = go.Figure(go.Pie(
        labels=tier_counts.index,
        values=tier_counts.values,
        marker=dict(colors=[colors.get(t, '#999') for t in tier_counts.index]),
        hole=0.4,
        texttemplate='%{label}<br>%{value}位(%{percent})',
    ))
    fig.update_layout(title='客户层级构成', height=400)
    return fig

def plot_refund_by_region(df):
    """Refund analysis by region"""
    region_stats = df.groupby('收货地址').agg(
        订单数=('订单编号', 'count'),
        总金额=('总金额', 'sum'),
        实付金额=('实付金额', 'sum'),
        退款金额=('退款金额', 'sum'),
    ).reset_index()
    region_stats['退款率'] = region_stats['退款金额'] / region_stats['实付金额'].replace(0, np.nan) * 100
    region_stats['未付款率'] = (1 - region_stats['实付金额'] / region_stats['总金额'].replace(0, np.nan)) * 100
    region_stats = region_stats[region_stats['订单数'] >= 50].sort_values('退款率', ascending=False)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=region_stats['收货地址'].head(15), x=region_stats['退款率'].head(15),
        orientation='h', name='退款率(%)', marker_color='#e74c3c',
        text=region_stats['退款率'].head(15).apply(lambda x: f'{x:.1f}%'),
        textposition='outside',
    ))
    fig.update_layout(
        title='各地区退款率排名（剔除<50单的地区）',
        height=500,
        xaxis_title='退款率(%)',
        yaxis=dict(categoryorder='total ascending'),
    )
    return fig

def plot_refund_daily(df):
    """Daily refund trend"""
    df_clean = df[df['创建时间'].notna()].copy()
    df_clean['日期'] = df_clean['创建时间'].apply(lambda x: x.date() if isinstance(x, datetime) else None)
    daily = df_clean.groupby('日期').agg(
        订单数=('订单编号', 'count'),
        实付总额=('实付金额', 'sum'),
        退款总额=('退款金额', 'sum'),
    ).reset_index()
    daily['退款率'] = daily['退款总额'] / daily['实付总额'].replace(0, np.nan) * 100
    daily = daily.dropna(subset=['日期'])

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Scatter(x=daily['日期'], y=daily['实付总额'], name='实付总额',
                   mode='lines', line=dict(color='steelblue', width=2), fill='tozeroy'),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(x=daily['日期'], y=daily['退款率'], name='退款率%',
                   mode='lines', line=dict(color='red', width=2)),
        secondary_y=True,
    )
    fig.update_layout(title='每日销售额 & 退款率趋势', height=450, hovermode='x unified')
    fig.update_yaxes(title_text='销售额(¥)', secondary_y=False)
    fig.update_yaxes(title_text='退款率(%)', secondary_y=True)
    return fig

def plot_price_distribution(df):
    """Price band distribution pie"""
    bands = []
    for p in df['price']:
        if p <= 50: bands.append('0-50元')
        elif p <= 100: bands.append('50-100元')
        elif p <= 200: bands.append('100-200元')
        elif p <= 500: bands.append('200-500元')
        elif p <= 1000: bands.append('500-1000元')
        else: bands.append('1000元以上')
    band_counts = Counter(bands)
    order = ['0-50元', '50-100元', '100-200元', '200-500元', '500-1000元', '1000元以上']
    values = [band_counts.get(k, 0) for k in order]

    fig = go.Figure(go.Pie(
        labels=order, values=values, hole=0.4,
        texttemplate='%{label}<br>%{value} SKU(%{percent})',
        marker=dict(colors=px.colors.sequential.Blues_r),
    ))
    fig.update_layout(title='SKU价格带分布', height=400)
    return fig

# ============================================================
# MAIN APP
# ============================================================
def main():
    # Sidebar
    with st.sidebar:
        st.image("https://img.icons8.com/color/96/000000/geography.png", width=60)
        st.title("📊 分析平台")
        st.markdown("---")
        st.markdown("### 📁 数据文件")
        st.markdown("""
        - `tmall_order_report.xlsx`
        - `双十一淘宝美妆数据.xlsx`
        - `日化.xlsx`
        """)
        st.markdown("---")
        st.markdown("### ⚙️ 全局设置")

        # Global filters
        selected_brands_global = st.multiselect(
            "筛选品牌（影响Tab2）",
            options=[],
            help="加载数据后自动填充品牌列表"
        )

        st.markdown("---")
        st.markdown("### 📥 数据导出")
        if st.button("📄 导出分析报告(HTML)", use_container_width=True):
            st.info("请切换到对应Tab导出")

        st.markdown("---")
        st.caption(f"© 2025 项目一组 | L4 系统化分析平台")

    # ============================================================
    # LOAD DATA
    # ============================================================
    import os
    base_dir = os.path.dirname(os.path.abspath(__file__))

    tmall_path = os.path.join(base_dir, 'tmall_order_report.xlsx')
    beauty_path = os.path.join(base_dir, '双十一淘宝美妆数据.xlsx')
    rihua_path = os.path.join(base_dir, '日化.xlsx')

    # Check files exist
    missing = []
    for name, path in [('天猫订单', tmall_path), ('双十一美妆', beauty_path), ('日化数据', rihua_path)]:
        if not os.path.exists(path):
            missing.append(name)

    if missing:
        st.error(f"❌ 缺少数据文件: {', '.join(missing)}")
        st.info("请确保三个 Excel 文件放在与本程序相同的目录下")
        st.stop()

    with st.spinner("正在加载数据..."):
        df_tmall = load_tmall_data(tmall_path)
        df_beauty = load_beauty_data(beauty_path)
        df_sales, df_product = load_rihua_data(rihua_path)

    # Update brand selector
    all_brands = sorted(df_beauty['品牌'].unique().tolist())

    # ============================================================
    # TABS
    # ============================================================
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📋 数据概览",
        "🏷️ 品牌竞争力",
        "🏭 渠道健康度",
        "🛒 零售效率",
        "🔗 联动洞察"
    ])

    # ============================================================
    # TAB 1: 数据概览
    # ============================================================
    with tab1:
        st.title("📋 数据概览")
        st.markdown("> 三份数据覆盖**品牌→渠道→零售**完整产业链，共计 **86,960** 条记录")

        # KPI Row
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("📦 数据文件", "3", "Excel")
        with col2:
            st.metric("🏷️ 品牌数", f"{len(all_brands)}", "双十一美妆")
        with col3:
            st.metric("🏭 经销商客户", f"{df_sales['客户编码'].nunique()}", "日化数据")
        with col4:
            st.metric("🛒 天猫订单", f"{len(df_tmall):,}", "2025年2月")
        with col5:
            total_gmv = df_beauty['gmv'].sum() / 1e8
            total_b2b = df_sales['金额'].sum() / 1e8
            st.metric("💰 总交易规模", f"¥{total_gmv + total_b2b:.0f}亿", "估算")

        st.markdown("---")

        # Dataset summaries in 3 columns
        c1, c2, c3 = st.columns(3)

        with c1:
            st.subheader("🐱 天猫订单报告")
            paid = df_tmall['实付金额'].sum()
            refund = df_tmall['退款金额'].sum()
            refund_rate = refund / paid * 100 if paid > 0 else 0
            zero_pay = (df_tmall['实付金额'] == 0).sum()
            st.markdown(f"""
            | 指标 | 数值 |
            |------|------|
            | 订单数 | **{len(df_tmall):,}** |
            | 时间范围 | 2025-02-01 ~ 2025-02-28 |
            | 订单总额 | ¥{df_tmall['总金额'].sum():,.0f} |
            | 实付金额 | ¥{paid:,.0f} |
            | 退款金额 | ¥{refund:,.0f} |
            | **退款率(金额)** | **{refund_rate:.1f}%** 🔴 |
            | 零支付订单 | {zero_pay} ({zero_pay/len(df_tmall)*100:.1f}%) |
            """)

        with c2:
            st.subheader("💄 双十一美妆数据")
            st.markdown(f"""
            | 指标 | 数值 |
            |------|------|
            | SKU总数 | **{len(df_beauty):,}** |
            | 品牌数 | {len(all_brands)} |
            | 总销量 | {df_beauty['sale_count'].sum()/1e8:.2f}亿件 |
            | GMV估算 | ¥{df_beauty['gmv'].sum()/1e8:,.1f}亿 |
            | 价格区间 | ¥{df_beauty['price'].min():.0f} ~ ¥{df_beauty['price'].max():.0f} |
            | 均价 | ¥{df_beauty['price'].mean():.0f} |
            | 评论率 | {df_beauty['comment_count'].sum()/max(df_beauty['sale_count'].sum(),1)*100:.1f}% |
            """)

        with c3:
            st.subheader("🧴 日化渠道数据")
            monthly = compute_monthly_trend(df_sales)
            months = len(monthly)
            st.markdown(f"""
            | 指标 | 数值 |
            |------|------|
            | 销售记录 | **{len(df_sales):,}** |
            | 订单数 | {df_sales['订单编码'].nunique():,} |
            | 经销商客户 | **{df_sales['客户编码'].nunique():,}** |
            | 商品SKU | {df_product['商品编号'].nunique()} |
            | 覆盖省份 | {df_sales['所在省份'].nunique()} |
            | 覆盖城市 | {df_sales['所在地市'].nunique()} |
            | 总金额 | **¥{df_sales['金额'].sum()/1e8:,.1f}亿** |
            | 月数 | {months}个月 |
            | 品类 | 护肤品({len(df_product[df_product['商品大类']=='护肤品'])}) + 彩妆({len(df_product[df_product['商品大类']=='彩妆'])}) |
            """)

        st.markdown("---")
        st.subheader("🔍 数据样本预览")
        sample_tab1, sample_tab2, sample_tab3 = st.tabs(["天猫订单", "双十一美妆", "日化销售"])

        with sample_tab1:
            st.dataframe(df_tmall.head(10), use_container_width=True, hide_index=True)

        with sample_tab2:
            st.dataframe(df_beauty[['品牌', 'title', 'price', 'sale_count', 'comment_count']].head(10),
                         use_container_width=True, hide_index=True)

        with sample_tab3:
            st.dataframe(df_sales.head(10), use_container_width=True, hide_index=True)

    # ============================================================
    # TAB 2: 品牌竞争力
    # ============================================================
    with tab2:
        st.title("🏷️ 品牌竞争力分析")
        st.markdown("> 基于双十一美妆数据，从6个维度评估22个品牌的竞争力")

        # Compute scores
        scores_df = compute_brand_scores(df_beauty)

        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            selected_brands = st.multiselect(
                "选择品牌（多选）",
                options=all_brands,
                default=all_brands[:5],
                key="tab2_brands"
            )
        with col_f2:
            price_range = st.slider(
                "价格区间(¥)",
                min_value=0, max_value=int(df_beauty['price'].max()),
                value=(0, int(df_beauty['price'].max())),
                step=10,
                key="tab2_price"
            )
        with col_f3:
            sort_by = st.selectbox(
                "排序方式",
                options=['GMV(亿)', 'SKU数', '均售价(¥)', '爆款率(%)', '评论率(%)'],
                key="tab2_sort"
            )

        # Filter data
        mask = df_beauty['品牌'].isin(selected_brands)
        mask &= (df_beauty['price'] >= price_range[0]) & (df_beauty['price'] <= price_range[1])
        filtered_beauty = df_beauty[mask]

        # --- Charts Row 1 ---
        st.markdown("### 📊 品牌GMV排名 & 价格-销量分布")
        c_left, c_right = st.columns([1, 1])

        with c_left:
            fig_gmv = plot_brand_gmv_bar(
                scores_df[scores_df['品牌'].isin(selected_brands)],
                n=min(15, len(selected_brands)),
                sort_by=sort_by
            )
            st.plotly_chart(fig_gmv, use_container_width=True)

        with c_right:
            fig_scatter = plot_price_vs_sales_scatter(filtered_beauty, selected_brands)
            st.plotly_chart(fig_scatter, use_container_width=True)

        # --- Charts Row 2 ---
        st.markdown("### 🎯 品牌雷达图对比 & 评分卡明细")
        c_left2, c_right2 = st.columns([1, 1])

        with c_left2:
            if len(selected_brands) >= 1:
                radar_brands = selected_brands[:6]
                fig_radar = plot_brand_radar(scores_df, radar_brands)
                st.plotly_chart(fig_radar, use_container_width=True)
            else:
                st.warning("请至少选择一个品牌")

        with c_right2:
            st.markdown("#### 品牌竞争力评分卡")
            display_scores = scores_df[scores_df['品牌'].isin(selected_brands)].sort_values(
                sort_by, ascending=False
            )
            st.dataframe(
                display_scores.style
                .background_gradient(subset=['GMV(亿)', '爆款率(%)', '品牌溢价指数'], cmap='Blues')
                .format({
                    '均售价(¥)': '¥{:.0f}',
                    'GMV(亿)': '{:.2f}',
                    '评论率(%)': '{:.1f}%',
                    '爆款率(%)': '{:.1f}%',
                    '品牌溢价指数': '{:.0f}',
                    '销量效率指数': '{:.0f}',
                }),
                use_container_width=True, hide_index=True,
            )

        # Price distribution
        st.markdown("### 📐 价格带分布")
        fig_price_band = plot_price_distribution(filtered_beauty)
        st.plotly_chart(fig_price_band, use_container_width=True, height=400)

        # Export
        st.markdown("---")
        if st.button("📥 导出品牌评分卡 (CSV)", key="export_brand"):
            csv = scores_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button("下载CSV", csv, "品牌竞争力评分卡.csv", "text/csv")

    # ============================================================
    # TAB 3: 渠道健康度
    # ============================================================
    with tab3:
        st.title("🏭 渠道健康度诊断")
        st.markdown("> 日化B2B经销商网络分析：地图分布、客户RFM分层、月度趋势")

        # Compute stats
        rfm_df = compute_rfm(df_sales)
        monthly_df = compute_monthly_trend(df_sales)
        prov_stats = compute_province_stats(df_sales)
        region_tree = compute_region_tree(df_sales)

        # --- KPI Row ---
        k1, k2, k3, k4 = st.columns(4)
        risk_count = len(rfm_df[rfm_df['客户层级'] == '流失风险']) if '客户层级' in rfm_df.columns else 0
        core_count = len(rfm_df[rfm_df['客户层级'] == '核心客户']) if '客户层级' in rfm_df.columns else 0

        with k1:
            latest_m = monthly_df['销售额'].iloc[-1] if len(monthly_df) > 0 else 0
            prev_m = monthly_df['销售额'].iloc[-2] if len(monthly_df) > 1 else latest_m
            delta = f"{(latest_m/prev_m - 1)*100:+.1f}%" if prev_m > 0 else "N/A"
            st.metric("📈 最新月销售额", f"¥{latest_m/1e8:.1f}亿", delta)
        with k2:
            st.metric("👥 总客户数", f"{len(rfm_df)}", "")
        with k3:
            st.metric("🟢 核心客户", f"{core_count}", f"{core_count/len(rfm_df)*100:.0f}%")
        with k4:
            st.metric("🔴 流失风险", f"{risk_count}", f"{risk_count/len(rfm_df)*100:.0f}%" if len(rfm_df) > 0 else "")

        # --- Charts Row 1: Map + Bar ---
        st.markdown("### 🗺️ 省域销售分布")
        c_map, c_bar = st.columns([1.2, 1])

        with c_map:
            if len(prov_stats) > 0:
                fig_map = plot_province_map(prov_stats)
                st.plotly_chart(fig_map, use_container_width=True)
            else:
                st.warning("暂无足够的地图数据")

        with c_bar:
            fig_bar = plot_province_bar(prov_stats)
            st.plotly_chart(fig_bar, use_container_width=True)

        # --- Charts Row 2: Treemap + Monthly Trend ---
        st.markdown("### 🌳 区域构成 & 月度趋势")
        c_tree, c_month = st.columns([1, 1])

        with c_tree:
            fig_tree = plot_region_treemap(region_tree)
            st.plotly_chart(fig_tree, use_container_width=True)

        with c_month:
            fig_month = plot_monthly_trend(monthly_df)
            st.plotly_chart(fig_month, use_container_width=True)

        # --- RFM Analysis ---
        st.markdown("### 🎯 客户RFM分层")
        c_rfm1, c_rfm2 = st.columns([1.2, 1])

        with c_rfm1:
            fig_rfm3d = plot_rfm_scatter(rfm_df)
            st.plotly_chart(fig_rfm3d, use_container_width=True)

        with c_rfm2:
            fig_pie = plot_rfm_pie(rfm_df)
            st.plotly_chart(fig_pie, use_container_width=True)

            # High-risk list
            if '客户层级' in rfm_df.columns:
                risk_list = rfm_df[rfm_df['客户层级'] == '流失风险'].nlargest(10, 'Monetary(¥)')
                if len(risk_list) > 0:
                    st.markdown("#### ⚠️ 高价值但流失风险客户")
                    st.dataframe(
                        risk_list[['客户编码', 'Recency(天)', 'Frequency', 'Monetary(¥)', 'RFM总分']]
                        .style.format({'Monetary(¥)': '¥{:,.0f}'}),
                        use_container_width=True, hide_index=True,
                    )

        # --- Province Detail ---
        st.markdown("### 📋 省份数据明细")
        prov_display = prov_stats.sort_values('销售额', ascending=False)
        st.dataframe(
            prov_display.style
            .format({'销售额': '¥{:,.0f}', '客单价': '¥{:,.0f}'})
            .background_gradient(subset=['销售额', '客户数'], cmap='Reds'),
            use_container_width=True, hide_index=True,
            height=400,
        )

        # Export
        st.markdown("---")
        if st.button("📥 导出RFM分析 (CSV)", key="export_rfm"):
            csv = rfm_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button("下载CSV", csv, "客户RFM分析.csv", "text/csv")

    # ============================================================
    # TAB 4: 零售效率
    # ============================================================
    with tab4:
        st.title("🛒 零售效率分析")
        st.markdown("> 天猫订单报告：退款归因、支付行为、地域效率")

        # KPI Row
        paid = df_tmall['实付金额'].sum()
        refund = df_tmall['退款金额'].sum()
        refund_rate = refund / paid * 100 if paid > 0 else 0
        zero_pay = (df_tmall['实付金额'] == 0).sum()

        k1, k2, k3, k4 = st.columns(4)
        with k1:
            st.metric("💰 实付总额", f"¥{paid:,.0f}", "")
        with k2:
            st.metric("💸 退款总额", f"¥{refund:,.0f}", f"-{refund_rate:.1f}%")
        with k3:
            st.metric("❌ 退款率(金额)", f"{refund_rate:.1f}%", "⚠️ 偏高")
        with k4:
            st.metric("📭 未付款订单", f"{zero_pay:,}", f"{zero_pay/len(df_tmall)*100:.1f}%")

        # --- Charts Row 1: Daily trend + Region refund ---
        st.markdown("### 📈 每日趋势 & 区域退款分析")
        c1, c2 = st.columns([1, 1])

        with c1:
            fig_daily = plot_refund_daily(df_tmall)
            st.plotly_chart(fig_daily, use_container_width=True)

        with c2:
            fig_region = plot_refund_by_region(df_tmall)
            st.plotly_chart(fig_region, use_container_width=True)

        # --- Payment behavior ---
        st.markdown("### 💳 支付行为分析")
        df_tmall_clean = df_tmall[df_tmall['创建时间'].notna()].copy()

        # Payment delay analysis
        same_day = 0; delayed = 0; no_pay = 0
        pay_delays = []
        for _, row in df_tmall_clean.iterrows():
            c = row['创建时间']; p = row['付款时间']
            if isinstance(c, datetime) and isinstance(p, datetime):
                delay = (p - c).total_seconds() / 3600
                pay_delays.append(delay)
                if delay < 24: same_day += 1
                else: delayed += 1
            elif isinstance(c, datetime):
                no_pay += 1

        c_p1, c_p2 = st.columns(2)
        with c_p1:
            pay_data = pd.DataFrame({
                '支付行为': ['当日付款', '延迟付款', '未付款'],
                '订单数': [same_day, delayed, no_pay],
            })
            fig_pay = px.pie(pay_data, values='订单数', names='支付行为', hole=0.4,
                             title='支付行为分布',
                             color='支付行为',
                             color_discrete_map={'当日付款':'#2ecc71', '延迟付款':'#f39c12', '未付款':'#e74c3c'})
            fig_pay.update_traces(texttemplate='%{label}<br>%{value}单(%{percent})')
            st.plotly_chart(fig_pay, use_container_width=True)

        with c_p2:
            # Refund by amount band
            df_tmall_clean['金额段'] = pd.cut(
                df_tmall_clean['总金额'],
                bins=[0, 50, 100, 200, 500, 1000, 999999],
                labels=['0-50', '50-100', '100-200', '200-500', '500-1000', '1000+']
            )
            band_refund = df_tmall_clean.groupby('金额段', observed=False).agg(
                订单数=('订单编号', 'count'),
                退款率=('退款金额', lambda x: (x.sum() / df_tmall_clean.loc[x.index, '实付金额'].sum() * 100) if df_tmall_clean.loc[x.index, '实付金额'].sum() > 0 else 0),
            ).reset_index()

            fig_band = px.bar(
                band_refund, x='金额段', y='退款率',
                title='各金额段退款率', text=band_refund['退款率'].apply(lambda x: f'{x:.1f}%'),
                color='退款率', color_continuous_scale='Reds',
            )
            fig_band.update_traces(textposition='outside')
            st.plotly_chart(fig_band, use_container_width=True)

        # --- Refund prediction model ---
        st.markdown("### 🤖 退款风险预测模型（逻辑回归）")

        st.markdown("""
        **变量定义：**
        - **因变量 Y**：是否退款（`退款金额 > 0` → 1，否则 → 0），二分类问题
        - **自变量 X**：订单金额、下单时段、是否周末、付款延迟时长、所在地区退款率
        """)

        # Pre-compute address risk for feature engineering
        addr_risk = df_tmall_clean.groupby('收货地址').apply(
            lambda g: (g['退款金额'] > 0).sum() / max((g['实付金额'] > 0).sum(), 1)
        ).to_dict()

        if st.button("🔄 训练/刷新预测模型", key="train_model"):
            with st.spinner("正在训练模型..."):
                from sklearn.linear_model import LogisticRegression
                from sklearn.preprocessing import StandardScaler
                from sklearn.model_selection import train_test_split
                from sklearn.metrics import classification_report, confusion_matrix

                # Only train on PAID orders (unpaid can't have refunds)
                model_df = df_tmall_clean[df_tmall_clean['实付金额'] > 0].copy()

                # ---- Feature Engineering ----
                model_df['退款标记'] = (model_df['退款金额'] > 0).astype(int)
                model_df['下单小时'] = model_df['创建时间'].apply(lambda x: x.hour if isinstance(x, datetime) else 12)
                model_df['下单星期'] = model_df['创建时间'].apply(lambda x: x.weekday() if isinstance(x, datetime) else 3)
                model_df['是否周末'] = model_df['下单星期'].apply(lambda x: 1 if x >= 5 else 0)
                model_df['金额'] = model_df['总金额']
                # Payment delay (hours): longer delay = more hesitation = higher risk
                model_df['付款延迟'] = model_df.apply(
                    lambda r: max((r['付款时间'] - r['创建时间']).total_seconds() / 3600, 0)
                    if isinstance(r['创建时间'], datetime) and isinstance(r['付款时间'], datetime) else 0,
                    axis=1
                )
                # Address historical refund rate
                model_df['地区风险'] = model_df['收货地址'].map(addr_risk).fillna(0.2)

                features = ['金额', '下单小时', '是否周末', '付款延迟', '地区风险']
                feature_labels = ['订单金额', '下单时段', '是否周末', '付款延迟(h)', '地区退款率']

                X = model_df[features].fillna(0)
                y = model_df['退款标记']

                scaler = StandardScaler()
                X_scaled = scaler.fit_transform(X)

                X_train, X_test, y_train, y_test = train_test_split(
                    X_scaled, y, test_size=0.3, random_state=42, stratify=y
                )

                lr = LogisticRegression(class_weight='balanced', max_iter=2000, C=0.5)
                lr.fit(X_train, y_train)

                # ---- Evaluation ----
                y_pred = lr.predict(X_test)
                train_acc = lr.score(X_train, y_train)
                test_acc = lr.score(X_test, y_test)

                from sklearn.metrics import precision_score, recall_score, f1_score
                precision = precision_score(y_test, y_pred, zero_division=0)
                recall = recall_score(y_test, y_pred, zero_division=0)
                f1 = f1_score(y_test, y_pred, zero_division=0)

                # Store in session
                st.session_state['model'] = lr
                st.session_state['scaler'] = scaler
                st.session_state['features'] = features

                # Metrics display
                c_m1, c_m2, c_m3, c_m4, c_m5 = st.columns(5)
                c_m1.metric("训练集准确率", f"{train_acc:.1%}")
                c_m2.metric("测试集准确率", f"{test_acc:.1%}")
                c_m3.metric("精确率(Precision)", f"{precision:.1%}",
                           "预测退款中真正退款的比例")
                c_m4.metric("召回率(Recall)", f"{recall:.1%}",
                           "实际退款中被识别出的比例")
                c_m5.metric("F1分数", f"{f1:.1%}", "精确率与召回率的调和")

                # Confusion matrix
                cm = confusion_matrix(y_test, y_pred)
                st.markdown("#### 混淆矩阵")
                cm_df = pd.DataFrame(
                    cm,
                    columns=['预测=未退款', '预测=退款'],
                    index=['实际=未退款', '实际=退款']
                )
                st.dataframe(cm_df, use_container_width=True)

                # Feature importance
                coef_df = pd.DataFrame({
                    '特征': feature_labels,
                    '系数': lr.coef_[0],
                    '影响方向': ['🔺 增加退款风险' if c > 0 else '🔻 降低退款风险' for c in lr.coef_[0]]
                })
                fig_coef = px.bar(
                    coef_df, x='系数', y='特征', color='影响方向',
                    title='特征对退款风险的影响（系数>0 → 该特征值越大，退款概率越高）',
                    text=coef_df['系数'].apply(lambda x: f'{x:.3f}'),
                    color_discrete_map={'🔺 增加退款风险': '#e74c3c', '🔻 降低退款风险': '#2ecc71'},
                    height=350,
                )
                st.plotly_chart(fig_coef, use_container_width=True)

        # Interactive prediction
        st.markdown("#### 🎛️ 实时退款风险预测")
        col_pred1, col_pred2, col_pred3, col_pred4 = st.columns(4)
        with col_pred1:
            pred_amount = st.number_input("订单金额(¥)", min_value=0.0, max_value=10000.0, value=200.0, step=10.0)
        with col_pred2:
            pred_hour = st.slider("下单小时（0-23）", 0, 23, 20,
                                  help="凌晨下单 → 冲动消费 → 退款风险高")
        with col_pred3:
            pred_weekend = st.selectbox("是否周末", options=['工作日', '周末'],
                                        help="周末下单 → 闲逛型消费")
        with col_pred4:
            pred_delay = st.slider("付款延迟(小时)", 0.0, 72.0, 0.0, 1.0,
                                   help="下单后多久付款，延迟越久越可能退款")

        if st.button("🔮 预测退款风险", key="predict_btn"):
            if 'model' in st.session_state:
                weekend_val = 1 if pred_weekend == '周末' else 0
                # Use default region risk
                X_pred = np.array([[pred_amount, pred_hour, weekend_val, pred_delay, 0.2]])
                X_pred_scaled = st.session_state['scaler'].transform(X_pred)
                proba = st.session_state['model'].predict_proba(X_pred_scaled)[0][1]

                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    risk_level = '🔴 高风险' if proba > 0.4 else ('🟡 中风险' if proba > 0.2 else '🟢 低风险')
                    st.metric("退款概率", f"{proba:.1%}", delta=risk_level)
                with col_r2:
                    if proba > 0.4:
                        st.error("⚠️ 建议：该订单退款风险较高，建议发送定向优惠券或客服主动跟进挽留")
                    elif proba > 0.2:
                        st.warning("⚡ 该订单存在一定退款风险，建议关注物流进度，提前发送使用指南")
                    else:
                        st.success("✅ 该订单退款风险较低，正常履约即可")
            else:
                st.warning("请先点击上方「训练/刷新预测模型」按钮")

    # ============================================================
    # TAB 5: 联动洞察
    # ============================================================
    with tab5:
        st.title("🔗 跨数据联动洞察")
        st.markdown("> 将三份数据拼接成**品牌→渠道→零售**产业链全景图")

        # Framework diagram
        st.markdown("""
        ### 🧩 分析框架
        ```
        ┌──────────────────────────────────────────────────────────┐
        │                     产业链分析引擎                          │
        │                                                          │
        │  双十一美妆数据         日化渠道数据         天猫零售数据     │
        │  (品牌市场竞争)    →    (经销商健康度)    →   (终端效率)     │
        │                                                          │
        │  • 品牌GMV排名           • RFM分层           • 退款归因    │
        │  • 价格带策略            • 区域渗透率         • 支付行为    │
        │  • 竞品对标              • 月度趋势           • 风险预测    │
        │                                                          │
        │              └──────── 跨数据联动 ─────────┘               │
        │                                                          │
        │  🎯 面膜品类链路：品牌卖多少 → 经销商铺多少 → 终端退多少        │
        │  🎯 区域穿透力：品类热力图 × 经销商密度 × 零售退货率            │
        └──────────────────────────────────────────────────────────┘
        ```
        """)

        st.markdown("---")

        # ============================================================
        # Insight 1: Category Chain Analysis
        # ============================================================
        st.subheader("🔍 洞察一：品类全链路透视（以面膜为例）")

        # Beauty: find mask-related brands
        mask_keywords = ['面膜', '面贴膜', '眼膜', '面贴']
        df_beauty['is_mask'] = df_beauty['title'].apply(
            lambda x: any(kw in str(x) for kw in mask_keywords)
        )
        mask_beauty = df_beauty[df_beauty['is_mask']]

        # Rihua: mask products
        mask_products = df_product[df_product['商品小类'] == '面膜']
        mask_codes = mask_products['商品编号'].tolist()
        mask_sales = df_sales[df_sales['商品编号'].isin(mask_codes)]

        col_i1, col_i2, col_i3 = st.columns(3)
        with col_i1:
            st.metric("🏷️ 双十一面膜SKU", f"{len(mask_beauty):,}", f"占总SKU {len(mask_beauty)/len(df_beauty)*100:.1f}%")
            st.metric("💰 面膜GMV估算", f"¥{mask_beauty['gmv'].sum()/1e8:.1f}亿")

        with col_i2:
            st.metric("📦 经销商面膜SKU", f"{len(mask_codes)}", f"占总商品 {len(mask_codes)/len(df_product)*100:.0f}%")
            st.metric("💵 渠道面膜销售额", f"¥{mask_sales['金额'].sum()/1e8:.1f}亿")

        with col_i3:
            # Compare ratio
            beauty_gmv = mask_beauty['gmv'].sum()
            channel_gmv = mask_sales['金额'].sum()
            ratio = channel_gmv / beauty_gmv * 100 if beauty_gmv > 0 else 0
            st.metric("📊 渠道/品牌比", f"{ratio:.0f}%",
                     "渠道>品牌：经销商压货" if ratio > 100 else "品牌>渠道：直营为主")

        st.markdown("---")

        # ============================================================
        # Insight 2: Regional Cross-Analysis
        # ============================================================
        st.subheader("🗺️ 洞察二：区域穿透力矩阵")

        # Province overlap between tmall addresses and rihua sales
        tmall_addrs = set(df_tmall['收货地址'].unique())
        rihua_provs = set(df_sales['所在省份'].unique())

        common = tmall_addrs & rihua_provs
        only_tmall = tmall_addrs - rihua_provs
        only_rihua = rihua_provs - tmall_addrs

        cc1, cc2, cc3 = st.columns(3)
        cc1.metric("共同覆盖", f"{len(common)}")
        cc2.metric("仅零售覆盖", f"{len(only_tmall)}")
        cc3.metric("仅渠道覆盖", f"{len(only_rihua)}", "→ 渠道空白市场")

        # Combined region analysis
        st.markdown("#### 省份级别：渠道销售额 vs 零售订单数")
        tmall_prov = df_tmall.groupby('收货地址').agg(
            零售订单=('订单编号', 'count'),
            零售退款率=('退款金额', lambda x: x.sum() / max(df_tmall.loc[x.index, '实付金额'].sum(), 1) * 100),
        ).reset_index()
        tmall_prov.columns = ['省份', '零售订单', '零售退款率']

        rihua_prov_agg = df_sales.groupby('所在省份').agg(
            渠道销售额=('金额', 'sum'),
            渠道客户=('客户编码', 'nunique'),
        ).reset_index()
        rihua_prov_agg.columns = ['省份', '渠道销售额', '渠道客户']

        merged_prov = pd.merge(rihua_prov_agg, tmall_prov, on='省份', how='outer').fillna(0)
        if len(merged_prov) > 0:
            fig_cross = px.scatter(
                merged_prov,
                x='渠道销售额', y='零售订单',
                size='渠道客户',
                color='零售退款率',
                hover_name='省份',
                title='省份交叉矩阵：渠道实力 × 零售规模 × 退款率',
                labels={'渠道销售额': '渠道销售额(¥)', '零售订单': '零售订单数', '零售退款率': '零售退款率(%)'},
                color_continuous_scale='RdYlGn_r',
                height=500,
            )
            fig_cross.update_layout(legend=dict(orientation='h', y=-0.15))
            st.plotly_chart(fig_cross, use_container_width=True)

        st.markdown("---")

        # ============================================================
        # Insight 3: Key Findings Summary
        # ============================================================
        st.subheader("📌 核心发现汇总")

        f1, f2, f3 = st.columns(3)
        with f1:
            st.info("""
            **发现一：国货品牌完成大众价位带封锁**
            - ¥100-200价位带占SKU的23.2%
            - 相宜本草以均价¥123做到¥61亿GMV
            - 国际品牌被困高端价位，市占率受限
            """)
        with f2:
            st.warning("""
            **发现二：天猫零售存在严重的"冲动消费"问题**
            - 退款率30.1%（行业正常水平约5-10%）
            - 32.3%的订单从未付款
            - 高价段(¥500+)退款率显著偏高
            """)
        with f3:
            st.success("""
            **发现三：渠道有明确的区域扩张机会**
            - 东区占36%份额，过度集中
            - 西区12省仅占13%份额，但人口基数大
            - 核心客户仅占客户群的少数，尾部客户活跃度不足
            """)

        f4, f5, f6 = st.columns(3)
        with f4:
            st.info("""
            **发现四：面膜是"引流-退货"双刃剑**
            - 面膜是SKU最多的品类（16个）
            - 双十一面膜GMV占比极高
            - 但也可能是高退货率品类
            """)
        with f5:
            st.warning("""
            **发现五：江苏省是"完美市场"**
            - 渠道销售额¥6.7亿（第一）
            - 零售订单数2,126单（第三）
            - 渠道-零售协同效应最优
            """)
        with f6:
            st.success("""
            **发现六：系统化平台的价值**
            - 可替换数据源（换一批Excel即可复用）
            - 交互式探索（比静态报告发现更多问题）
            - 模型可迭代（持续优化RFM/退款预测）
            """)

        st.markdown("---")

        # ============================================================
        # Overall Summary
        # ============================================================
        st.subheader("📝 分析总结")

        st.markdown("""
        ### 一句话结论

        > **国货美妆品牌在大众市场已建立压倒性优势，但渠道效率和终端零售体验仍是瓶颈——**
        > **30%的退款率意味着每三个客户就有一个不满意，这是我们面临的最大挑战。**

        ### 行动建议

        | 优先级 | 领域 | 建议 | 预期影响 |
        |--------|------|------|----------|
        | 🔴 P0 | 零售退款 | 建立退款预警模型，对高风险订单主动介入（优惠券/客服） | 降低退款率5-10个百分点 |
        | 🟡 P1 | 渠道扩张 | 西区（四川、陕西为核心）新增经销商招募，复制江苏模式 | 增长15-20% |
        | 🟡 P1 | 品类优化 | 减少同质化面膜SKU，增加差异化产品（防晒、精华） | 提升客单价和复购 |
        | 🟢 P2 | 客户运营 | 对RFM"流失风险"客户启动召回计划 | 挽回10%高危客户 |

        ---
        *本分析报告由「日化美妆产业链分析平台」自动生成 | L4 系统化数据分析项目*
        """)

        # Export full report
        st.download_button(
            "📥 导出完整分析报告 (CSV数据包)",
            data=df_sales.head(1000).to_csv(index=False).encode('utf-8-sig'),
            file_name="analysis_export_sample.csv",
            mime="text/csv",
        )


# ============================================================
# RUN
# ============================================================
if __name__ == "__main__":
    main()
