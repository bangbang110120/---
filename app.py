# -*- coding: utf-8 -*-
"""日化美妆产业链分析平台 - L4系统化答辩项目"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from collections import Counter
from datetime import datetime
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="日化美妆产业链分析平台",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# PROVINCE NAME MAPPING (for GeoJSON matching)
# ============================================================
def normalize_province(name):
    """Normalize province names to match DataV GeoJSON format"""
    if not isinstance(name, str): return name
    name = name.strip()
    mapping = {
        '北京': '北京市', '上海': '上海市', '天津': '天津市', '重庆': '重庆市',
        '广东省': '广东省', '浙江省': '浙江省', '江苏省': '江苏省',
        '山东省': '山东省', '四川省': '四川省', '福建省': '福建省',
        '安徽省': '安徽省', '湖南省': '湖南省', '辽宁省': '辽宁省',
        '河北省': '河北省', '河南省': '河南省', '江西省': '江西省',
        '云南省': '云南省', '贵州省': '贵州省', '湖北省': '湖北省',
        '陕西省': '陕西省', '山西省': '山西省', '甘肃省': '甘肃省',
        '吉林省': '吉林省', '黑龙江省': '黑龙江省', '海南省': '海南省',
        '青海省': '青海省', '台湾省': '台湾省',
        '内蒙古自治区': '内蒙古自治区',
        '广西壮族自治区': '广西壮族自治区',
        '新疆维吾尔自治区': '新疆维吾尔自治区',
        '宁夏回族自治区': '宁夏回族自治区',
        '西藏自治区': '西藏自治区',
        '香港特别行政区': '香港特别行政区',
        '澳门特别行政区': '澳门特别行政区',
    }
    return mapping.get(name, name)

# ============================================================
# UTILITIES
# ============================================================
def safe_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        s = str(v).replace('元','').replace('¥','').replace(',','').replace(' ','').strip()
        return float(s) if s else 0.0
    except:
        return 0.0

# Province coordinates for map viz (approximate centroids)
PROVINCE_COORDS = {
    '北京': (116.4, 39.9), '上海': (121.5, 31.2), '天津': (117.2, 39.1),
    '重庆': (106.5, 29.5), '广东省': (113.3, 23.1), '浙江省': (120.2, 30.3),
    '江苏省': (118.8, 32.1), '山东省': (117.0, 36.7), '四川省': (104.1, 30.6),
    '福建省': (119.3, 26.1), '安徽省': (117.3, 31.9), '湖南省': (113.0, 28.2),
    '辽宁省': (123.4, 41.8), '河北省': (114.5, 38.0), '河南省': (113.7, 34.8),
    '江西省': (115.9, 28.7), '云南省': (102.7, 25.0), '贵州省': (106.7, 26.6),
    '广西壮族自治区': (108.3, 22.8), '湖北省': (114.4, 30.6), '陕西省': (108.9, 34.3),
    '山西省': (112.6, 37.9), '甘肃省': (103.8, 36.1), '吉林省': (125.3, 43.9),
    '黑龙江省': (126.6, 45.8), '新疆维吾尔自治区': (87.6, 43.8), '海南省': (110.3, 20.0),
    '内蒙古自治区': (111.7, 40.8), '宁夏回族自治区': (106.3, 38.5),
    '青海省': (101.8, 36.6), '西藏自治区': (91.1, 29.6),
    '香港特别行政区': (114.2, 22.3), '台湾省': (121.0, 23.5),
}

# ============================================================
# DATA LOADING (cached)
# ============================================================
@st.cache_data
def load_tmall_data(filepath):
    """Load 天猫订单报告"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        create_dt = row[4] if isinstance(row[4], datetime) else None
        pay_dt = row[5] if isinstance(row[5], datetime) else None
        rows.append({
            '订单编号': str(row[0]) if row[0] else '',
            '总金额': safe_float(row[1]),
            '实付金额': safe_float(row[2]),
            '收货地址': str(row[3]).strip() if row[3] else '',
            '创建时间': create_dt,
            '付款时间': pay_dt,
            '退款金额': safe_float(row[6]),
        })
    return pd.DataFrame(rows)

@st.cache_data
def load_beauty_data(filepath):
    """Load 双十一淘宝美妆数据"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        brand = str(row[6]).strip() if row[6] else 'Unknown'
        price = safe_float(row[3])
        sale = int(row[4]) if row[4] else 0
        comment = int(row[5]) if row[5] else 0
        rows.append({
            'update_time': row[0],
            'id': str(row[1]) if row[1] else '',
            'title': str(row[2]) if row[2] else '',
            'price': price,
            'sale_count': sale,
            'comment_count': comment,
            '品牌': brand,
            'gmv': price * sale,
        })
    return pd.DataFrame(rows)

@st.cache_data
def load_rihua_data(filepath):
    """Load 日化数据 (sales detail + product info)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # Sheet 0: 销售明细
    ws_sales = wb[wb.sheetnames[0]]
    sales_rows = []
    for row in ws_sales.iter_rows(min_row=2, values_only=True):
        dt_val = row[1] if isinstance(row[1], datetime) else None
        sales_rows.append({
            '订单编码': str(row[0]) if row[0] else '',
            '订单日期': dt_val,
            '客户编码': str(row[2]).strip() if row[2] else '',
            '所在区域': str(row[3]).strip() if row[3] else '',
            '所在省份': str(row[4]).strip() if row[4] else '',
            '所在地市': str(row[5]).strip() if row[5] else '',
            '商品编号': str(row[6]).strip() if row[6] else '',
            '订购数量': safe_float(row[7]),
            '订购单价': safe_float(row[8]),
            '金额': safe_float(row[9]),
        })
    df_sales = pd.DataFrame(sales_rows)

    # Remove the 2050 outlier
    df_sales = df_sales[df_sales['订单日期'].notna()]
    df_sales = df_sales[df_sales['订单日期'].apply(lambda x: x.year < 2030 if isinstance(x, datetime) else True)]

    # Sheet 1: 商品信息表
    ws_prod = wb[wb.sheetnames[1]]
    prod_rows = []
    for row in ws_prod.iter_rows(min_row=2, values_only=True):
        prod_rows.append({
            '商品编号': str(row[0]).strip() if row[0] else '',
            '商品名称': str(row[1]).strip() if row[1] else '',
            '商品小类': str(row[2]).strip() if row[2] else '',
            '商品大类': str(row[3]).strip() if row[3] else '',
            '销售单价': safe_float(row[4]),
        })
    df_prod = pd.DataFrame(prod_rows)

    return df_sales, df_prod

# ============================================================
# ANALYSIS FUNCTIONS
# ============================================================
def compute_brand_scores(df):
    """Compute brand competitiveness scores (6 dimensions)"""
    brands = df['品牌'].unique()
    scores = []
    all_avg_price = df['price'].mean()
    all_avg_sales = df['sale_count'].mean()

    for brand in brands:
        bd = df[df['品牌'] == brand]
        skus = len(bd)
        avg_price = bd['price'].mean()
        total_gmv = bd['gmv'].sum() / 1e8  # in 100M
        total_sales = bd['sale_count'].sum()
        comment_rate = bd['comment_count'].sum() / max(total_sales, 1)
        hit_rate = len(bd[bd['sale_count'] > 10000]) / max(skus, 1)

        # Price coverage: count of price bands covered
        bands = {'0-50':0, '50-100':0, '100-200':0, '200-500':0, '500-1000':0, '1000+':0}
        for p in bd['price']:
            if p <= 50: bands['0-50'] += 1
            elif p <= 100: bands['50-100'] += 1
            elif p <= 200: bands['100-200'] += 1
            elif p <= 500: bands['200-500'] += 1
            elif p <= 1000: bands['500-1000'] += 1
            else: bands['1000+'] += 1
        coverage = sum(1 for v in bands.values() if v > 0)

        scores.append({
            '品牌': brand,
            'SKU数': skus,
            '均售价(¥)': round(avg_price, 0),
            'GMV(亿)': round(total_gmv, 2),
            '总销量(万)': round(total_sales/1e4, 1),
            '评论率(%)': round(comment_rate*100, 1),
            '爆款率(%)': round(hit_rate*100, 1),
            '价格带覆盖': coverage,
            '品牌溢价指数': round(avg_price / max(all_avg_price, 1) * 100, 0),
            '销量效率指数': round((total_sales/skus) / max(all_avg_sales, 1) * 100, 0),
        })

    return pd.DataFrame(scores)

def compute_rfm(df):
    """Compute RFM for each customer"""
    now = df['订单日期'].max()
    if not isinstance(now, datetime):
        now = datetime(2025, 10, 1)

    rfm_list = []
    for cust, grp in df.groupby('客户编码'):
        recency = (now - grp['订单日期'].max()).days
        frequency = len(grp['订单编码'].unique())
        monetary = grp['金额'].sum()
        rfm_list.append({
            '客户编码': cust,
            'Recency(天)': recency,
            'Frequency': frequency,
            'Monetary(¥)': monetary,
        })

    rfm_df = pd.DataFrame(rfm_list)
    if len(rfm_df) >= 3:
        rfm_df['R_Score'] = pd.qcut(rfm_df['Recency(天)'], 4, labels=[4,3,2,1]).astype(int)
        rfm_df['F_Score'] = pd.qcut(rfm_df['Frequency'].rank(method='first'), 4, labels=[1,2,3,4]).astype(int)
        rfm_df['M_Score'] = pd.qcut(rfm_df['Monetary(¥)'].rank(method='first'), 4, labels=[1,2,3,4]).astype(int)
        rfm_df['RFM总分'] = rfm_df['R_Score'] + rfm_df['F_Score'] + rfm_df['M_Score']

        def tier(r, f, m):
            total = r + f + m
            if total >= 10: return '核心客户'
            elif total >= 7: return '重要客户'
            elif total >= 5: return '一般客户'
            else: return '流失风险'

        rfm_df['客户层级'] = rfm_df.apply(lambda x: tier(x['R_Score'], x['F_Score'], x['M_Score']), axis=1)
    else:
        rfm_df['R_Score'] = 0; rfm_df['F_Score'] = 0; rfm_df['M_Score'] = 0
        rfm_df['RFM总分'] = 0; rfm_df['客户层级'] = '数据不足'

    return rfm_df

def compute_monthly_trend(df):
    """Compute monthly sales trend with MoM growth"""
    df = df.copy()
    df['月份'] = df['订单日期'].apply(lambda x: f'{x.year}-{x.month:02d}' if isinstance(x, datetime) else None)
    monthly = df.groupby('月份').agg(
        销售额=('金额', 'sum'),
        订单数=('订单编码', 'nunique'),
        销售数量=('订购数量', 'sum'),
    ).reset_index().dropna()
    monthly['环比增长(%)'] = monthly['销售额'].pct_change() * 100
    return monthly

def compute_province_stats(df):
    """Compute province-level statistics with coordinates for mapping"""
    province_stats = df.groupby('所在省份').agg(
        销售额=('金额', 'sum'),
        订单数=('订单编码', 'nunique'),
        客户数=('客户编码', 'nunique'),
        销售数量=('订购数量', 'sum'),
    ).reset_index()
    province_stats['客单价'] = province_stats['销售额'] / province_stats['客户数']
    province_stats['lat'] = province_stats['所在省份'].map(lambda x: PROVINCE_COORDS.get(x, (None, None))[1])
    province_stats['lon'] = province_stats['所在省份'].map(lambda x: PROVINCE_COORDS.get(x, (None, None))[0])
    province_stats = province_stats.dropna(subset=['lat', 'lon'])
    return province_stats

def compute_region_tree(df):
    """Compute region → province hierarchy for treemap"""
    # Clean region names: merge near-duplicates
    df_clean = df.copy()
    region_map = {}
    for r in df_clean['所在区域'].unique():
        cleaned = str(r).strip().replace(' ', '')
        region_map[r] = cleaned
    df_clean['区域_clean'] = df_clean['所在区域'].map(region_map)

    data = []
    for (reg, prov), grp in df_clean.groupby(['区域_clean', '所在省份']):
        amt = round(grp['金额'].sum() / 1e8, 2)
        if amt <= 0:
            continue
        data.append({
            '区域': reg,
            '省份': prov,
            '销售额(亿)': amt,
            '客户数': grp['客户编码'].nunique(),
        })
    result = pd.DataFrame(data)
    # Drop rows that may cause treemap issues
    result = result[result['销售额(亿)'] > 0]
    return result

# ============================================================
# VISUALIZATION FUNCTIONS
# ============================================================
def plot_brand_gmv_bar(scores_df, n=15, sort_by='GMV(亿)'):
    """Horizontal bar chart of brand ranking, sortable"""
    top = scores_df.nlargest(n, sort_by).sort_values(sort_by)
    fig = px.bar(
        top, x=sort_by, y='品牌', orientation='h',
        text=sort_by, color=sort_by,
        color_continuous_scale='Blues',
        title=f'Top {n} 品牌排名（按{sort_by}）'
    )
    if 'GMV' in sort_by or '金额' in sort_by:
        fig.update_traces(texttemplate='¥%{text:.1f}亿', textposition='outside')
    elif '率' in sort_by:
        fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
    elif '价' in sort_by:
        fig.update_traces(texttemplate='¥%{text:.0f}', textposition='outside')
    else:
        fig.update_traces(texttemplate='%{text}', textposition='outside')
    fig.update_layout(height=500, yaxis={'categoryorder': 'total ascending'})
    return fig

def plot_brand_radar(scores_df, selected_brands):
    """Radar chart for selected brands"""
    dimensions = ['均售价(¥)', 'GMV(亿)', 'SKU数', '爆款率(%)', '评论率(%)', '价格带覆盖']

    # Normalize each dimension to 0-100
    norm_df = scores_df.copy()
    for dim in dimensions:
        max_v = max(norm_df[dim].max(), 1)
        norm_df[f'{dim}_norm'] = norm_df[dim] / max_v * 100

    fig = go.Figure()
    for brand in selected_brands:
        bd = norm_df[norm_df['品牌'] == brand]
        if len(bd) == 0: continue
        values = [bd[f'{dim}_norm'].values[0] for dim in dimensions]
        values_raw = [bd[dim].values[0] for dim in dimensions]

        fig.add_trace(go.Scatterpolar(
            r=values,
            theta=dimensions,
            fill='toself',
            name=f'{brand}',
            hovertemplate='%{theta}: %{customdata}<extra></extra>',
            customdata=[f'{v:.1f}' for v in values_raw],
        ))

    fig.update_layout(
        polar=dict(radialaxis=dict(range=[0, 105], showticklabels=False)),
        height=500,
        title='品牌竞争力雷达图（归一化对比）'
    )
    return fig

def plot_price_vs_sales_scatter(df, selected_brands):
    """Scatter plot: price vs sales, bubble size = GMV"""
    plot_df = df[df['品牌'].isin(selected_brands)].copy()
    plot_df['gmv_m'] = plot_df['gmv'] / 1e6

    fig = px.scatter(
        plot_df, x='price', y='sale_count', size='gmv_m',
        color='品牌', hover_name='title',
        log_y=True,
        title='价格-销量散点图（气泡大小=GMV）',
        labels={'price': '价格(¥)', 'sale_count': '销量(件)', 'gmv_m': 'GMV(百万)'},
        height=550,
    )
    fig.update_layout(legend=dict(orientation='h', y=-0.2))
    return fig

def plot_province_map(prov_stats):
    """China province bubble map"""
    china_provs = prov_stats[
        (prov_stats['lat'] >= 18) & (prov_stats['lat'] <= 54) &
        (prov_stats['lon'] >= 73) & (prov_stats['lon'] <= 135)
    ].copy()

    if len(china_provs) == 0:
        fig = go.Figure()
        fig.update_layout(title='省域销售分布地图（无数据）', height=400)
        return fig

    max_sales = china_provs['销售额'].max()
    sizes = [18 + (s / max_sales) * 42 for s in china_provs['销售额']]

    fig = go.Figure()
    fig.add_trace(go.Scattergeo(
        lat=china_provs['lat'], lon=china_provs['lon'],
        marker=dict(
            size=sizes, color=china_provs['销售额'],
            colorscale='Reds', showscale=True,
            colorbar=dict(title='销售额(¥)', x=0.85, len=0.7),
            line=dict(width=1.5, color='#c0392b'),
        ),
        text=china_provs.apply(
            lambda r: f"<b>{r['所在省份']}</b><br>销售额: ¥{r['销售额']/1e8:.2f}亿<br>客户数: {int(r['客户数'])}",
            axis=1
        ),
        hoverinfo='text', mode='markers',
    ))
    fig.update_geos(
        scope='asia', center=dict(lat=35, lon=104),
        projection_scale=3.5, showcountries=True,
        countrycolor='#ddd', showcoastlines=True,
        coastlinecolor='#999', showland=True, landcolor='#fafafa',
        showocean=True, oceancolor='#eef5fb', fitbounds=False,
    )
    fig.update_layout(
        title='省域销售分布地图',
        height=600, margin=dict(l=10, r=10, t=40, b=10),
        geo=dict(lataxis=dict(range=[15,55], showgrid=False),
                 lonaxis=dict(range=[70,140], showgrid=False)),
    )
    return fig

def plot_province_bar(prov_stats):
    """Province bar chart sorted by sales"""
    sorted_df = prov_stats.nlargest(15, '销售额').sort_values('销售额')
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=sorted_df['所在省份'], x=sorted_df['销售额'],
        orientation='h',
        marker=dict(color=sorted_df['销售额'], colorscale='Reds', showscale=True,
                     colorbar=dict(title='销售额(¥)')),
        text=sorted_df['销售额'].apply(lambda x: f'¥{x/1e8:.1f}亿'),
        textposition='outside',
    ))
    fig.update_layout(
        title='省份销售额排名（Top 15）',
        height=500,
        xaxis_title='销售额(¥)',
        yaxis=dict(categoryorder='total ascending'),
    )
    return fig

def plot_region_treemap(region_df):
    """Treemap showing region → province hierarchy"""
    # Clean data: remove rows with zero/NaN values and empty paths
    clean_df = region_df.copy()
    clean_df = clean_df[clean_df['销售额(亿)'] > 0]
    clean_df = clean_df[clean_df['区域'].notna() & (clean_df['区域'] != '')]
    clean_df = clean_df[clean_df['省份'].notna() & (clean_df['省份'] != '')]

    if len(clean_df) == 0:
        fig = go.Figure()
        fig.update_layout(title='区域-省份销售额构成（无数据）', height=400)
        return fig

    try:
        fig = px.treemap(
            clean_df,
            path=['区域', '省份'],
            values='销售额(亿)',
            color='销售额(亿)',
            color_continuous_scale='RdYlGn',
            title='区域-省份销售额构成（亿元）',
            height=550,
        )
        fig.update_traces(texttemplate='%{label}<br>¥%{value:.1f}亿')
        return fig
    except Exception:
        # Fallback: simple bar chart
        fig = px.bar(
            clean_df.groupby('区域')['销售额(亿)'].sum().reset_index(),
            x='区域', y='销售额(亿)',
            title='区域销售额构成（亿元）',
            color='销售额(亿)',
            color_continuous_scale='RdYlGn',
        )
        return fig

def plot_monthly_trend(monthly_df):
    """Monthly trend with bar + line for MoM growth"""
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Bar(x=monthly_df['月份'], y=monthly_df['销售额'], name='销售额',
               marker_color='steelblue', hovertemplate='%{x}<br>¥%{y:,.0f}'),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(x=monthly_df['月份'], y=monthly_df['环比增长(%)'], name='环比增长%',
                   mode='lines+markers', line=dict(color='red', width=2, dash='dash'),
                   marker=dict(size=8)),
        secondary_y=True,
    )
    fig.update_layout(
        title='月度销售额趋势 & 环比增长',
        height=450,
        hovermode='x unified',
        legend=dict(orientation='h', y=1.1),
    )
    fig.update_yaxes(title_text='销售额(¥)', secondary_y=False)
    fig.update_yaxes(title_text='环比增长(%)', secondary_y=True)
    return fig

def plot_rfm_scatter(rfm_df):
    """RFM 3D scatter plot"""
    color_map = {'核心客户': '#2ecc71', '重要客户': '#3498db', '一般客户': '#f39c12', '流失风险': '#e74c3c'}

    fig = go.Figure()
    for tier in ['核心客户', '重要客户', '一般客户', '流失风险']:
        td = rfm_df[rfm_df['客户层级'] == tier]
        if len(td) == 0: continue
        fig.add_trace(go.Scatter3d(
            x=td['Recency(天)'], y=td['Frequency'], z=td['Monetary(¥)']/1e4,
            mode='markers', name=f'{tier}({len(td)})',
            marker=dict(size=6, color=color_map.get(tier, '#999'), opacity=0.7),
            hovertemplate='%{text}<extra></extra>',
            text=[f'{c}<br>R:{r}天 F:{f} M:¥{m/1e4:.1f}万' for c, r, f, m in
                  zip(td['客户编码'], td['Recency(天)'], td['Frequency'], td['Monetary(¥)'])],
        ))
    fig.update_layout(
        title='客户RFM三维分布',
        height=550,
        scene=dict(
            xaxis_title='Recency(距最后购买天数)',
            yaxis_title='Frequency(购买频次)',
            zaxis_title='Monetary(万元)',
        ),
        legend=dict(orientation='h', y=1.1),
    )
    return fig

def plot_rfm_pie(rfm_df):
    """Pie chart of customer tiers"""
    tier_counts = rfm_df['客户层级'].value_counts()
    colors = {'核心客户': '#2ecc71', '重要客户': '#3498db', '一般客户': '#f39c12', '流失风险': '#e74c3c'}
    fig = go.Figure(go.Pie(
        labels=tier_counts.index,
        values=tier_counts.values,
        marker=dict(colors=[colors.get(t, '#999') for t in tier_counts.index]),
        hole=0.4,
        texttemplate='%{label}<br>%{value}位(%{percent})',
    ))
    fig.update_layout(title='客户层级构成', height=400)
    return fig

def plot_refund_by_region(df):
    """Refund analysis by region"""
    region_stats = df.groupby('收货地址').agg(
        订单数=('订单编号', 'count'),
        总金额=('总金额', 'sum'),
        实付金额=('实付金额', 'sum'),
        退款金额=('退款金额', 'sum'),
    ).reset_index()
    region_stats['退款率'] = region_stats['退款金额'] / region_stats['实付金额'].replace(0, np.nan) * 100
    region_stats['未付款率'] = (1 - region_stats['实付金额'] / region_stats['总金额'].replace(0, np.nan)) * 100
    region_stats = region_stats[region_stats['订单数'] >= 50].sort_values('退款率', ascending=False)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=region_stats['收货地址'].head(15), x=region_stats['退款率'].head(15),
        orientation='h', name='退款率(%)', marker_color='#e74c3c',
        text=region_stats['退款率'].head(15).apply(lambda x: f'{x:.1f}%'),
        textposition='outside',
    ))
    fig.update_layout(
        title='各地区退款率排名（剔除<50单的地区）',
        height=500,
        xaxis_title='退款率(%)',
        yaxis=dict(categoryorder='total ascending'),
    )
    return fig

def plot_refund_daily(df):
    """Daily refund trend"""
    df_clean = df[df['创建时间'].notna()].copy()
    df_clean['日期'] = df_clean['创建时间'].apply(lambda x: x.date() if isinstance(x, datetime) else None)
    daily = df_clean.groupby('日期').agg(
        订单数=('订单编号', 'count'),
        实付总额=('实付金额', 'sum'),
        退款总额=('退款金额', 'sum'),
    ).reset_index()
    daily['退款率'] = daily['退款总额'] / daily['实付总额'].replace(0, np.nan) * 100
    daily = daily.dropna(subset=['日期'])

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Scatter(x=daily['日期'], y=daily['实付总额'], name='实付总额',
                   mode='lines', line=dict(color='steelblue', width=2), fill='tozeroy'),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(x=daily['日期'], y=daily['退款率'], name='退款率%',
                   mode='lines', line=dict(color='red', width=2)),
        secondary_y=True,
    )
    fig.update_layout(title='每日销售额 & 退款率趋势', height=450, hovermode='x unified')
    fig.update_yaxes(title_text='销售额(¥)', secondary_y=False)
    fig.update_yaxes(title_text='退款率(%)', secondary_y=True)
    return fig

def plot_price_distribution(df):
    """Price band distribution pie"""
    bands = []
    for p in df['price']:
        if p <= 50: bands.append('0-50元')
        elif p <= 100: bands.append('50-100元')
        elif p <= 200: bands.append('100-200元')
        elif p <= 500: bands.append('200-500元')
        elif p <= 1000: bands.append('500-1000元')
        else: bands.append('1000元以上')
    band_counts = Counter(bands)
    order = ['0-50元', '50-100元', '100-200元', '200-500元', '500-1000元', '1000元以上']
    values = [band_counts.get(k, 0) for k in order]

    fig = go.Figure(go.Pie(
        labels=order, values=values, hole=0.4,
        texttemplate='%{label}<br>%{value} SKU(%{percent})',
        marker=dict(colors=px.colors.sequential.Blues_r),
    ))
    fig.update_layout(title='SKU价格带分布', height=400)
    return fig

# ============================================================
# MAIN APP
# ============================================================
def main():
    # Sidebar
    with st.sidebar:
        st.image("https://img.icons8.com/color/96/000000/geography.png", width=60)
        st.title("📊 分析平台")
        st.markdown("---")
        st.markdown("### 📁 数据文件")
        st.markdown("""
        - `tmall_order_report.xlsx`
        - `双十一淘宝美妆数据.xlsx`
        - `日化.xlsx`
        """)
        st.markdown("---")
        st.markdown("### ⚙️ 全局设置")

        # Global filters
        selected_brands_global = st.multiselect(
            "筛选品牌（影响Tab2）",
            options=[],
            help="加载数据后自动填充品牌列表"
        )

        st.markdown("---")
        st.markdown("### 📥 数据导出")
        if st.button("📄 导出分析报告(HTML)", use_container_width=True):
            st.info("请切换到对应Tab导出")

        st.markdown("---")
        st.caption(f"© 2025 项目一组 | L4 系统化分析平台")

    # ============================================================
    # LOAD DATA
    # ============================================================
    import os
    base_dir = os.path.dirname(os.path.abspath(__file__))

    tmall_path = os.path.join(base_dir, 'tmall_order_report.xlsx')
    beauty_path = os.path.join(base_dir, '双十一淘宝美妆数据.xlsx')
    rihua_path = os.path.join(base_dir, '日化.xlsx')

    # Check files exist
    missing = []
    for name, path in [('天猫订单', tmall_path), ('双十一美妆', beauty_path), ('日化数据', rihua_path)]:
        if not os.path.exists(path):
            missing.append(name)

    if missing:
        st.error(f"❌ 缺少数据文件: {', '.join(missing)}")
        st.info("请确保三个 Excel 文件放在与本程序相同的目录下")
        st.stop()

    with st.spinner("正在加载数据..."):
        df_tmall = load_tmall_data(tmall_path)
        df_beauty = load_beauty_data(beauty_path)
        df_sales, df_product = load_rihua_data(rihua_path)

    # Update brand selector
    all_brands = sorted(df_beauty['品牌'].unique().tolist())

    # ============================================================
    # TABS
    # ============================================================
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "📋 数据概览",
        "🏷️ 品牌竞争力",
        "🏭 渠道健康度",
        "🛒 零售效率",
        "👤 用户画像",
        "🔗 联动洞察",
        "📊 销售策略",
        "🎁 套餐推荐"
    ])

    # ============================================================
    # TAB 1: 数据概览
    # ============================================================
    with tab1:
        st.title("📋 数据概览")
        st.markdown("> 三份数据覆盖**品牌→渠道→零售**完整产业链，共计 **86,960** 条记录")

        # KPI Row
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("📦 数据文件", "3", "Excel")
        with col2:
            st.metric("🏷️ 品牌数", f"{len(all_brands)}", "双十一美妆")
        with col3:
            st.metric("🏭 经销商客户", f"{df_sales['客户编码'].nunique()}", "日化数据")
        with col4:
            st.metric("🛒 天猫订单", f"{len(df_tmall):,}", "2025年2月")
        with col5:
            total_gmv = df_beauty['gmv'].sum() / 1e8
            total_b2b = df_sales['金额'].sum() / 1e8
            st.metric("💰 总交易规模", f"¥{total_gmv + total_b2b:.0f}亿", "估算")

        st.markdown("---")

        # Dataset summaries in 3 columns
        c1, c2, c3 = st.columns(3)

        with c1:
            st.subheader("🐱 天猫订单报告")
            paid = df_tmall['实付金额'].sum()
            refund = df_tmall['退款金额'].sum()
            refund_rate = refund / paid * 100 if paid > 0 else 0
            zero_pay = (df_tmall['实付金额'] == 0).sum()
            st.markdown(f"""
            | 指标 | 数值 |
            |------|------|
            | 订单数 | **{len(df_tmall):,}** |
            | 时间范围 | 2025-02-01 ~ 2025-02-28 |
            | 订单总额 | ¥{df_tmall['总金额'].sum():,.0f} |
            | 实付金额 | ¥{paid:,.0f} |
            | 退款金额 | ¥{refund:,.0f} |
            | **退款率(金额)** | **{refund_rate:.1f}%** 🔴 |
            | 零支付订单 | {zero_pay} ({zero_pay/len(df_tmall)*100:.1f}%) |
            """)

        with c2:
            st.subheader("💄 双十一美妆数据")
            st.markdown(f"""
            | 指标 | 数值 |
            |------|------|
            | SKU总数 | **{len(df_beauty):,}** |
            | 品牌数 | {len(all_brands)} |
            | 总销量 | {df_beauty['sale_count'].sum()/1e8:.2f}亿件 |
            | GMV估算 | ¥{df_beauty['gmv'].sum()/1e8:,.1f}亿 |
            | 价格区间 | ¥{df_beauty['price'].min():.0f} ~ ¥{df_beauty['price'].max():.0f} |
            | 均价 | ¥{df_beauty['price'].mean():.0f} |
            | 评论率 | {df_beauty['comment_count'].sum()/max(df_beauty['sale_count'].sum(),1)*100:.1f}% |
            """)

        with c3:
            st.subheader("🧴 日化渠道数据")
            monthly = compute_monthly_trend(df_sales)
            months = len(monthly)
            st.markdown(f"""
            | 指标 | 数值 |
            |------|------|
            | 销售记录 | **{len(df_sales):,}** |
            | 订单数 | {df_sales['订单编码'].nunique():,} |
            | 经销商客户 | **{df_sales['客户编码'].nunique():,}** |
            | 商品SKU | {df_product['商品编号'].nunique()} |
            | 覆盖省份 | {df_sales['所在省份'].nunique()} |
            | 覆盖城市 | {df_sales['所在地市'].nunique()} |
            | 总金额 | **¥{df_sales['金额'].sum()/1e8:,.1f}亿** |
            | 月数 | {months}个月 |
            | 品类 | 护肤品({len(df_product[df_product['商品大类']=='护肤品'])}) + 彩妆({len(df_product[df_product['商品大类']=='彩妆'])}) |
            """)

        st.markdown("---")
        st.subheader("🔍 数据样本预览")
        sample_tab1, sample_tab2, sample_tab3 = st.tabs(["天猫订单", "双十一美妆", "日化销售"])

        with sample_tab1:
            st.dataframe(df_tmall.head(10), use_container_width=True, hide_index=True)

        with sample_tab2:
            st.dataframe(df_beauty[['品牌', 'title', 'price', 'sale_count', 'comment_count']].head(10),
                         use_container_width=True, hide_index=True)

        with sample_tab3:
            st.dataframe(df_sales.head(10), use_container_width=True, hide_index=True)

    # ============================================================
    # TAB 2: 品牌竞争力
    # ============================================================
    with tab2:
        st.title("🏷️ 品牌竞争力分析")
        st.markdown("> 基于双十一美妆数据，从6个维度评估22个品牌的竞争力")

        # Compute scores
        scores_df = compute_brand_scores(df_beauty)

        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            selected_brands = st.multiselect(
                "选择品牌（多选）",
                options=all_brands,
                default=all_brands[:5],
                key="tab2_brands"
            )
        with col_f2:
            price_range = st.slider(
                "价格区间(¥)",
                min_value=0, max_value=int(df_beauty['price'].max()),
                value=(0, int(df_beauty['price'].max())),
                step=10,
                key="tab2_price"
            )
        with col_f3:
            sort_by = st.selectbox(
                "排序方式",
                options=['GMV(亿)', 'SKU数', '均售价(¥)', '爆款率(%)', '评论率(%)'],
                key="tab2_sort"
            )

        # Filter data
        mask = df_beauty['品牌'].isin(selected_brands)
        mask &= (df_beauty['price'] >= price_range[0]) & (df_beauty['price'] <= price_range[1])
        filtered_beauty = df_beauty[mask]

        # --- Charts Row 1 ---
        st.markdown("### 📊 品牌GMV排名 & 价格-销量分布")
        c_left, c_right = st.columns([1, 1])

        with c_left:
            fig_gmv = plot_brand_gmv_bar(
                scores_df[scores_df['品牌'].isin(selected_brands)],
                n=min(15, len(selected_brands)),
                sort_by=sort_by
            )
            st.plotly_chart(fig_gmv, use_container_width=True)

        with c_right:
            fig_scatter = plot_price_vs_sales_scatter(filtered_beauty, selected_brands)
            st.plotly_chart(fig_scatter, use_container_width=True)

        # --- Charts Row 2 ---
        st.markdown("### 🎯 品牌雷达图对比 & 评分卡明细")
        c_left2, c_right2 = st.columns([1, 1])

        with c_left2:
            if len(selected_brands) >= 1:
                radar_brands = selected_brands[:6]
                fig_radar = plot_brand_radar(scores_df, radar_brands)
                st.plotly_chart(fig_radar, use_container_width=True)
            else:
                st.warning("请至少选择一个品牌")

        with c_right2:
            st.markdown("#### 品牌竞争力评分卡")
            display_scores = scores_df[scores_df['品牌'].isin(selected_brands)].sort_values(
                sort_by, ascending=False
            )
            st.dataframe(
                display_scores.style
                .background_gradient(subset=['GMV(亿)', '爆款率(%)', '品牌溢价指数'], cmap='Blues')
                .format({
                    '均售价(¥)': '¥{:.0f}',
                    'GMV(亿)': '{:.2f}',
                    '评论率(%)': '{:.1f}%',
                    '爆款率(%)': '{:.1f}%',
                    '品牌溢价指数': '{:.0f}',
                    '销量效率指数': '{:.0f}',
                }),
                use_container_width=True, hide_index=True,
            )

        # Price distribution
        st.markdown("### 📐 价格带分布")
        fig_price_band = plot_price_distribution(filtered_beauty)
        st.plotly_chart(fig_price_band, use_container_width=True, height=400)

        # Export
        st.markdown("---")
        if st.button("📥 导出品牌评分卡 (CSV)", key="export_brand"):
            csv = scores_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button("下载CSV", csv, "品牌竞争力评分卡.csv", "text/csv")

        # ============================================================
        # 爆款基因拆解服务
        # ============================================================
        st.markdown("---")
        st.markdown("### 🔬 爆款基因拆解服务")
        st.markdown("> 提取Top爆款的核心词汇、价格带特征，帮助商家快速找到高转化产品方向")

        # 1. 词频分析
        st.markdown("#### 1️⃣ 爆款高频词汇分析")
        
        # 筛选爆款（销量>1000）
        top_products = filtered_beauty[filtered_beauty['sale_count'] > 1000].nlargest(50, 'sale_count')
        
        if len(top_products) > 0:
            # 提取标题关键词
            keywords_list = []
            for title in top_products['title']:
                # 关键词提取逻辑
                keywords = ['补水', '保湿', '美白', '抗皱', '紧致', '舒缓', '控油', '祛痘',
                           '防晒', '隔离', '粉底', '口红', '眼影', '睫毛', '眉笔',
                           '面膜', '精华', '面霜', '爽肤水', '洁面', '眼霜', '乳液',
                           '套装', '礼盒', '男士', '孕妇', '敏感肌', '修护',
                           '补水保湿', '控油祛痘', '美白淡斑', '抗衰老']
                for kw in keywords:
                    if kw in str(title):
                        keywords_list.append(kw)
            
            # 词频统计
            keyword_counts = Counter(keywords_list)
            keyword_df = pd.DataFrame(keyword_counts.most_common(15), columns=['关键词', '出现频次'])
            
            col_kw1, col_kw2 = st.columns([1, 1])
            with col_kw1:
                fig_kw = px.bar(keyword_df, x='出现频次', y='关键词',
                               title='Top50爆款高频词汇排名',
                               color='出现频次', color_continuous_scale='Reds',
                               orientation='h')
                fig_kw.update_layout(height=400)
                st.plotly_chart(fig_kw, use_container_width=True)
            
            with col_kw2:
                st.markdown("**💡 商家赋能建议：**")
                top_kw = keyword_df.head(5)['关键词'].tolist()
                st.info(f"当前爆款热词：{', '.join(top_kw)}")
                st.markdown("""
                **可执行建议**：
                - 新品开发时优先使用上述热词作为标题核心词汇
                - 详情页文案重点强调这些功效卖点
                - 广告投放关键词覆盖这些词汇
                """)
        else:
            st.warning("当前筛选条件下无爆款商品")

        # 2. 价格带-销量关系分析
        st.markdown("#### 2️⃣ 价格带与销量关系拆解")
        
        # 价格带划分
        filtered_beauty['价格带'] = filtered_beauty['price'].apply(
            lambda x: '50元以下引流款' if x < 50 else 
                     ('50-100元大众款' if x < 100 else 
                     ('100-200元品质款' if x < 200 else 
                     ('200-500元中高端' if x < 500 else '500元以上高端款')))
        )
        
        price_band_stats = filtered_beauty.groupby('价格带').agg({
            'sale_count': ['sum', 'mean', 'count'],
            'gmv': 'sum',
            'price': 'mean'
        }).reset_index()
        price_band_stats.columns = ['价格带', '总销量', '平均销量', 'SKU数量', '总GMV', '平均单价']
        price_band_stats = price_band_stats.sort_values('总GMV', ascending=False)
        
        col_pb1, col_pb2 = st.columns([1.2, 1])
        with col_pb1:
            fig_pb = px.bar(price_band_stats, x='价格带', y=['总销量', '总GMV'],
                           title='各价格带销量与GMV对比',
                           barmode='group',
                           color_discrete_map={'总销量': '#3498db', '总GMV': '#e74c3c'})
            st.plotly_chart(fig_pb, use_container_width=True)
        
        with col_pb2:
            st.markdown("**📊 价格带分析结论：**")
            st.dataframe(price_band_stats.style
                        .format({'总销量': '{:,}', '平均销量': '{:.0f}', '总GMV': '¥{:.0f}', '平均单价': '¥{:.0f}'})
                        .background_gradient(subset=['总GMV'], cmap='Greens'),
                        use_container_width=True)
            
            # 最佳价格带建议
            best_band = price_band_stats.nlargest(1, '总GMV')['价格带'].values[0]
            st.success(f"🎯 最佳价格带：{best_band}，建议主推该价格区间产品")

        # 3. 价格弹性测算
        st.markdown("#### 3️⃣ 大促价格弹性测算")
        
        # 按日期统计价格与销量变化
        if 'update_time' in filtered_beauty.columns:
            daily_stats = filtered_beauty.groupby('update_time').agg({
                'price': 'mean',
                'sale_count': 'sum',
                'gmv': 'sum'
            }).reset_index()
            
            if len(daily_stats) > 1:
                daily_stats['日期'] = pd.to_datetime(daily_stats['update_time']).dt.strftime('%m-%d')
                daily_stats['销量环比'] = daily_stats['sale_count'].pct_change()
                daily_stats['价格环比'] = daily_stats['price'].pct_change()
                
                # 价格弹性系数 = 销量变化率 / 价格变化率
                daily_stats['弹性系数'] = daily_stats.apply(
                    lambda r: abs(r['销量环比'] / r['价格环比']) if r['价格环比'] != 0 and pd.notna(r['销量环比']) else 0,
                    axis=1
                )
                
                fig_elastic = make_subplots(specs=[[{"secondary_y": True}]])
                fig_elastic.add_trace(
                    go.Scatter(x=daily_stats['日期'], y=daily_stats['price'], name='平均价格',
                              line=dict(color='#e74c3c', width=2)),
                    secondary_y=False
                )
                fig_elastic.add_trace(
                    go.Scatter(x=daily_stats['日期'], y=daily_stats['sale_count'], name='销量',
                              line=dict(color='#3498db', width=2)),
                    secondary_y=True
                )
                fig_elastic.update_layout(title='价格-销量双轴趋势图', height=400)
                fig_elastic.update_yaxes(title_text="平均价格(¥)", secondary_y=False)
                fig_elastic.update_yaxes(title_text="销量", secondary_y=True)
                st.plotly_chart(fig_elastic, use_container_width=True)
                
                avg_elasticity = daily_stats['弹性系数'].mean()
                st.markdown(f"**平均价格弹性系数：{avg_elasticity:.2f}**")
                if avg_elasticity > 1:
                    st.info(f"弹性>1：消费者对价格敏感，降价1%可提升销量{avg_elasticity:.1f}%，建议大促期间适当降价")
                else:
                    st.info(f"弹性<1：消费者对价格不敏感，降价效果有限，建议维持原价或小幅优惠")
            else:
                st.warning("数据日期跨度不足，无法计算价格弹性")
        else:
            st.warning("数据缺少时间维度，无法进行价格弹性分析")

        # 4. 品类生命周期判断
        st.markdown("#### 4️⃣ 品类生命周期判断")
        
        # 从标题提取品类关键词
        def extract_category(title):
            categories = {
                '面膜': ['面膜', '面贴膜', '眼膜'],
                '面霜': ['面霜', '霜'],
                '精华': ['精华', '精华液'],
                '爽肤水': ['爽肤水', '水', '乳液', '乳'],
                '洁面': ['洁面', '洗面奶', '洁面乳'],
                '眼霜': ['眼霜'],
                '防晒': ['防晒', '防晒霜'],
                '口红': ['口红', '唇膏'],
                '粉底': ['粉底', '粉底液', '粉饼'],
                '套装': ['套装', '礼盒', '组合']
            }
            for cat, keywords in categories.items():
                for kw in keywords:
                    if kw in str(title):
                        return cat
            return '其他'
        
        filtered_beauty['商品品类'] = filtered_beauty['title'].apply(extract_category)
        
        # 品类销量对比
        category_stats = filtered_beauty.groupby('商品品类').agg({
            'sale_count': 'sum',
            'gmv': 'sum',
            'price': 'mean'
        }).reset_index()
        category_stats.columns = ['商品品类', '总销量', '总GMV', '平均单价']
        category_stats = category_stats.sort_values('总GMV', ascending=False)
        category_stats['销量占比'] = category_stats['总销量'] / category_stats['总销量'].sum() * 100
        
        fig_cat = px.bar(category_stats, x='商品品类', y='总销量',
                        title='各品类销量分布',
                        color='总GMV', color_continuous_scale='Purples')
        st.plotly_chart(fig_cat, use_container_width=True)
        
        # 品类建议
        top_cat = category_stats.nlargest(3, '总销量')['商品品类'].tolist()
        st.markdown(f"**🔥 当前热销品类：{', '.join(top_cat)}**")
        st.markdown("""
        **商家赋能建议**：
        - 将资源倾斜至热销品类，增加库存备货
        - 冷门品类可考虑清仓或捆绑促销
        - 关注新兴品类趋势，提前布局
        """)

    # ============================================================
    # TAB 3: 渠道健康度
    # ============================================================
    with tab3:
        st.title("🏭 渠道健康度诊断")
        st.markdown("> 日化B2B经销商网络分析：地图分布、客户RFM分层、月度趋势")

        # Compute stats
        rfm_df = compute_rfm(df_sales)
        monthly_df = compute_monthly_trend(df_sales)
        prov_stats = compute_province_stats(df_sales)
        region_tree = compute_region_tree(df_sales)

        # --- KPI Row ---
        k1, k2, k3, k4 = st.columns(4)
        risk_count = len(rfm_df[rfm_df['客户层级'] == '流失风险']) if '客户层级' in rfm_df.columns else 0
        core_count = len(rfm_df[rfm_df['客户层级'] == '核心客户']) if '客户层级' in rfm_df.columns else 0

        with k1:
            latest_m = monthly_df['销售额'].iloc[-1] if len(monthly_df) > 0 else 0
            prev_m = monthly_df['销售额'].iloc[-2] if len(monthly_df) > 1 else latest_m
            delta = f"{(latest_m/prev_m - 1)*100:+.1f}%" if prev_m > 0 else "N/A"
            st.metric("📈 最新月销售额", f"¥{latest_m/1e8:.1f}亿", delta)
        with k2:
            st.metric("👥 总客户数", f"{len(rfm_df)}", "")
        with k3:
            st.metric("🟢 核心客户", f"{core_count}", f"{core_count/len(rfm_df)*100:.0f}%")
        with k4:
            st.metric("🔴 流失风险", f"{risk_count}", f"{risk_count/len(rfm_df)*100:.0f}%" if len(rfm_df) > 0 else "")

        # --- Charts Row 1: Map + Bar ---
        st.markdown("### 🗺️ 省域销售分布")
        c_map, c_bar = st.columns([1.2, 1])

        with c_map:
            if len(prov_stats) > 0:
                fig_map = plot_province_map(prov_stats)
                st.plotly_chart(fig_map, use_container_width=True)
            else:
                st.warning("暂无足够的地图数据")

        with c_bar:
            fig_bar = plot_province_bar(prov_stats)
            st.plotly_chart(fig_bar, use_container_width=True)

        # --- Charts Row 2: Treemap + Monthly Trend ---
        st.markdown("### 🌳 区域构成 & 月度趋势")
        c_tree, c_month = st.columns([1, 1])

        with c_tree:
            fig_tree = plot_region_treemap(region_tree)
            st.plotly_chart(fig_tree, use_container_width=True)

        with c_month:
            fig_month = plot_monthly_trend(monthly_df)
            st.plotly_chart(fig_month, use_container_width=True)

        # --- Province Detail ---
        st.markdown("### 📋 省份数据明细")
        prov_display = prov_stats.sort_values('销售额', ascending=False)
        st.dataframe(
            prov_display.style
            .format({'销售额': '¥{:,.0f}', '客单价': '¥{:,.0f}'})
            .background_gradient(subset=['销售额', '客户数'], cmap='Reds'),
            use_container_width=True, hide_index=True,
            height=400,
        )

        # ============================================================
        # 区域空白市场扫描
        # ============================================================
        st.markdown("---")
        st.markdown("### 🗺️ 区域空白市场扫描")
        st.markdown("> 通过省份→地市下钻，识别空白覆盖区域，指导地推团队开荒")

        # 获取所有省份和地市
        all_provinces = df_sales['所在省份'].unique()
        all_cities = df_sales['所在地市'].unique() if '所在地市' in df_sales.columns else []

        # 统计每个省份的客户覆盖
        province_coverage = df_sales.groupby('所在省份').agg({
            '客户编码': 'nunique',
            '金额': 'sum'
        }).reset_index()
        province_coverage.columns = ['省份', '客户数', '销售额']

        # 地市覆盖统计
        if '所在地市' in df_sales.columns:
            city_coverage = df_sales.groupby(['所在省份', '所在地市']).agg({
                '客户编码': 'nunique',
                '金额': 'sum'
            }).reset_index()
            city_coverage.columns = ['省份', '地市', '客户数', '销售额']
            city_coverage['覆盖状态'] = city_coverage.apply(
                lambda r: '🟢 已覆盖' if r['客户数'] > 0 else '🔴 空白',
                axis=1
            )
            
            # 空白地市列表
            blank_cities = city_coverage[city_coverage['客户数'] == 0]
            covered_cities = city_coverage[city_coverage['客户数'] > 0]

        col_blank1, col_blank2 = st.columns([1, 1])
        
        with col_blank1:
            st.markdown("#### 📊 省域覆盖热力图")
            fig_prov_cov = px.bar(province_coverage.sort_values('销售额', ascending=False).head(15),
                                 x='省份', y='销售额',
                                 title='TOP15省份销售额',
                                 color='客户数', color_continuous_scale='Blues')
            st.plotly_chart(fig_prov_cov, use_container_width=True)

        with col_blank2:
            st.markdown("#### 📋 覆盖状态统计")
            total_cities = len(city_coverage) if '所在地市' in df_sales.columns else 0
            blank_count = len(blank_cities) if '所在地市' in df_sales.columns else 0
            covered_count = len(covered_cities) if '所在地市' in df_sales.columns else 0
            
            c_stat1, c_stat2, c_stat3 = st.columns(3)
            c_stat1.metric("总地市数", total_cities)
            c_stat2.metric("已覆盖", covered_count, f"{covered_count/total_cities*100:.0f}%" if total_cities > 0 else "")
            c_stat3.metric("空白区域", blank_count, f"{blank_count/total_cities*100:.0f}%" if total_cities > 0 else "")

        # 空白市场明细表
        if '所在地市' in df_sales.columns and len(blank_cities) > 0:
            st.markdown("#### 🔴 空白市场清单（建议开荒）")
            st.dataframe(blank_cities[['省份', '地市', '覆盖状态']],
                        use_container_width=True, height=300)
            
            st.markdown("""
            **💡 商家赋能建议**：
            - 上述空白地市建议安排地推团队进行客户开发
            - 优先选择销售额较高省份的空白地市
            - 可配合线上广告投放进行区域定向
            """)
        else:
            st.success("✅ 所有地市均已覆盖，市场渗透率良好")

        # ============================================================
        # 窜货预警逻辑模型
        # ============================================================
        st.markdown("---")
        st.markdown("### ⚠️ 窜货预警逻辑模型")
        st.markdown("> 线上线下数据联动，识别疑似窜货行为")

        st.markdown("""
        **预警逻辑**：
        1. B2B端某省某品类提货量环比增长 > 30%
        2. AND 天猫端该省收货订单中，客单价 < 天猫官方均价 × 0.8
        3. AND 天猫端该省退款率 > 整体均值
        4. → 预警等级 = HIGH，疑似窜货地 = '某省'
        """)

        # 合并 df_sales 和 df_product 获取品类信息
        df_sales_merged = df_sales.merge(df_product[['商品编号', '商品大类']], 
                                         on='商品编号', how='left')
        
        # 计算B2B端省份品类数据
        if '商品大类' in df_sales_merged.columns:
            b2b_region_cat = df_sales_merged.groupby(['所在省份', '商品大类']).agg({
                '金额': 'sum',
                '订购数量': 'sum',
                '客户编码': 'nunique'
            }).reset_index()
            b2b_region_cat.columns = ['省份', '品类', 'B2B进货额', 'B2B进货量', 'B2B客户数']
            
            # 天猫端省份数据
            df_tmall_for_alert = df_tmall[df_tmall['创建时间'].notna()].copy()
            tmall_region = df_tmall_for_alert.groupby('收货地址').agg({
                '实付金额': ['sum', 'mean'],
                '退款金额': 'sum',
                '订单编号': 'count'
            }).reset_index()
            tmall_region.columns = ['省份', '天猫销售额', '天猫均价', '天猫退款额', '天猫订单数']
            tmall_region['天猫退款率'] = tmall_region['天猫退款额'] / tmall_region['天猫销售额']
            
            # 整体退款率均值
            overall_refund_rate = df_tmall_for_alert['退款金额'].sum() / df_tmall_for_alert['实付金额'].sum()
            overall_avg_price = df_tmall_for_alert['总金额'].mean()
            
            # 窜货预警检测
            alert_list = []
            
            for prov in b2b_region_cat['省份'].unique():
                b2b_data = b2b_region_cat[b2b_region_cat['省份'] == prov]
                tmall_data = tmall_region[tmall_region['省份'] == prov]
                
                if len(tmall_data) > 0:
                    # 检测条件
                    prov_avg_price = tmall_data['天猫均价'].values[0]
                    prov_refund_rate = tmall_data['天猫退款率'].values[0]
                    
                    # 条件1: 价格低于官方均价80%
                    price_alert = prov_avg_price < overall_avg_price * 0.8
                    
                    # 条件2: 退款率高于整体均值
                    refund_alert = prov_refund_rate > overall_refund_rate
                    
                    # 条件3: B2B进货量较大
                    b2b_amount = b2b_data['B2B进货额'].sum()
                    high_b2b = b2b_amount > b2b_region_cat['B2B进货额'].quantile(0.75)
                    
                    # 综合判断
                    if price_alert and refund_alert and high_b2b:
                        alert_list.append({
                            '省份': prov,
                            '天猫均价': prov_avg_price,
                            '官方均价': overall_avg_price,
                            '价格偏离': f"{(prov_avg_price/overall_avg_price - 1)*100:.1f}%",
                            '退款率': prov_refund_rate,
                            '整体退款率': overall_refund_rate,
                            'B2B进货额': b2b_amount,
                            '预警等级': '🔴 HIGH'
                        })
                    elif price_alert and refund_alert:
                        alert_list.append({
                            '省份': prov,
                            '天猫均价': prov_avg_price,
                            '官方均价': overall_avg_price,
                            '价格偏离': f"{(prov_avg_price/overall_avg_price - 1)*100:.1f}%",
                            '退款率': prov_refund_rate,
                            '整体退款率': overall_refund_rate,
                            'B2B进货额': b2b_amount,
                            '预警等级': '🟡 MEDIUM'
                        })
            
            if len(alert_list) > 0:
                alert_df = pd.DataFrame(alert_list)
                st.markdown("#### 🔴 窜货预警报告")
                st.dataframe(alert_df.style
                            .format({'天猫均价': '¥{:.0f}', '官方均价': '¥{:.0f}', 
                                    '退款率': '{:.1%}', '整体退款率': '{:.1%}',
                                    'B2B进货额': '¥{:.0f}'})
                            .background_gradient(subset=['退款率'], cmap='Reds'),
                            use_container_width=True)
                
                st.error(f"⚠️ 发现 {len(alert_df)} 个疑似窜货省份，建议核实线下经销商价格管控情况")
                
                st.markdown("""
                **💡 商家赋能建议**：
                1. 对预警省份的经销商进行价格稽查
                2. 加强经销商协议中的窜货处罚条款
                3. 限制高风险省份的发货数量
                """)
            else:
                st.success("✅ 未发现窜货预警，线上线下价格体系正常")
        else:
            st.warning("B2B数据缺少品类维度，无法进行窜货预警分析")

    # ============================================================
    # TAB 4: 零售效率
    # ============================================================
    with tab4:
        st.title("🛒 零售效率分析")
        st.markdown("> 天猫订单报告：退款归因、支付行为、地域效率")

        # KPI Row
        paid = df_tmall['实付金额'].sum()
        refund = df_tmall['退款金额'].sum()
        refund_rate = refund / paid * 100 if paid > 0 else 0
        zero_pay = (df_tmall['实付金额'] == 0).sum()

        k1, k2, k3, k4 = st.columns(4)
        with k1:
            st.metric("💰 实付总额", f"¥{paid:,.0f}", "")
        with k2:
            st.metric("💸 退款总额", f"¥{refund:,.0f}", f"-{refund_rate:.1f}%")
        with k3:
            st.metric("❌ 退款率(金额)", f"{refund_rate:.1f}%", "⚠️ 偏高")
        with k4:
            st.metric("📭 未付款订单", f"{zero_pay:,}", f"{zero_pay/len(df_tmall)*100:.1f}%")

        # --- Charts Row 1: Daily trend + Region refund ---
        st.markdown("### 📈 每日趋势 & 区域退款分析")
        c1, c2 = st.columns([1, 1])

        with c1:
            fig_daily = plot_refund_daily(df_tmall)
            st.plotly_chart(fig_daily, use_container_width=True)

        with c2:
            fig_region = plot_refund_by_region(df_tmall)
            st.plotly_chart(fig_region, use_container_width=True)

        # --- Payment behavior ---
        st.markdown("### 💳 支付行为分析")
        df_tmall_clean = df_tmall[df_tmall['创建时间'].notna()].copy()

        # Payment delay analysis
        same_day = 0; delayed = 0; no_pay = 0
        pay_delays = []
        for _, row in df_tmall_clean.iterrows():
            c = row['创建时间']; p = row['付款时间']
            if isinstance(c, datetime) and isinstance(p, datetime):
                delay = (p - c).total_seconds() / 3600
                pay_delays.append(delay)
                if delay < 24: same_day += 1
                else: delayed += 1
            elif isinstance(c, datetime):
                no_pay += 1

        c_p1, c_p2 = st.columns(2)
        with c_p1:
            pay_data = pd.DataFrame({
                '支付行为': ['当日付款', '延迟付款', '未付款'],
                '订单数': [same_day, delayed, no_pay],
            })
            fig_pay = px.pie(pay_data, values='订单数', names='支付行为', hole=0.4,
                             title='支付行为分布',
                             color='支付行为',
                             color_discrete_map={'当日付款':'#2ecc71', '延迟付款':'#f39c12', '未付款':'#e74c3c'})
            fig_pay.update_traces(texttemplate='%{label}<br>%{value}单(%{percent})')
            st.plotly_chart(fig_pay, use_container_width=True)

        with c_p2:
            # Refund by amount band
            df_tmall_clean['金额段'] = pd.cut(
                df_tmall_clean['总金额'],
                bins=[0, 50, 100, 200, 500, 1000, 999999],
                labels=['0-50', '50-100', '100-200', '200-500', '500-1000', '1000+']
            )
            band_refund = df_tmall_clean.groupby('金额段', observed=False).agg(
                订单数=('订单编号', 'count'),
                退款率=('退款金额', lambda x: (x.sum() / df_tmall_clean.loc[x.index, '实付金额'].sum() * 100) if df_tmall_clean.loc[x.index, '实付金额'].sum() > 0 else 0),
            ).reset_index()

            fig_band = px.bar(
                band_refund, x='金额段', y='退款率',
                title='各金额段退款率', text=band_refund['退款率'].apply(lambda x: f'{x:.1f}%'),
                color='退款率', color_continuous_scale='Reds',
            )
            fig_band.update_traces(textposition='outside')
            st.plotly_chart(fig_band, use_container_width=True)

        # --- Refund prediction model ---
        st.markdown("### 🤖 退款风险预测模型（XGBoost）")

        st.markdown("""
        **变量定义：**
        - **因变量 Y**：是否退款（`退款金额 > 0` → 1，否则 → 0），二分类问题
        - **自变量 X**：总金额、实付金额、下单时段、是否周末、付款延迟时长、金额段、地区退款率
        - **模型**：XGBoost 分类器（预训练模型，保存在 `models/refund_xgboost.joblib`）
        """)

        # Load pre-trained model
        import joblib
        model_dir = os.path.join(base_dir, 'models')
        model_path = os.path.join(model_dir, 'refund_xgboost.joblib')
        encoder_path = os.path.join(model_dir, 'encoders.joblib')
        metrics_path = os.path.join(base_dir, 'model_metrics', 'refund_metrics.json')

        # Check if model exists
        model_loaded = False
        if os.path.exists(model_path) and os.path.exists(encoder_path):
            try:
                xgb_model = joblib.load(model_path)
                encoders = joblib.load(encoder_path)
                model_loaded = True
                st.success(f"✅ 已加载预训练 XGBoost 模型")
            except Exception as e:
                st.warning(f"模型加载失败: {e}")

        if model_loaded:
            # Display model metrics from saved file
            if os.path.exists(metrics_path):
                import json
                with open(metrics_path, 'r', encoding='utf-8') as f:
                    saved_metrics = json.load(f)
                if 'XGBoost' in saved_metrics:
                    metrics = saved_metrics['XGBoost']
                    c_m1, c_m2, c_m3, c_m4, c_m5 = st.columns(5)
                    c_m1.metric("准确率(Accuracy)", f"{metrics.get('accuracy', 0):.1%}")
                    c_m2.metric("精确率(Precision)", f"{metrics.get('precision', 0):.1%}",
                               "预测退款中真正退款的比例")
                    c_m3.metric("召回率(Recall)", f"{metrics.get('recall', 0):.1%}",
                               "实际退款中被识别出的比例")
                    c_m4.metric("F1分数", f"{metrics.get('f1', 0):.1%}", "精确率与召回率的调和")
                    c_m5.metric("AUC", f"{metrics.get('auc', 0):.1%}", "模型区分能力")

            # Feature importance
            feature_labels = ['总金额', '实付金额', '下单小时', '是否周末', 
                         '付款延迟(小时)', '金额段', '地区编码', '地区退款率']
            importances = xgb_model.feature_importances_
            importance_df = pd.DataFrame({
                '特征': feature_labels,
                '重要性': importances,
            }).sort_values('重要性', ascending=True)

            fig_imp = px.bar(
                importance_df, x='重要性', y='特征',
                title='XGBoost 特征重要性排名',
                color='重要性', color_continuous_scale='Blues',
                height=350,
            )
            st.plotly_chart(fig_imp, use_container_width=True)

            # Store in session for prediction
            st.session_state['xgb_model'] = xgb_model
            st.session_state['encoders'] = encoders
        else:
            st.error("❌ 未找到预训练模型，请先运行 `python prediction_models.py` 训练模型")
            if st.button("🔄 重新训练模型", key="retrain_model"):
                st.info("请在终端运行: python prediction_models.py")

        # ============================================================
        # Batch Refund Prediction (导入表格批量预测)
        # ============================================================
        st.markdown("---")
        st.markdown("#### 📊 批量退款风险预测（导入订单表格）")
        st.markdown("""
        **使用说明**：
        1. 上传订单表格文件（CSV 或 Excel 格式）
        2. 表格需包含以下字段：`订单编号`、`总金额`、`实付金额`、`创建时间`、`付款时间`、`收货地址`
        3. 系统自动计算退款概率，输出风险预警报告
        """)

        # File upload
        uploaded_file = st.file_uploader(
            "📤 上传订单数据文件", 
            type=['csv', 'xlsx'],
            help="支持 CSV 和 Excel 格式"
        )

        if uploaded_file is not None and 'xgb_model' in st.session_state:
            try:
                # Read uploaded file
                if uploaded_file.name.endswith('.csv'):
                    df_upload = pd.read_csv(uploaded_file)
                else:
                    df_upload = pd.read_excel(uploaded_file)
                
                st.success(f"✅ 成功加载 {len(df_upload)} 条订单数据")
                
                # Show preview
                st.markdown("**数据预览：**")
                st.dataframe(df_upload.head(10), use_container_width=True)
                
                # Check required columns
                required_cols = ['订单编号', '总金额', '实付金额', '创建时间', '付款时间', '收货地址']
                missing_cols = [col for col in required_cols if col not in df_upload.columns]
                
                if missing_cols:
                    st.warning(f"⚠️ 缺少必要字段：{missing_cols}，请检查表格格式")
                else:
                    if st.button("🔮 开始批量预测", key="batch_predict_btn"):
                        with st.spinner("正在批量预测退款风险..."):
                            # Feature engineering
                            df_pred = df_upload.copy()
                            
                            # Parse datetime
                            df_pred['创建时间'] = pd.to_datetime(df_pred['创建时间'], errors='coerce')
                            df_pred['付款时间'] = pd.to_datetime(df_pred['付款时间'], errors='coerce')
                            
                            # Calculate features
                            df_pred['下单小时'] = df_pred['创建时间'].dt.hour.fillna(12).astype(int)
                            df_pred['下单星期'] = df_pred['创建时间'].dt.weekday.fillna(3).astype(int)
                            df_pred['是否周末'] = (df_pred['下单星期'] >= 5).astype(int)
                            
                            # Payment delay (hours)
                            df_pred['付款延迟'] = (
                                (df_pred['付款时间'] - df_pred['创建时间']).dt.total_seconds() / 3600
                            ).fillna(0).clip(lower=0)
                            
                            # Amount band
                            df_pred['金额段'] = df_pred['总金额'].apply(
                                lambda x: 0 if x <= 50 else (1 if x <= 100 else (2 if x <= 200 else (3 if x <= 500 else (4 if x <= 1000 else 5))))
                            )
                            
                            # Region refund rate (use default)
                            df_pred['地区退款率'] = 0.2
                            df_pred['地区编码'] = 0
                            
                            # Build feature matrix
                            features = ['总金额', '实付金额', '下单小时', '是否周末', 
                                       '付款延迟', '金额段', '地区编码', '地区退款率']
                            X_batch = df_pred[features].fillna(0).values
                            
                            # Predict
                            probs = st.session_state['xgb_model'].predict_proba(X_batch)[:, 1]
                            
                            # Add prediction results
                            df_pred['退款概率'] = probs
                            df_pred['风险等级'] = df_pred['退款概率'].apply(
                                lambda p: '🔴 高风险' if p > 0.4 else ('🟡 中风险' if p > 0.2 else '🟢 低风险')
                            )
                            df_pred['建议措施'] = df_pred['退款概率'].apply(
                                lambda p: '发送定向优惠券，客服主动跟进' if p > 0.4 else 
                                          ('关注物流进度，发送使用指南' if p > 0.2 else '正常履约即可')
                            )
                            
                            # Output result
                            st.markdown("---")
                            st.markdown("### 📋 退款风险预警报告")
                            
                            # Statistics
                            high_risk_count = (df_pred['退款概率'] > 0.4).sum()
                            mid_risk_count = ((df_pred['退款概率'] > 0.2) & (df_pred['退款概率'] <= 0.4)).sum()
                            low_risk_count = (df_pred['退款概率'] <= 0.2).sum()
                            
                            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                            col_stat1.metric("总订单数", len(df_pred))
                            col_stat2.metric("🔴 高风险", high_risk_count, f"{high_risk_count/len(df_pred):.1%}")
                            col_stat3.metric("🟡 中风险", mid_risk_count, f"{mid_risk_count/len(df_pred):.1%}")
                            col_stat4.metric("🟢 低风险", low_risk_count, f"{low_risk_count/len(df_pred):.1%}")
                            
                            # Risk distribution chart
                            risk_dist = df_pred['风险等级'].value_counts()
                            fig_risk = px.pie(
                                values=risk_dist.values, 
                                names=risk_dist.index,
                                title='退款风险等级分布',
                                color=risk_dist.index,
                                color_discrete_map={
                                    '🔴 高风险': '#e74c3c',
                                    '🟡 中风险': '#f39c12',
                                    '🟢 低风险': '#27ae60'
                                },
                                hole=0.4
                            )
                            st.plotly_chart(fig_risk, use_container_width=True)
                            
                            # Show result table
                            result_cols = ['订单编号', '总金额', '实付金额', '收货地址', '退款概率', '风险等级', '建议措施']
                            st.markdown("**详细预测结果：**")
                            st.dataframe(
                                df_pred[result_cols].sort_values('退款概率', ascending=False),
                                use_container_width=True,
                                height=400
                            )
                            
                            # Download button
                            csv_result = df_pred[result_cols].to_csv(index=False).encode('utf-8-sig')
                            st.download_button(
                                label="📥 下载预测报告（CSV）",
                                data=csv_result,
                                file_name="refund_risk_report.csv",
                                mime='text/csv'
                            )
                            
                            # High risk orders alert
                            if high_risk_count > 0:
                                st.error(f"⚠️ 发现 {high_risk_count} 条高风险订单，建议优先处理！")
                                st.markdown("**高风险订单列表：**")
                                high_risk_orders = df_pred[df_pred['退款概率'] > 0.4][result_cols]
                                st.dataframe(high_risk_orders, use_container_width=True)
            
            except Exception as e:
                st.error(f"❌ 文件处理失败：{e}")
        
        elif uploaded_file is not None and 'xgb_model' not in st.session_state:
            st.warning("⚠️ 请先加载预训练模型（确保 models/refund_xgboost.joblib 存在）")

        # ============================================================
        # 广告时段优化建议表
        # ============================================================
        st.markdown("---")
        st.markdown("### 📢 广告分时折扣建议表")
        st.markdown("> 绘制全天24小时的\"真实支付订单\"曲线，指导广告投放时段")

        # 按小时统计支付订单
        df_paid_orders = df_tmall_clean[df_tmall_clean['实付金额'] > 0]
        if len(df_paid_orders) > 0:
            hourly_stats = df_paid_orders.groupby(df_paid_orders['创建时间'].dt.hour).agg({
                '订单编号': 'count',
                '实付金额': 'sum',
                '退款金额': 'sum'
            }).reset_index()
            hourly_stats.columns = ['小时', '订单数', '支付金额', '退款金额']
            hourly_stats['退款率'] = hourly_stats['退款金额'] / hourly_stats['支付金额'].replace(0, 1)
            hourly_stats['订单占比'] = hourly_stats['订单数'] / hourly_stats['订单数'].sum() * 100
            
            # 确保所有小时都有数据
            all_hours = pd.DataFrame({'小时': range(24)})
            hourly_stats = all_hours.merge(hourly_stats, on='小时', how='left').fillna(0)
            hourly_stats['订单占比'] = hourly_stats['订单数'] / hourly_stats['订单数'].sum() * 100
            hourly_stats = hourly_stats.sort_values('小时')
            
            # 划分时段建议
            peak_hours = hourly_stats.nlargest(3, '订单数')['小时'].tolist()
            low_hours = hourly_stats.nsmallest(6, '订单数')['小时'].tolist()
            
            # 广告折扣建议
            hourly_stats['广告折扣建议'] = hourly_stats['小时'].apply(
                lambda h: '🔥 满折扣(100%)' if h in peak_hours else 
                         ('⚡ 高折扣(80%)' if hourly_stats[hourly_stats['小时']==h]['订单占比'].values[0] > 5 else 
                         ('💤 低折扣(50%)' if hourly_stats[hourly_stats['小时']==h]['订单占比'].values[0] < 2 else 
                         ('⚡ 正常折扣(70%)' if h in low_hours else '⚡ 正常折扣(70%)')))
            )
            
            col_ad1, col_ad2 = st.columns([1.2, 1])
            
            with col_ad1:
                fig_hourly = px.bar(hourly_stats, x='小时', y='订单数',
                                   title='24小时订单分布',
                                   color='支付金额', color_continuous_scale='Viridis')
                fig_hourly.update_xaxes(tickvals=list(range(24)))
                st.plotly_chart(fig_hourly, use_container_width=True)
            
            with col_ad2:
                st.markdown("**📊 广告分时折扣建议表：**")
                st.dataframe(hourly_stats[['小时', '订单数', '订单占比', '广告折扣建议']]
                            .style.format({'订单占比': '{:.1f}%'})
                            .background_gradient(subset=['订单数'], cmap='Blues'),
                            use_container_width=True, height=400)
                
                st.markdown("""
                **💡 商家赋能建议**：
                - 🔥 **高峰时段**（{peak}点）：广告预算拉满，争取最大曝光
                - ⚡ **正常时段**：维持基础投放，关注ROI
                - 💤 **低谷时段**（凌晨等）：广告暂停或降至最低折扣，节省预算
                """.format(peak=', '.join(map(str, peak_hours))))
                
                # 导出建议表
                csv_ad = hourly_stats[['小时', '订单数', '订单占比', '广告折扣建议']].to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 导出广告折扣建议表", csv_ad, "广告分时折扣建议.csv", "text/csv")

        # ============================================================
        # 客单价拉升诊断
        # ============================================================
        st.markdown("---")
        st.markdown("### 💰 客单价拉升诊断")
        st.markdown("> 分析实际支付订单的金额分布，提供客单价提升策略")

        # 金额分布分析
        df_paid = df_tmall_clean[df_tmall_clean['实付金额'] > 0]
        if len(df_paid) > 0:
            # 金额分段
            df_paid['金额段'] = df_paid['实付金额'].apply(
                lambda x: '0-50元' if x < 50 else 
                         ('50-100元' if x < 100 else 
                         ('100-200元' if x < 200 else 
                         ('200-500元' if x < 500 else '500元以上')))
            )
            
            amount_dist = df_paid.groupby('金额段').agg({
                '订单编号': 'count',
                '实付金额': ['sum', 'mean']
            }).reset_index()
            amount_dist.columns = ['金额段', '订单数', '总金额', '平均金额']
            amount_dist['订单占比'] = amount_dist['订单数'] / amount_dist['订单数'].sum() * 100
            amount_dist['金额占比'] = amount_dist['总金额'] / amount_dist['总金额'].sum() * 100
            
            # 计算整体客单价
            avg_order_value = df_paid['实付金额'].mean()
            
            col_aov1, col_aov2 = st.columns([1, 1])
            
            with col_aov1:
                fig_aov = px.pie(amount_dist, values='订单数', names='金额段',
                               title='订单金额分布',
                               hole=0.4,
                               color_discrete_sequence=px.colors.sequential.Blues_r)
                st.plotly_chart(fig_aov, use_container_width=True)
            
            with col_aov2:
                st.metric("整体客单价(AOV)", f"¥{avg_order_value:.0f}")
                
                # 低客单价诊断
                low_orders_pct = amount_dist[amount_dist['金额段'] == '0-50元']['订单占比'].values[0] if len(amount_dist[amount_dist['金额段'] == '0-50元']) > 0 else 0
                
                if low_orders_pct > 50:
                    st.error(f"⚠️ {low_orders_pct:.0f}%订单集中在50元以下，客单价过低")
                    st.markdown("""
                    **💡 商家赋能建议**：
                    1. 在店铺首页推「加¥39换购」策略
                    2. 设置「第二件半价」或「满¥99减¥10」门槛
                    3. 推出套装组合（如洁面+爽肤水套装）
                    4. 详情页引导用户升级规格（如大容量版）
                    """)
                elif avg_order_value < 100:
                    st.warning(f"客单价¥{avg_order_value:.0f}偏低，建议提升")
                    st.markdown("""
                    **可执行建议**：
                    - 优化商品定价策略，适当提高基础款价格
                    - 推出更多中高端产品线
                    - 设置满减门槛（如满¥199减¥30）
                    """)
                else:
                    st.success(f"✅ 客单价¥{avg_order_value:.0f}处于健康水平")
                
                st.dataframe(amount_dist.style
                            .format({'总金额': '¥{:,.0f}', '平均金额': '¥{:.0f}', 
                                    '订单占比': '{:.1f}%', '金额占比': '{:.1f}%'})
                            .background_gradient(subset=['订单占比'], cmap='Oranges'),
                            use_container_width=True)

        # ============================================================
        # 弃单挽回概率预测
        # ============================================================
        st.markdown("---")
        st.markdown("### 📭 弃单挽回概率预测")
        st.markdown("> 定位\"下单未付款\"的高发时段和金额区间，提供催付优先级列表")

        # 筛选未付款订单
        df_unpaid = df_tmall_clean[df_tmall_clean['实付金额'] == 0]
        
        if len(df_unpaid) > 0:
            st.metric("弃单总数", len(df_unpaid), f"{len(df_unpaid)/len(df_tmall_clean)*100:.1f}%")
            
            # 弃单时段分析
            unpaid_hourly = df_unpaid.groupby(df_unpaid['创建时间'].dt.hour).size().reset_index(name='弃单数')
            unpaid_hourly.columns = ['小时', '弃单数']
            
            # 弃单金额区间分析
            df_unpaid['金额段'] = df_unpaid['总金额'].apply(
                lambda x: '0-50元' if x < 50 else 
                         ('50-100元' if x < 100 else 
                         ('100-200元' if x < 200 else 
                         ('200-500元' if x < 500 else '500元以上')))
            )
            unpaid_amount = df_unpaid.groupby('金额段').agg({
                '订单编号': 'count',
                '总金额': 'sum'
            }).reset_index()
            unpaid_amount.columns = ['金额段', '弃单数', '潜在损失']
            unpaid_amount = unpaid_amount.sort_values('潜在损失', ascending=False)
            
            col_unp1, col_unp2 = st.columns([1, 1])
            
            with col_unp1:
                fig_unp_hour = px.bar(unpaid_hourly, x='小时', y='弃单数',
                                     title='弃单时段分布',
                                     color='弃单数', color_continuous_scale='Reds')
                fig_unp_hour.update_xaxes(tickvals=list(range(24)))
                st.plotly_chart(fig_unp_hour, use_container_width=True)
            
            with col_unp2:
                fig_unp_amt = px.bar(unpaid_amount, x='金额段', y='弃单数',
                                     title='弃单金额区间分布',
                                     color='潜在损失', color_continuous_scale='Oranges')
                st.plotly_chart(fig_unp_amt, use_container_width=True)
            
            # 催付优先级建议
            st.markdown("#### 📋 催付优先级列表")
            
            # 高价值弃单
            high_value_unpaid = df_unpaid[df_unpaid['总金额'] > 100].nlargest(20, '总金额')
            if len(high_value_unpaid) > 0:
                st.markdown("**高价值弃单（金额>¥100）建议优先催付：**")
                st.dataframe(high_value_unpaid[['订单编号', '总金额', '创建时间', '收货地址']]
                            .sort_values('总金额', ascending=False)
                            .style.format({'总金额': '¥{:.0f}'}),
                            use_container_width=True, height=300)
                
                st.markdown("""
                **💡 商家赋能建议**：
                1. 高金额弃单优先安排客服催付（如¥200以上）
                2. 配置「限时5元催付券」发送给弃单用户
                3. 高发时段（如20-22点）配置自动催付短信
                4. 催付黄金窗口：下单后30分钟内效果最佳
                """)
                
                # 计算挽回潜力
                potential_recovery = unpaid_amount['潜在损失'].sum() * 0.3  # 假设30%挽回率
                st.info(f"💰 若挽回30%弃单，可增收约 ¥{potential_recovery:,.0f}")
            
        else:
            st.success("✅ 无弃单数据，支付转化率良好")

    # ============================================================
    # TAB 5: 用户画像
    # ============================================================
    with tab5:
        st.title("👤 用户画像与客户价值分析")
        st.markdown("> 通过 RFM 模型量化客户价值，指导精准营销策略")
        
        # ============================================================
        # RFM 评分权重设置
        # ============================================================
        st.markdown("### 📊 RFM 评分模型配置")
        st.markdown("""
        > **评分公式**：客户价值总分 = R×20% + F×30% + M×50%
        
        | 维度 | 含义 | 权重 | 评分逻辑 |
        |------|------|------|----------|
        | **R-Recency** | 最近购买时间 | 20% | 越近得分越高（反向排名） |
        | **F-Frequency** | 购买频次 | 30% | 频次越高得分越高 |
        | **M-Monetary** | 消费金额 | 50% | 金额越高得分越高 |
        """)
        
        # 固定权重
        r_weight, f_weight, m_weight = 20, 30, 50
        
        # ============================================================
        # RFM 评分计算
        # ============================================================
        st.markdown("---")
        st.markdown("### 🎯 客户 RFM 价值评分")
        
        # 重新计算带权重的 RFM 评分
        if len(rfm_df) > 0 and '客户编码' in rfm_df.columns:
            # 获取 RFM 原始值
            rfm_score_df = rfm_df.copy()
            
            # 计算各维度的百分制得分（基于排名）
            if 'Recency(天)' in rfm_score_df.columns:
                # Recency: 越小越好，反向排名
                rfm_score_df['R_百分制'] = 100 - (rfm_score_df['Recency(天)'].rank(pct=True) * 100)
            else:
                rfm_score_df['R_百分制'] = 50
            
            if 'Frequency' in rfm_score_df.columns:
                # Frequency: 越大越好，正向排名
                rfm_score_df['F_百分制'] = rfm_score_df['Frequency'].rank(pct=True) * 100
            else:
                rfm_score_df['F_百分制'] = 50
            
            if 'Monetary(¥)' in rfm_score_df.columns:
                # Monetary: 越大越好，正向排名
                rfm_score_df['M_百分制'] = rfm_score_df['Monetary(¥)'].rank(pct=True) * 100
            else:
                rfm_score_df['M_百分制'] = 50
            
            # 计算加权总分
            rfm_score_df['加权总分'] = (
                rfm_score_df['R_百分制'] * r_weight / 100 +
                rfm_score_df['F_百分制'] * f_weight / 100 +
                rfm_score_df['M_百分制'] * m_weight / 100
            )
            
            # 客户标签划分
            def assign_customer_label(score):
                if score >= 80:
                    return '优质客户'
                elif score >= 60:
                    return '潜力客户'
                elif score >= 40:
                    return '普通客户'
                elif score >= 20:
                    return '待激活客户'
                else:
                    return '流失风险'
            
            rfm_score_df['客户标签'] = rfm_score_df['加权总分'].apply(assign_customer_label)
            
            # 统计各标签客户数
            label_stats = rfm_score_df['客户标签'].value_counts().reset_index()
            label_stats.columns = ['客户标签', '客户数']
            label_stats['占比'] = label_stats['客户数'] / label_stats['客户数'].sum() * 100
            
            # 展示评分结果
            col_score1, col_score2, col_score3 = st.columns([1.5, 1, 1])
            
            with col_score1:
                # 客户标签分布
                fig_label = px.bar(
                    label_stats,
                    x='客户标签',
                    y='客户数',
                    color='客户标签',
                    color_discrete_map={
                        '优质客户': '#2ecc71',
                        '潜力客户': '#3498db',
                        '普通客户': '#f39c12',
                        '待激活客户': '#e67e22',
                        '流失风险': '#e74c3c'
                    },
                    title='客户标签分布'
                )
                fig_label.update_traces(text=label_stats['占比'].apply(lambda x: f'{x:.1f}%'), textposition='outside')
                st.plotly_chart(fig_label, use_container_width=True)
            
            with col_score2:
                # 优质客户统计
                premium_count = len(rfm_score_df[rfm_score_df['客户标签'] == '优质客户'])
                premium_monetary = rfm_score_df[rfm_score_df['客户标签'] == '优质客户']['Monetary(¥)'].sum()
                st.metric("🏆 优质客户数", f"{premium_count}", f"占比 {premium_count/len(rfm_score_df)*100:.1f}%")
                st.metric("💰 优质客户贡献", f"¥{premium_monetary/1e4:.1f}万", 
                         f"贡献率 {premium_monetary/rfm_score_df['Monetary(¥)'].sum()*100:.1f}%")
            
            with col_score3:
                # 流失风险统计
                risk_count = len(rfm_score_df[rfm_score_df['客户标签'] == '流失风险'])
                risk_monetary = rfm_score_df[rfm_score_df['客户标签'] == '流失风险']['Monetary(¥)'].sum()
                st.metric("⚠️ 流失风险数", f"{risk_count}", f"占比 {risk_count/len(rfm_score_df)*100:.1f}%")
                st.metric("💸 流失潜在损失", f"¥{risk_monetary/1e4:.1f}万")
            
            # 客户详情列表
            st.markdown("---")
            st.markdown("### 📋 客户价值明细")
            
            # 筛选器
            filter_label = st.selectbox(
                "筛选客户标签",
                ['全部', '优质客户', '潜力客户', '普通客户', '待激活客户', '流失风险'],
                key="filter_label"
            )
            
            if filter_label != '全部':
                display_df = rfm_score_df[rfm_score_df['客户标签'] == filter_label]
            else:
                display_df = rfm_score_df
            
            # 排序
            display_df = display_df.sort_values('加权总分', ascending=False)
            
            # 展示表格
            st.dataframe(
                display_df[['客户编码', 'Recency(天)', 'Frequency', 'Monetary(¥)', 
                           'R_百分制', 'F_百分制', 'M_百分制', '加权总分', '客户标签']]
                .style
                .format({
                    'Recency(天)': '{:.0f}',
                    'Frequency': '{:.0f}',
                    'Monetary(¥)': '¥{:,.0f}',
                    'R_百分制': '{:.1f}',
                    'F_百分制': '{:.1f}',
                    'M_百分制': '{:.1f}',
                    '加权总分': '{:.1f}'
                })
                .background_gradient(subset=['加权总分'], cmap='RdYlGn')
                .map(lambda x: 'color: #2ecc71; font-weight: bold' if x == '优质客户' else 
                                ('color: #e74c3c' if x == '流失风险' else ''), subset=['客户标签']),
                use_container_width=True,
                hide_index=True,
                height=400
            )
            
            # 导出功能
            if st.button("📥 导出客户评分结果 (CSV)", key="export_rfm_score"):
                csv = display_df.to_csv(index=False).encode('utf-8-sig')
                st.download_button("下载CSV", csv, "客户RFM评分.csv", "text/csv")
            
            # ============================================================
            # RFM 三维分布图
            # ============================================================
            st.markdown("---")
            st.markdown("### 🌐 RFM 三维分布图")
            
            col_3d1, col_3d2 = st.columns([1.5, 1])
            
            with col_3d1:
                # 3D 散点图（带标签颜色）
                fig_rfm_3d = px.scatter_3d(
                    rfm_score_df,
                    x='Recency(天)',
                    y='Frequency',
                    z='Monetary(¥)',
                    color='客户标签',
                    color_discrete_map={
                        '优质客户': '#2ecc71',
                        '潜力客户': '#3498db',
                        '普通客户': '#f39c12',
                        '待激活客户': '#e67e22',
                        '流失风险': '#e74c3c'
                    },
                    hover_data=['客户编码', '加权总分'],
                    title='客户RFM三维分布（按标签分类）'
                )
                fig_rfm_3d.update_layout(scene=dict(
                    xaxis_title='R-最近购买天数',
                    yaxis_title='F-购买频次',
                    zaxis_title='M-消费金额'
                ))
                st.plotly_chart(fig_rfm_3d, use_container_width=True)
            
            with col_3d2:
                # 饼图：客户标签占比
                fig_pie_label = px.pie(
                    label_stats,
                    values='客户数',
                    names='客户标签',
                    color='客户标签',
                    color_discrete_map={
                        '优质客户': '#2ecc71',
                        '潜力客户': '#3498db',
                        '普通客户': '#f39c12',
                        '待激活客户': '#e67e22',
                        '流失风险': '#e74c3c'
                    },
                    title='客户标签占比',
                    hole=0.4
                )
                fig_pie_label.update_traces(text=label_stats['占比'].apply(lambda x: f'{x:.1f}%'))
                st.plotly_chart(fig_pie_label, use_container_width=True)
            
            # ============================================================
            # 营销策略建议
            # ============================================================
            st.markdown("---")
            st.markdown("### 📌 分层营销策略建议")
            
            strategy_data = [
                {'客户标签': '优质客户', '策略': 'VIP专属服务、新品优先体验、积分加倍', '优先级': 'P0', '资源分配': '40%'},
                {'客户标签': '潜力客户', '策略': '促销定向推送、满减券刺激、会员升级', '优先级': 'P1', '资源分配': '30%'},
                {'客户标签': '普通客户', '策略': '节日问候、常规促销、批量触达', '优先级': 'P2', '资源分配': '15%'},
                {'客户标签': '待激活客户', '策略': '唤醒短信、限时优惠、老客回归礼包', '优先级': 'P3', '资源分配': '10%'},
                {'客户标签': '流失风险', '策略': '深度回访、问题诊断、挽回激励', '优先级': 'P4', '资源分配': '5%'},
            ]
            
            strategy_df = pd.DataFrame(strategy_data)
            st.table(strategy_df)
            
            # ============================================================
            # 高价值流失风险客户预警
            # ============================================================
            st.markdown("---")
            st.markdown("### ⚠️ 高价值流失风险预警")
            
            # 高价值但流失风险的客户
            high_value_risk = rfm_score_df[
                (rfm_score_df['客户标签'] == '流失风险') & 
                (rfm_score_df['Monetary(¥)'] > rfm_score_df['Monetary(¥)'].median())
            ].sort_values('Monetary(¥)', ascending=False)
            
            if len(high_value_risk) > 0:
                st.warning(f"发现 {len(high_value_risk)} 位高价值客户存在流失风险，建议优先挽回")
                
                st.dataframe(
                    high_value_risk[['客户编码', 'Recency(天)', 'Frequency', 'Monetary(¥)', '加权总分']]
                    .style
                    .format({
                        'Recency(天)': '{:.0f}天',
                        'Frequency': '{:.0f}次',
                        'Monetary(¥)': '¥{:,.0f}',
                        '加权总分': '{:.1f}分'
                    })
                    .background_gradient(subset=['Monetary(¥)'], cmap='Reds'),
                    use_container_width=True,
                    hide_index=True
                )
                
                # 挽回建议
                total_loss = high_value_risk['Monetary(¥)'].sum()
                potential_recovery = total_loss * 0.3
                st.info(f"💡 若挽回30%高危客户，预计可挽回 ¥{potential_recovery:,.0f} 销售额")
            else:
                st.success("✅ 无高价值流失风险客户")
            
        else:
            st.warning("⚠️ 无 RFM 数据，请先运行数据预处理")
        
        # ============================================================
        # 客户消费行为分析
        # ============================================================
        st.markdown("---")
        st.markdown("### 📈 客户消费行为分析")
        
        if len(df_sales) > 0:
            # 客户消费频次分布
            freq_dist = df_sales.groupby('客户编码').size().reset_index(name='订单数')
            freq_dist['频次区间'] = freq_dist['订单数'].apply(
                lambda x: '1次' if x == 1 else 
                         ('2-3次' if x <= 3 else 
                         ('4-5次' if x <= 5 else 
                         ('6-10次' if x <= 10 else '10次以上')))
            )
            
            freq_stats = freq_dist.groupby('频次区间').size().reset_index(name='客户数')
            
            col_freq1, col_freq2 = st.columns(2)
            
            with col_freq1:
                fig_freq = px.bar(
                    freq_stats,
                    x='频次区间',
                    y='客户数',
                    title='客户购买频次分布',
                    color='频次区间',
                    color_discrete_sequence=px.colors.sequential.Blues
                )
                st.plotly_chart(fig_freq, use_container_width=True)
            
            with col_freq2:
                # 客户首次购买时间分布
                if '订单日期' in df_sales.columns:
                    first_purchase = df_sales.groupby('客户编码')['订单日期'].min().reset_index()
                    first_purchase.columns = ['客户编码', '首次购买时间']
                    first_purchase['首次购买月份'] = pd.to_datetime(first_purchase['首次购买时间']).dt.to_period('M').astype(str)
                    
                    month_dist = first_purchase.groupby('首次购买月份').size().reset_index(name='新客户数')
                    
                    fig_new = px.line(
                        month_dist,
                        x='首次购买月份',
                        y='新客户数',
                        title='新客户增长趋势',
                        markers=True
                    )
                    st.plotly_chart(fig_new, use_container_width=True)
            
            # 客户平均消费金额分布
            avg_amount = df_sales.groupby('客户编码')['金额'].mean().reset_index()
            avg_amount.columns = ['客户编码', '平均消费']
            avg_amount['金额区间'] = avg_amount['平均消费'].apply(
                lambda x: '0-100' if x <= 100 else 
                         ('100-500' if x <= 500 else 
                         ('500-1000' if x <= 1000 else 
                         ('1000-5000' if x <= 5000 else '5000以上')))
            )
            
            amount_dist = avg_amount.groupby('金额区间').size().reset_index(name='客户数')
            
            fig_amount = px.pie(
                amount_dist,
                values='客户数',
                names='金额区间',
                title='客户平均消费金额分布',
                hole=0.4
            )
            st.plotly_chart(fig_amount, use_container_width=True)

    # ============================================================
    # TAB 6: 联动洞察
    # ============================================================
    with tab6:
        st.title("🔗 跨数据联动洞察")
        st.markdown("> 将三份数据拼接成**品牌→渠道→零售**产业链全景图")

        # Framework diagram
        st.markdown("""
        ### 🧩 分析框架
        ```
        ┌──────────────────────────────────────────────────────────┐
        │                     产业链分析引擎                          │
        │                                                          │
        │  双十一美妆数据         日化渠道数据         天猫零售数据     │
        │  (品牌市场竞争)    →    (经销商健康度)    →   (终端效率)     │
        │                                                          │
        │  • 品牌GMV排名           • RFM分层           • 退款归因    │
        │  • 价格带策略            • 区域渗透率         • 支付行为    │
        │  • 竞品对标              • 月度趋势           • 风险预测    │
        │                                                          │
        │              └──────── 跨数据联动 ─────────┘               │
        │                                                          │
        │  🎯 面膜品类链路：品牌卖多少 → 经销商铺多少 → 终端退多少        │
        │  🎯 区域穿透力：品类热力图 × 经销商密度 × 零售退货率            │
        └──────────────────────────────────────────────────────────┘
        ```
        """)

        st.markdown("---")

        # ============================================================
        # Insight 1: Category Chain Analysis
        # ============================================================
        st.subheader("🔍 洞察一：品类全链路透视（以面膜为例）")

        # Beauty: find mask-related brands
        mask_keywords = ['面膜', '面贴膜', '眼膜', '面贴']
        df_beauty['is_mask'] = df_beauty['title'].apply(
            lambda x: any(kw in str(x) for kw in mask_keywords)
        )
        mask_beauty = df_beauty[df_beauty['is_mask']]

        # Rihua: mask products
        mask_products = df_product[df_product['商品小类'] == '面膜']
        mask_codes = mask_products['商品编号'].tolist()
        mask_sales = df_sales[df_sales['商品编号'].isin(mask_codes)]

        col_i1, col_i2, col_i3 = st.columns(3)
        with col_i1:
            st.metric("🏷️ 双十一面膜SKU", f"{len(mask_beauty):,}", f"占总SKU {len(mask_beauty)/len(df_beauty)*100:.1f}%")
            st.metric("💰 面膜GMV估算", f"¥{mask_beauty['gmv'].sum()/1e8:.1f}亿")

        with col_i2:
            st.metric("📦 经销商面膜SKU", f"{len(mask_codes)}", f"占总商品 {len(mask_codes)/len(df_product)*100:.0f}%")
            st.metric("💵 渠道面膜销售额", f"¥{mask_sales['金额'].sum()/1e8:.1f}亿")

        with col_i3:
            # Compare ratio
            beauty_gmv = mask_beauty['gmv'].sum()
            channel_gmv = mask_sales['金额'].sum()
            ratio = channel_gmv / beauty_gmv * 100 if beauty_gmv > 0 else 0
            st.metric("📊 渠道/品牌比", f"{ratio:.0f}%",
                     "渠道>品牌：经销商压货" if ratio > 100 else "品牌>渠道：直营为主")

        st.markdown("---")

        # ============================================================
        # Insight 2: Regional Cross-Analysis
        # ============================================================
        st.subheader("🗺️ 洞察二：区域穿透力矩阵")

        # Province overlap between tmall addresses and rihua sales
        tmall_addrs = set(df_tmall['收货地址'].unique())
        rihua_provs = set(df_sales['所在省份'].unique())

        common = tmall_addrs & rihua_provs
        only_tmall = tmall_addrs - rihua_provs
        only_rihua = rihua_provs - tmall_addrs

        cc1, cc2, cc3 = st.columns(3)
        cc1.metric("共同覆盖", f"{len(common)}")
        cc2.metric("仅零售覆盖", f"{len(only_tmall)}")
        cc3.metric("仅渠道覆盖", f"{len(only_rihua)}", "→ 渠道空白市场")

        # Combined region analysis
        st.markdown("#### 省份级别：渠道销售额 vs 零售订单数")
        tmall_prov = df_tmall.groupby('收货地址').agg(
            零售订单=('订单编号', 'count'),
            零售退款率=('退款金额', lambda x: x.sum() / max(df_tmall.loc[x.index, '实付金额'].sum(), 1) * 100),
        ).reset_index()
        tmall_prov.columns = ['省份', '零售订单', '零售退款率']

        rihua_prov_agg = df_sales.groupby('所在省份').agg(
            渠道销售额=('金额', 'sum'),
            渠道客户=('客户编码', 'nunique'),
        ).reset_index()
        rihua_prov_agg.columns = ['省份', '渠道销售额', '渠道客户']

        merged_prov = pd.merge(rihua_prov_agg, tmall_prov, on='省份', how='outer').fillna(0)
        if len(merged_prov) > 0:
            fig_cross = px.scatter(
                merged_prov,
                x='渠道销售额', y='零售订单',
                size='渠道客户',
                color='零售退款率',
                hover_name='省份',
                title='省份交叉矩阵：渠道实力 × 零售规模 × 退款率',
                labels={'渠道销售额': '渠道销售额(¥)', '零售订单': '零售订单数', '零售退款率': '零售退款率(%)'},
                color_continuous_scale='RdYlGn_r',
                height=500,
            )
            fig_cross.update_layout(legend=dict(orientation='h', y=-0.15))
            st.plotly_chart(fig_cross, use_container_width=True)

        st.markdown("---")

        # ============================================================
        # Insight 3: Key Findings Summary
        # ============================================================
        st.subheader("📌 核心发现汇总")

        f1, f2, f3 = st.columns(3)
        with f1:
            st.info("""
            **发现一：国货品牌完成大众价位带封锁**
            - ¥100-200价位带占SKU的23.2%
            - 相宜本草以均价¥123做到¥61亿GMV
            - 国际品牌被困高端价位，市占率受限
            """)
        with f2:
            st.warning("""
            **发现二：天猫零售存在严重的"冲动消费"问题**
            - 退款率30.1%（行业正常水平约5-10%）
            - 32.3%的订单从未付款
            - 高价段(¥500+)退款率显著偏高
            """)
        with f3:
            st.success("""
            **发现三：渠道有明确的区域扩张机会**
            - 东区占36%份额，过度集中
            - 西区12省仅占13%份额，但人口基数大
            - 核心客户仅占客户群的少数，尾部客户活跃度不足
            """)

        f4, f5, f6 = st.columns(3)
        with f4:
            st.info("""
            **发现四：面膜是"引流-退货"双刃剑**
            - 面膜是SKU最多的品类（16个）
            - 双十一面膜GMV占比极高
            - 但也可能是高退货率品类
            """)
        with f5:
            st.warning("""
            **发现五：江苏省是"完美市场"**
            - 渠道销售额¥6.7亿（第一）
            - 零售订单数2,126单（第三）
            - 渠道-零售协同效应最优
            """)
        with f6:
            st.success("""
            **发现六：系统化平台的价值**
            - 可替换数据源（换一批Excel即可复用）
            - 交互式探索（比静态报告发现更多问题）
            - 模型可迭代（持续优化RFM/退款预测）
            """)

        st.markdown("---")

        # ============================================================
        # Overall Summary
        # ============================================================
        st.subheader("📝 分析总结")

        st.markdown("""
        ### 一句话结论

        > **国货美妆品牌在大众市场已建立压倒性优势，但渠道效率和终端零售体验仍是瓶颈——**
        > **30%的退款率意味着每三个客户就有一个不满意，这是我们面临的最大挑战。**

        ### 行动建议

        | 优先级 | 领域 | 建议 | 预期影响 |
        |--------|------|------|----------|
        | 🔴 P0 | 零售退款 | 建立退款预警模型，对高风险订单主动介入（优惠券/客服） | 降低退款率5-10个百分点 |
        | 🟡 P1 | 渠道扩张 | 西区（四川、陕西为核心）新增经销商招募，复制江苏模式 | 增长15-20% |
        | 🟡 P1 | 品类优化 | 减少同质化面膜SKU，增加差异化产品（防晒、精华） | 提升客单价和复购 |
        | 🟢 P2 | 客户运营 | 对RFM"流失风险"客户启动召回计划 | 挽回10%高危客户 |

        ---
        *本分析报告由「日化美妆产业链分析平台」自动生成 | L4 系统化数据分析项目*
        """)

        # Export full report
        st.download_button(
            "📥 导出完整分析报告 (CSV数据包)",
            data=df_sales.head(1000).to_csv(index=False).encode('utf-8-sig'),
            file_name="analysis_export_sample.csv",
            mime="text/csv",
        )

    # ============================================================
    # TAB 7: 销售策略（动态调整仓库存储 - 使用ML模型）
    # ============================================================
    with tab7:
        st.title("📊 销售策略 - 智能库存调整")
        st.markdown("> 使用机器学习模型预测热度等级和销量，动态调整仓库存储")
        
        # 加载库存数据
        inventory_path = os.path.join(base_dir, 'inventory_stock.csv')
        
        if os.path.exists(inventory_path):
            df_inventory = pd.read_csv(inventory_path)
            
            # 数据预处理
            df_inventory['当前库存量'] = pd.to_numeric(df_inventory['当前库存量'], errors='coerce').fillna(0)
            df_inventory['历史总销量'] = pd.to_numeric(df_inventory['历史总销量'], errors='coerce').fillna(0)
            df_inventory['订单数'] = pd.to_numeric(df_inventory['订单数'], errors='coerce').fillna(0)
            df_inventory['最近补货日期'] = pd.to_datetime(df_inventory['最近补货日期'], errors='coerce')
            
            # ===== 加载训练好的模型 =====
            model_dir = os.path.join(base_dir, 'models')
            hotness_model_path = os.path.join(model_dir, 'product_hotness_best.joblib')
            hotness_encoder_path = os.path.join(model_dir, 'product_hotness_encoders.joblib')
            sales_model_path = os.path.join(model_dir, 'sales_forecast_best_optimized.joblib')
            sales_encoder_path = os.path.join(model_dir, 'sales_forecast_encoders_optimized.joblib')
            
            models_loaded = False
            
            if os.path.exists(hotness_model_path) and os.path.exists(hotness_encoder_path):
                try:
                    hotness_model = joblib.load(hotness_model_path)
                    hotness_encoders = joblib.load(hotness_encoder_path)
                    models_loaded = True
                    st.success("✅ 商品热度模型已加载")
                except Exception as e:
                    st.warning(f"⚠️ 热度模型加载失败: {e}")
            
            if os.path.exists(sales_model_path) and os.path.exists(sales_encoder_path):
                try:
                    sales_model = joblib.load(sales_model_path)
                    sales_encoders = joblib.load(sales_encoder_path)
                    st.success("✅ 销量预测模型已加载")
                except Exception as e:
                    st.warning(f"⚠️ 销量模型加载失败: {e}")
            
            # ===== 1. 商品热度预测（ML模型） =====
            st.markdown("### 1️⃣ 商品热度预测")
            
            if models_loaded:
                st.markdown("**使用训练好的机器学习模型预测商品热度等级**")
                
                # 构建特征（适配模型需要的特征）
                # 计算库存特征
                df_inventory['平均销量'] = df_inventory['历史总销量'] / (df_inventory['订单数'] + 1)
                df_inventory['总销售额'] = df_inventory['历史总销量'] * 50  # 估算平均单价
                df_inventory['平均销售额'] = df_inventory['总销售额'] / (df_inventory['订单数'] + 1)
                df_inventory['覆盖省份'] = 5  # 默认值
                df_inventory['客户数'] = df_inventory['订单数']  # 假设每个订单一个客户
                df_inventory['近30天销量'] = df_inventory['历史总销量'] / 12  # 估算
                df_inventory['近60天销量'] = df_inventory['历史总销量'] / 6
                df_inventory['近90天销量'] = df_inventory['历史总销量'] / 4
                df_inventory['距今天数'] = (pd.Timestamp.now() - df_inventory['最近补货日期']).dt.days.fillna(30)
                df_inventory['销量趋势'] = df_inventory['近30天销量'] / df_inventory['近90天销量'].replace(0, 1)
                df_inventory['销量趋势'] = df_inventory['销量趋势'].replace([np.inf, -np.inf], 0)
                
                # 编码类别特征
                try:
                    le_category = hotness_encoders.get('category_encoder')
                    le_subcategory = hotness_encoders.get('subcategory_encoder')
                    
                    if le_category:
                        df_inventory['大类编码'] = df_inventory['商品大类'].apply(
                            lambda x: le_category.transform([str(x)])[0] if str(x) in le_category.classes_ else 0
                        )
                    else:
                        df_inventory['大类编码'] = 0
                    
                    if le_subcategory:
                        df_inventory['小类编码'] = df_inventory['商品小类'].apply(
                            lambda x: le_subcategory.transform([str(x)])[0] if str(x) in le_subcategory.classes_ else 0
                        )
                    else:
                        df_inventory['小类编码'] = 0
                    
                    # 构建特征矩阵
                    hotness_features = ['总销量', '平均销量', '订单次数', '总销售额', '平均销售额',
                                        '覆盖省份', '客户数', '近30天销量', '近60天销量', '近90天销量',
                                        '距今天数', '销量趋势', '大类编码', '小类编码']
                    
                    # 映射特征名
                    feature_mapping = {
                        '总销量': '历史总销量',
                        '订单次数': '订单数',
                        '覆盖省份': '覆盖省份',
                        '客户数': '客户数',
                        '距今天数': '距今天数'
                    }
                    
                    X_hotness = pd.DataFrame()
                    for feat in hotness_features:
                        if feat in feature_mapping:
                            X_hotness[feat] = df_inventory[feature_mapping[feat]]
                        elif feat in df_inventory.columns:
                            X_hotness[feat] = df_inventory[feat]
                        else:
                            X_hotness[feat] = 0
                    
                    X_hotness = X_hotness.fillna(0).values
                    
                    # 使用模型预测
                    hotness_pred = hotness_model.predict(X_hotness)
                    
                    # 解码预测结果
                    le_hotness = hotness_encoders.get('hotness_encoder')
                    if le_hotness:
                        hotness_labels = le_hotness.inverse_transform(hotness_pred)
                    else:
                        # 默认映射
                        hotness_map = {0: '呆滞', 1: '滞销预警', 2: '正常', 3: '热销'}
                        hotness_labels = [hotness_map.get(p, '正常') for p in hotness_pred]
                    
                    df_inventory['热度等级_ML'] = hotness_labels
                    
                except Exception as e:
                    st.warning(f"⚠️ 模型预测失败，使用规则判断: {e}")
                    # 使用规则判断作为备选
                    total_sales_q75 = df_inventory['历史总销量'].quantile(0.75)
                    total_sales_q50 = df_inventory['历史总销量'].quantile(0.50)
                    order_q75 = df_inventory['订单数'].quantile(0.75)
                    
                    def assign_hotness(row):
                        if row['历史总销量'] >= total_sales_q75 and row['订单数'] >= order_q75:
                            return '热销'
                        elif row['历史总销量'] >= total_sales_q50:
                            return '正常'
                        elif row['历史总销量'] < total_sales_q50 * 0.5:
                            return '滞销预警'
                        else:
                            return '呆滞'
                    
                    df_inventory['热度等级_ML'] = df_inventory.apply(assign_hotness, axis=1)
                
                # 添加emoji
                hotness_emoji = {'热销': '🔥 热销', '正常': '✅ 正常', '滞销预警': '⚠️ 滞销预警', '呆滞': '❌ 呆滞'}
                df_inventory['热度等级'] = df_inventory['热度等级_ML'].map(hotness_emoji)
                
            else:
                st.warning("⚠️ 模型未加载，使用规则判断")
                
                # 使用规则判断
                total_sales_q75 = df_inventory['历史总销量'].quantile(0.75)
                total_sales_q50 = df_inventory['历史总销量'].quantile(0.50)
                order_q75 = df_inventory['订单数'].quantile(0.75)
                
                def assign_hotness(row):
                    if row['历史总销量'] >= total_sales_q75 and row['订单数'] >= order_q75:
                        return '🔥 热销'
                    elif row['历史总销量'] >= total_sales_q50:
                        return '✅ 正常'
                    elif row['历史总销量'] < total_sales_q50 * 0.5:
                        return '⚠️ 滞销预警'
                    else:
                        return '❌ 呆滞'
                
                df_inventory['热度等级'] = df_inventory.apply(assign_hotness, axis=1)
            
            # 热度分布统计
            col1, col2 = st.columns(2)
            
            with col1:
                hotness_counts = df_inventory['热度等级'].value_counts()
                fig_hotness = px.pie(
                    values=hotness_counts.values,
                    names=hotness_counts.index,
                    title='商品热度分布（ML模型预测）',
                    color=hotness_counts.index,
                    color_discrete_map={
                        '🔥 热销': '#2ecc71',
                        '✅ 正常': '#3498db',
                        '⚠️ 滞销预警': '#f39c12',
                        '❌ 呆滞': '#e74c3c'
                    }
                )
                st.plotly_chart(fig_hotness, use_container_width=True)
            
            with col2:
                category_hotness = df_inventory.groupby(['商品大类', '热度等级']).size().reset_index(name='数量')
                fig_cat_hotness = px.bar(
                    category_hotness,
                    x='商品大类',
                    y='数量',
                    color='热度等级',
                    title='各品类热度分布',
                    color_discrete_map={
                        '🔥 热销': '#2ecc71',
                        '✅ 正常': '#3498db',
                        '⚠️ 滞销预警': '#f39c12',
                        '❌ 呆滞': '#e74c3c'
                    }
                )
                st.plotly_chart(fig_cat_hotness, use_container_width=True)
            
            # ===== 2. 销量预测（ML模型） =====
            st.markdown("### 2️⃣ 销量预测")
            
            # 使用简单估算（因为库存数据特征不足以支持销量预测模型）
            st.markdown("**基于历史销量预测未来30天销量**")
            
            # 简化预测：历史日均销量 × 30天
            df_inventory['日均销量'] = df_inventory['历史总销量'] / 365
            df_inventory['预测30天销量'] = df_inventory['日均销量'] * 30
            
            # 安全库存 = 预测销量 × 1.5（安全系数）
            df_inventory['安全库存'] = df_inventory['预测30天销量'] * 1.5
            
            # 补货建议量 = 安全库存 - 当前库存
            df_inventory['补货建议量'] = df_inventory['安全库存'] - df_inventory['当前库存量']
            df_inventory['补货建议量'] = df_inventory['补货建议量'].apply(lambda x: max(0, x))
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("预测30天总销量", f"{df_inventory['预测30天销量'].sum():,.0f}")
            with col2:
                st.metric("平均安全库存", f"{df_inventory['安全库存'].mean():,.0f}")
            with col3:
                need_replenish = df_inventory[df_inventory['补货建议量'] > 0]
                st.metric("需补货商品数", f"{len(need_replenish)}")
            with col4:
                st.metric("总补货量建议", f"{df_inventory['补货建议量'].sum():,.0f}")
            
            # 销量预测表
            st.markdown("**销量预测与补货建议**")
            
            pred_display = df_inventory[['商品编号', '商品名称', '热度等级', '当前库存量',
                                          '日均销量', '预测30天销量', '安全库存', '补货建议量']].copy()
            pred_display = pred_display.sort_values('预测30天销量', ascending=False).head(20)
            
            st.dataframe(
                pred_display.style.background_gradient(subset=['预测30天销量', '补货建议量'], cmap='Blues'),
                use_container_width=True,
                height=400
            )
            
            # ===== 3. 库存状态分析 =====
            st.markdown("### 3️⃣ 库存状态分析")
            
            # 库存周转天数
            df_inventory['库存周转天数'] = df_inventory['当前库存量'] / (df_inventory['日均销量'] + 1)
            
            # 库存健康度分级
            def inventory_health(row):
                turnover = row['库存周转天数']
                if turnover > 180:
                    return '🔴 库存积压'
                elif turnover > 90:
                    return '🟡 库存偏高'
                elif turnover > 30:
                    return '🟢 库存正常'
                else:
                    return '🔵 库存偏低'
            
            df_inventory['库存健康度'] = df_inventory.apply(inventory_health, axis=1)
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("总库存量", f"{df_inventory['当前库存量'].sum():,.0f}")
            with col2:
                avg_turnover = df_inventory['库存周转天数'].mean()
                st.metric("平均周转天数", f"{avg_turnover:.1f}")
            with col3:
                st.metric("商品数", f"{len(df_inventory)}")
            
            health_counts = df_inventory['库存健康度'].value_counts()
            fig_health = px.bar(
                x=health_counts.index,
                y=health_counts.values,
                title='库存健康度分布',
                color=health_counts.index,
                color_discrete_map={
                    '🔴 库存积压': '#e74c3c',
                    '🟡 库存偏高': '#f39c12',
                    '🟢 库存正常': '#2ecc71',
                    '🔵 库存偏低': '#3498db'
                }
            )
            st.plotly_chart(fig_health, use_container_width=True)
            
            # ===== 4. 动态调整策略（热度+库存） =====
            st.markdown("### 4️⃣ 动态库存调整策略")
            
            def adjustment_strategy(row):
                hotness = row['热度等级']
                health = row['库存健康度']
                
                if hotness == '🔥 热销':
                    if health == '🔵 库存偏低':
                        return '⚡ 紧急补货：立即增加50%库存'
                    elif health == '🟢 库存正常':
                        return '📈 预防性补货：建议增加20%库存'
                    else:
                        return '✅ 维持现状：热销商品库存充足'
                
                elif hotness == '✅ 正常':
                    if health == '🔴 库存积压':
                        return '📉 减少补货：暂停补货，消化库存'
                    elif health == '🟡 库存偏高':
                        return '⏸️ 延迟补货：暂缓补货计划'
                    else:
                        return '✅ 维持现状：正常补货节奏'
                
                elif hotness == '⚠️ 滞销预警':
                    if health == '🔴 库存积压':
                        return '🧹 清仓促销：打折清仓，释放库存'
                    elif health == '🟡 库存偏高':
                        return '📉 减少库存：降低库存水位'
                    else:
                        return '⏸️ 观察等待：暂不补货'
                
                else:  # 呆滞
                    return '🚫 建议下架：停止补货，考虑淘汰'
            
            df_inventory['调整策略'] = df_inventory.apply(adjustment_strategy, axis=1)
            
            # 调整策略分布
            strategy_counts = df_inventory['调整策略'].value_counts()
            fig_strategy = px.bar(
                x=strategy_counts.index,
                y=strategy_counts.values,
                title='调整策略分布',
                color=strategy_counts.values,
                color_continuous_scale='RdYlGn'
            )
            st.plotly_chart(fig_strategy, use_container_width=True)
            
            # ===== 5. 调整策略明细 =====
            st.markdown("### 5️⃣ 调整策略明细表")
            
            # 筛选器
            col1, col2, col3 = st.columns(3)
            with col1:
                selected_hotness = st.multiselect(
                    "热度等级",
                    ['🔥 热销', '✅ 正常', '⚠️ 滞销预警', '❌ 呆滞'],
                    default=['🔥 热销', '❌ 呆滞']
                )
            with col2:
                selected_health = st.multiselect(
                    "库存健康度",
                    ['🔴 库存积压', '🟡 库存偏高', '🟢 库存正常', '🔵 库存偏低'],
                    default=['🔴 库存积压', '🔵 库存偏低']
                )
            with col3:
                selected_category = st.multiselect(
                    "商品大类",
                    df_inventory['商品大类'].unique().tolist(),
                    default=df_inventory['商品大类'].unique().tolist()[:2]
                )
            
            # 筛选数据
            filtered_inventory = df_inventory[
                df_inventory['热度等级'].isin(selected_hotness) &
                df_inventory['库存健康度'].isin(selected_health) &
                df_inventory['商品大类'].isin(selected_category)
            ]
            
            st.markdown(f"**共筛选出 {len(filtered_inventory)} 个商品需要调整**")
            
            if len(filtered_inventory) > 0:
                priority_order = {
                    '⚡ 紧急补货：立即增加50%库存': 1,
                    '📈 预防性补货：建议增加20%库存': 2,
                    '🧹 清仓促销：打折清仓，释放库存': 3,
                    '📉 减少库存：降低库存水位': 4,
                    '📉 减少补货：暂停补货，消化库存': 5,
                    '⏸️ 延迟补货：暂缓补货计划': 6,
                    '⏸️ 观察等待：暂不补货': 7,
                    '🚫 建议下架：停止补货，考虑淘汰': 8,
                    '✅ 维持现状：热销商品库存充足': 9,
                    '✅ 维持现状：正常补货节奏': 10
                }
                
                filtered_inventory['优先级'] = filtered_inventory['调整策略'].map(priority_order)
                filtered_inventory = filtered_inventory.sort_values('优先级')
                
                display_cols = ['商品编号', '商品名称', '商品大类', '热度等级', '库存健康度',
                                '当前库存量', '预测30天销量', '补货建议量', '调整策略']
                
                st.dataframe(
                    filtered_inventory[display_cols].style
                    .background_gradient(subset=['预测30天销量', '补货建议量'], cmap='Blues')
                    .map(lambda x: 'color: #e74c3c; font-weight: bold' if '紧急' in str(x) or '清仓' in str(x) else
                                    ('color: #2ecc71' if '补货' in str(x) else ''), subset=['调整策略']),
                    use_container_width=True,
                    height=400
                )
                
                st.download_button(
                    "📥 导出调整清单",
                    data=filtered_inventory[display_cols].to_csv(index=False).encode('utf-8-sig'),
                    file_name="inventory_adjustment_plan.csv",
                    mime="text/csv"
                )
            
            # ===== 6. 品类调整汇总 =====
            st.markdown("### 6️⃣ 品类调整汇总")
            
            category_summary = df_inventory.groupby('商品大类').agg({
                '当前库存量': 'sum',
                '历史总销量': 'sum',
                '预测30天销量': 'sum',
                '补货建议量': 'sum',
                '商品编号': 'count'
            }).reset_index()
            
            category_summary['品类周转天数'] = category_summary['当前库存量'] / (category_summary['历史总销量'] / 365 + 1)
            
            hot_items = df_inventory[df_inventory['热度等级'] == '🔥 热销'].groupby('商品大类').size().reset_index(name='热销商品数')
            category_summary = category_summary.merge(hot_items, on='商品大类', how='left').fillna(0)
            
            dead_items = df_inventory[df_inventory['热度等级'] == '❌ 呆滞'].groupby('商品大类').size().reset_index(name='呆滞商品数')
            category_summary = category_summary.merge(dead_items, on='商品大类', how='left').fillna(0)
            
            def category_strategy(row):
                if row['热销商品数'] > 0 and row['品类周转天数'] < 30:
                    return '🔥 重点补货：热销品类，库存偏低'
                elif row['呆滞商品数'] > 0 and row['品类周转天数'] > 180:
                    return '🧹 品类清仓：呆滞品类，库存积压'
                elif row['热销商品数'] > 0:
                    return '📈 增加备货：热销品类'
                elif row['呆滞商品数'] > 0:
                    return '📉 减少备货：呆滞品类'
                else:
                    return '✅ 维持现状'
            
            category_summary['品类策略'] = category_summary.apply(category_strategy, axis=1)
            
            st.dataframe(category_summary.style.background_gradient(subset=['品类周转天数', '补货建议量'], cmap='RdYlGn_r'),
                        use_container_width=True)
            
            # ===== 7. 执行优先级队列 =====
            st.markdown("### 7️⃣ 执行优先级队列")
            
            st.markdown("""
            **执行优先级建议：**
            
            | 优先级 | 调整类型 | 执行时限 | 负责人建议 |
            |--------|----------|----------|------------|
            | P0 紧急 | 🔥热销+库存偏低 | 立即（24h内） | 采购部 |
            | P1 重要 | ❌呆滞+库存积压 | 本周内 | 销售部（促销） |
            | P2 一般 | ⚠️滞销预警 | 两周内 | 库管部 |
            | P3 观察 | ✅正常商品 | 月度复盘 | 按常规流程 |
            """)
            
            urgent_items = df_inventory[df_inventory['调整策略'].str.contains('紧急|清仓')]
            
            if len(urgent_items) > 0:
                st.markdown(f"**🔴 P0级紧急调整商品 ({len(urgent_items)}个)**")
                urgent_display = urgent_items[['商品编号', '商品名称', '热度等级', '库存健康度', 
                                                '当前库存量', '预测30天销量', '调整策略']].head(10)
                st.dataframe(urgent_display, use_container_width=True)
            else:
                st.success("✅ 当前无P0级紧急调整商品")
            
            # ===== 8. 模型信息 =====
            st.markdown("---")
            st.markdown("### 📊 模型信息")
            
            metrics_path = os.path.join(base_dir, 'model_metrics')
            hotness_metrics_path = os.path.join(metrics_path, 'product_hotness_metrics.json')
            
            if os.path.exists(hotness_metrics_path):
                with open(hotness_metrics_path, 'r', encoding='utf-8') as f:
                    hotness_metrics = json.load(f)
                
                best_model = max(hotness_metrics.keys(), key=lambda k: hotness_metrics[k]['accuracy'])
                st.markdown(f"""
                **热度预测模型：{best_model}**
                - Accuracy: {hotness_metrics[best_model]['accuracy']:.4f}
                - F1 Score: {hotness_metrics[best_model]['f1']:.4f}
                """)
            
        else:
            st.error("⚠️ 未找到库存数据文件 inventory_stock.csv")
            st.markdown("请确保数据文件位于项目根目录下")

    # ============================================================
    # TAB 8: 套餐推荐（基于聚类模型）
    # ============================================================
    with tab8:
        st.title("🎁 套餐推荐 - 智能组合建议")
        st.markdown("> 使用聚类模型发现商品购买关联，自动生成套餐组合建议")
        
        # 加载训练好的模型
        model_dir = os.path.join(base_dir, 'models')
        clustering_model_path = os.path.join(model_dir, 'product_clustering_model.joblib')
        cluster_info_path = os.path.join(base_dir, 'model_metrics', 'product_clusters.json')
        bundle_path = os.path.join(base_dir, 'model_metrics', 'bundle_suggestions.json')
        
        # 1. 显示聚类结果
        st.markdown("### 1️⃣ 商品品类聚类分析")
        
        if os.path.exists(cluster_info_path):
            with open(cluster_info_path, 'r', encoding='utf-8') as f:
                cluster_info = json.load(f)
            
            st.success(f"✅ 已加载聚类结果（共 {len(cluster_info)} 个簇）")
            
            # 聚类概览
            col1, col2, col3, col4 = st.columns(4)
            
            total_cats = sum([v['品类数'] for v in cluster_info.values()])
            total_products = sum([v['商品数'] for v in cluster_info.values()])
            total_sales = sum([v['总销量'] for v in cluster_info.values()])
            
            with col1:
                st.metric("聚类数", len(cluster_info))
            with col2:
                st.metric("覆盖品类", total_cats)
            with col3:
                st.metric("覆盖商品", total_products)
            with col4:
                st.metric("总销量", f"{total_sales:,}")
            
            # 各簇详情
            st.markdown("**各簇品类分布：**")
            
            for cluster_id, info in cluster_info.items():
                with st.expander(f"簇 {cluster_id}：{info['品类数']} 个品类，总销量 {info['总销量']:,}", expanded=False):
                    st.markdown(f"""
                    - **品类列表**: {', '.join(info['品类列表'][:10])}{'...' if len(info['品类列表']) > 10 else ''}
                    - **商品数**: {info['商品数']} 个
                    - **簇内共现强度**: {info['簇内共现强度']:,} 次
                    - **平均单价**: ¥{info['平均单价']}
                    """)
            
            # 聚类可视化
            clustering_chart_path = os.path.join(base_dir, 'model_charts', 'product_clustering.png')
            if os.path.exists(clustering_chart_path):
                st.markdown("**聚类可视化：**")
                st.image(clustering_chart_path, use_container_width=True)
            
        else:
            st.warning("⚠️ 聚类结果文件未找到，请先运行训练脚本")
            st.code("python train/association_rules_model.py", language="python")
        
        # 2. 套餐组合推荐
        st.markdown("### 2️⃣ 套餐组合推荐")
        
        if os.path.exists(bundle_path):
            with open(bundle_path, 'r', encoding='utf-8') as f:
                bundle_suggestions = json.load(f)
            
            st.success(f"✅ 已加载 {len(bundle_suggestions)} 个套餐建议")
            
            # 筛选器
            col1, col2 = st.columns(2)
            with col1:
                min_support = st.slider("最小支持度", 0.0, 0.1, 0.0, 0.01, format="%.2f")
            with col2:
                selected_cluster = st.multiselect(
                    "选择簇",
                    list(range(len(cluster_info))) if cluster_info else [],
                    default=list(range(len(cluster_info))) if cluster_info else []
                )
            
            # 筛选套餐
            filtered_bundles = [
                b for b in bundle_suggestions 
                if b['support'] >= min_support and b['cluster_id'] in selected_cluster
            ]
            
            st.markdown(f"**符合条件的套餐：{len(filtered_bundles)} 个**")
            
            if len(filtered_bundles) > 0:
                # 套餐概览
                total_original = sum([b['original_price'] for b in filtered_bundles])
                total_bundle = sum([b['bundle_price'] for b in filtered_bundles])
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("套餐数量", len(filtered_bundles))
                with col2:
                    st.metric("原价总和", f"¥{total_original:,.0f}")
                with col3:
                    st.metric("套餐价总和", f"¥{total_bundle:,.0f}", delta=f"-¥{total_original-total_bundle:,.0f}")
                
                # 套餐详情表
                st.markdown("**套餐详情：**")
                
                bundle_df = pd.DataFrame(filtered_bundles)
                bundle_df = bundle_df[['bundle_id', 'bundle_name', 'items', 'co_occurrence', 
                                       'original_price', 'bundle_price', 'discount_rate', 'recommendation']]
                bundle_df['节省金额'] = bundle_df['original_price'] - bundle_df['bundle_price']
                
                st.dataframe(
                    bundle_df.style.background_gradient(subset=['co_occurrence', '节省金额'], cmap='Greens'),
                    use_container_width=True,
                    height=400
                )
                
                # 套餐可视化
                st.markdown("**套餐价格对比：**")
                
                # TOP套餐柱状图
                top_bundles = filtered_bundles[:10]
                
                fig = go.Figure()
                
                fig.add_trace(go.Bar(
                    name='原价',
                    x=[b['bundle_name'] for b in top_bundles],
                    y=[b['original_price'] for b in top_bundles],
                    marker_color='#e74c3c'
                ))
                
                fig.add_trace(go.Bar(
                    name='套餐价',
                    x=[b['bundle_name'] for b in top_bundles],
                    y=[b['bundle_price'] for b in top_bundles],
                    marker_color='#2ecc71'
                ))
                
                fig.update_layout(
                    title='TOP10 套餐价格对比',
                    barmode='group',
                    xaxis_tickangle=-45,
                    height=400
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
                # 导出套餐清单
                st.download_button(
                    "📥 导出套餐清单",
                    data=bundle_df.to_csv(index=False).encode('utf-8-sig'),
                    file_name="bundle_suggestions.csv",
                    mime="text/csv"
                )
                
                # 推荐等级分布
                st.markdown("**推荐等级分布：**")
                
                rec_counts = Counter([b['recommendation'] for b in filtered_bundles])
                
                fig_rec = px.pie(
                    values=list(rec_counts.values()),
                    names=list(rec_counts.keys()),
                    title='推荐等级分布',
                    color=list(rec_counts.keys()),
                    color_discrete_map={
                        '强烈推荐': '#2ecc71',
                        '推荐': '#3498db',
                        '常规': '#f39c12'
                    }
                )
                
                st.plotly_chart(fig_rec, use_container_width=True)
                
            else:
                st.info("暂无符合筛选条件的套餐")
        
        else:
            st.warning("⚠️ 套餐建议文件未找到，请先运行训练脚本")
            st.code("python train/association_rules_model.py", language="python")
        
        # 3. 套餐设计策略建议
        st.markdown("### 3️⃣ 套餐设计策略建议")
        
        st.markdown("""
        **套餐设计原则：**
        
        | 推荐等级 | 折扣策略 | 说明 |
        |----------|----------|------|
        | 🔥 强烈推荐 | 15%折扣 | 共现频率高，用户刚需组合 |
        | ✅ 推荐 | 20%折扣 | 有一定共现基础，适合促销 |
        | ⚠️ 常规 | 25%折扣 | 低频共现，尝试性组合 |
        
        **营销场景应用：**
        
        1. **首页推荐**：展示强烈推荐套餐，突出省钱效果
        2. **商品详情页**：显示"经常一起购买"组合
        3. **购物车页**：提示"加购XX元升级套餐"
        4. **结算页**：弹出"搭配购买更划算"弹窗
        5. **双十一大促**：集中推广高共现套餐
        """)
        
        # 4. 共现热力图
        st.markdown("### 4️⃣ 商品共现分析")
        
        heatmap_path = os.path.join(base_dir, 'model_charts', 'co_occurrence_heatmap.png')
        if os.path.exists(heatmap_path):
            st.markdown("**TOP品类共现热力图：**")
            st.image(heatmap_path, use_container_width=True)
        
        # 5. 模型信息
        st.markdown("---")
        st.markdown("### 📊 模型信息")
        
        if os.path.exists(clustering_model_path):
            st.markdown("""
            **聚类模型：K-Means**
            - 算法：基于购买共现模式的商品聚类
            - 特征：商品小类与其他品类的共现频率向量
            - 目的：发现购买行为相似的品类群
            
            **套餐生成逻辑：**
            1. K-Means聚类将品类分组
            2. 每组内按共现次数排序
            3. 选择组内共现最强的组合生成套餐
            """)
        else:
            st.warning("模型文件未找到")


# ============================================================
# RUN
# ============================================================
if __name__ == "__main__":
    main()
