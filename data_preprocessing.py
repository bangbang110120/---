# -*- coding: utf-8 -*-
"""
日化美妆数据预处理脚本
功能：数据清洗、数据合成、特征工程
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')


# ============================================================
# 数据加载函数
# ============================================================

def load_tmall_data(filepath):
    """加载天猫订单数据"""
    print("正在加载 tmall_order_report.xlsx...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            '订单编号': row[0],
            '总金额': row[1],
            '买家实际支付金额': row[2],
            '收货地址': str(row[3]).strip() if row[3] else '',
            '订单创建时间': row[4],
            '订单付款时间': row[5],
            '退款金额': row[6],
        })
    df = pd.DataFrame(rows)
    wb.close()
    print(f"  原始数据: {len(df)} 条记录")
    return df


def load_beauty_data(filepath):
    """加载双十一美妆数据"""
    print("正在加载 双十一淘宝美妆数据.xlsx...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            'update_time': row[0],
            'id': row[1],
            'title': row[2],
            'price': row[3],
            'sale_count': row[4],
            'comment_count': row[5],
            '店名': row[6],
        })
    df = pd.DataFrame(rows)
    wb.close()
    print(f"  原始数据: {len(df)} 条记录")
    return df


def load_rihua_data(filepath):
    """加载日化数据（销售订单表 + 商品信息表）"""
    print("正在加载 日化.xlsx...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # 销售订单表
    ws_sales = wb[wb.sheetnames[0]]
    sales_rows = []
    for row in ws_sales.iter_rows(min_row=2, values_only=True):
        sales_rows.append({
            '订单编码': row[0],
            '订单日期': row[1],
            '客户编码': row[2],
            '所在区域': row[3],
            '所在省份': row[4],
            '所在地市': row[5],
            '商品编号': row[6],
            '订购数量': row[7],
            '订购单价': row[8],
            '金额': row[9],
        })
    df_sales = pd.DataFrame(sales_rows)
    
    # 商品信息表
    ws_prod = wb[wb.sheetnames[1]]
    prod_rows = []
    for row in ws_prod.iter_rows(min_row=2, values_only=True):
        prod_rows.append({
            '商品编号': row[0],
            '商品名称': row[1],
            '商品小类': row[2],
            '商品大类': row[3],
            '销售单价': row[4],
        })
    df_prod = pd.DataFrame(prod_rows)
    wb.close()
    
    print(f"  销售订单表: {len(df_sales)} 条记录")
    print(f"  商品信息表: {len(df_prod)} 条记录")
    return df_sales, df_prod


# ============================================================
# 数据清洗函数
# ============================================================

def clean_tmall_data(df):
    """清洗天猫订单数据"""
    print("\n清洗 tmall_order_report 数据...")
    original_len = len(df)
    
    # 1. 处理缺失值
    # 订单编号不能为空
    df = df[df['订单编号'].notna()]
    
    # 金额缺失填充为0
    df['总金额'] = df['总金额'].fillna(0)
    df['买家实际支付金额'] = df['买家实际支付金额'].fillna(0)
    df['退款金额'] = df['退款金额'].fillna(0)
    
    # 收货地址缺失标记为"未知"
    df['收货地址'] = df['收货地址'].fillna('未知')
    df['收货地址'] = df['收货地址'].replace('', '未知')
    
    # 2. 处理异常值
    # 先进行类型转换（处理可能的字符串类型数值）
    df['总金额'] = pd.to_numeric(df['总金额'], errors='coerce').fillna(0)
    df['买家实际支付金额'] = pd.to_numeric(df['买家实际支付金额'], errors='coerce').fillna(0)
    df['退款金额'] = pd.to_numeric(df['退款金额'], errors='coerce').fillna(0)
    
    # 金额不能为负数
    df['总金额'] = df['总金额'].apply(lambda x: max(x, 0))
    df['买家实际支付金额'] = df['买家实际支付金额'].apply(lambda x: max(x, 0))
    df['退款金额'] = df['退款金额'].apply(lambda x: max(x, 0))
    
    # 退款金额不能超过实付金额（逻辑异常）
    df.loc[df['退款金额'] > df['买家实际支付金额'], '退款金额'] = df['买家实际支付金额']
    
    # 3. 去除重复订单
    df = df.drop_duplicates(subset=['订单编号'], keep='first')
    
    # 4. 类型转换
    df['订单编号'] = df['订单编号'].astype(str)
    df['总金额'] = df['总金额'].astype(float)
    df['买家实际支付金额'] = df['买家实际支付金额'].astype(float)
    df['退款金额'] = df['退款金额'].astype(float)
    
    # 时间字段转换
    df['订单创建时间'] = pd.to_datetime(df['订单创建时间'], errors='coerce')
    df['订单付款时间'] = pd.to_datetime(df['订单付款时间'], errors='coerce')
    
    # 5. 地址标准化（统一省份名称）
    province_mapping = {
        '北京': '北京市', '上海': '上海市', '天津': '天津市', '重庆': '重庆市',
        '内蒙古': '内蒙古自治区', '广西': '广西壮族自治区', 
        '新疆': '新疆维吾尔自治区', '宁夏': '宁夏回族自治区', '西藏': '西藏自治区',
        '香港': '香港特别行政区', '澳门': '澳门特别行政区',
    }
    df['收货地址'] = df['收货地址'].apply(
        lambda x: province_mapping.get(x, x) if x in province_mapping else x
    )
    
    print(f"  清洗后: {len(df)} 条记录 (删除 {original_len - len(df)} 条异常)")
    print(f"  缺失值统计:")
    print(f"    - 订单创建时间缺失: {df['订单创建时间'].isna().sum()}")
    print(f"    - 订单付款时间缺失: {df['订单付款时间'].isna().sum()}")
    
    return df


def clean_beauty_data(df):
    """清洗双十一美妆数据"""
    print("\n清洗 双十一淘宝美妆数据...")
    original_len = len(df)
    
    # 1. 处理缺失值
    # 商品ID不能为空
    df = df[df['id'].notna()]
    
    # 价格、销量、评论缺失填充为0
    df['price'] = df['price'].fillna(0)
    df['sale_count'] = df['sale_count'].fillna(0)
    df['comment_count'] = df['comment_count'].fillna(0)
    
    # 标题缺失标记为"未知商品"
    df['title'] = df['title'].fillna('未知商品')
    
    # 店名缺失标记为"未知品牌"
    df['店名'] = df['店名'].fillna('未知品牌')
    
    # 2. 处理异常值
    # 先进行类型转换（处理可能的字符串类型数值）
    df['price'] = pd.to_numeric(df['price'], errors='coerce').fillna(0)
    df['sale_count'] = pd.to_numeric(df['sale_count'], errors='coerce').fillna(0)
    df['comment_count'] = pd.to_numeric(df['comment_count'], errors='coerce').fillna(0)
    
    # 价格不能为负数或过高（超过10000可能是数据错误）
    df['price'] = df['price'].apply(lambda x: max(0, min(x, 10000)))
    
    # 销量、评论数不能为负数
    df['sale_count'] = df['sale_count'].apply(lambda x: max(0, x))
    df['comment_count'] = df['comment_count'].apply(lambda x: max(0, x))
    
    # 评论数不能超过销量（逻辑异常）
    df.loc[df['comment_count'] > df['sale_count'], 'comment_count'] = df['sale_count']
    
    # 3. 去除重复商品（同ID同店名）
    df = df.drop_duplicates(subset=['id', '店名'], keep='first')
    
    # 4. 类型转换
    df['id'] = df['id'].astype(str)
    df['price'] = df['price'].astype(float)
    df['sale_count'] = df['sale_count'].astype(int)
    df['comment_count'] = df['comment_count'].astype(int)
    df['店名'] = df['店名'].astype(str).str.strip()
    df['title'] = df['title'].astype(str).str.strip()
    
    # 时间转换
    df['update_time'] = pd.to_datetime(df['update_time'], errors='coerce')
    
    print(f"  清洗后: {len(df)} 条记录 (删除 {original_len - len(df)} 条异常)")
    
    return df


def clean_rihua_data(df_sales, df_prod):
    """清洗日化销售数据"""
    print("\n清洗 日化数据...")
    original_sales_len = len(df_sales)
    
    # ===== 销售订单表清洗 =====
    # 1. 处理缺失值
    # 订单编码不能为空
    df_sales = df_sales[df_sales['订单编码'].notna()]
    
    # 客户编码、商品编号缺失标记为"未知"
    df_sales['客户编码'] = df_sales['客户编码'].fillna('未知客户')
    df_sales['商品编号'] = df_sales['商品编号'].fillna('未知商品')
    
    # 数量、单价、金额缺失填充为0
    df_sales['订购数量'] = df_sales['订购数量'].fillna(0)
    df_sales['订购单价'] = df_sales['订购单价'].fillna(0)
    df_sales['金额'] = df_sales['金额'].fillna(0)
    
    # 地区信息缺失标记为"未知"
    df_sales['所在区域'] = df_sales['所在区域'].fillna('未知区域')
    df_sales['所在省份'] = df_sales['所在省份'].fillna('未知省份')
    df_sales['所在地市'] = df_sales['所在地市'].fillna('未知地市')
    
    # 2. 处理异常值
    # 去除未来日期的订单（如2050年等明显错误）
    df_sales['订单日期'] = pd.to_datetime(df_sales['订单日期'], errors='coerce')
    df_sales = df_sales[df_sales['订单日期'].notna()]
    df_sales = df_sales[df_sales['订单日期'].apply(lambda x: x.year < 2030 if pd.notna(x) else False)]
    
    # 先进行类型转换（处理可能的字符串类型数值）
    df_sales['订购数量'] = pd.to_numeric(df_sales['订购数量'], errors='coerce').fillna(0)
    df_sales['订购单价'] = pd.to_numeric(df_sales['订购单价'], errors='coerce').fillna(0)
    df_sales['金额'] = pd.to_numeric(df_sales['金额'], errors='coerce').fillna(0)
    
    # 数量、金额不能为负数
    df_sales['订购数量'] = df_sales['订购数量'].apply(lambda x: max(0, x))
    df_sales['订购单价'] = df_sales['订购单价'].apply(lambda x: max(0, x))
    df_sales['金额'] = df_sales['金额'].apply(lambda x: max(0, x))
    
    # 金额验证：金额 ≈ 数量 × 单价（误差超过50%视为异常，重新计算）
    calculated_amount = df_sales['订购数量'] * df_sales['订购单价']
    amount_diff_ratio = abs(df_sales['金额'] - calculated_amount) / df_sales['金额'].replace(0, 1)
    df_sales.loc[amount_diff_ratio > 0.5, '金额'] = calculated_amount
    
    # 3. 去除完全重复的订单行
    df_sales = df_sales.drop_duplicates(keep='first')
    
    # 4. 类型转换
    df_sales['订单编码'] = df_sales['订单编码'].astype(str).str.strip()
    df_sales['客户编码'] = df_sales['客户编码'].astype(str).str.strip()
    df_sales['商品编号'] = df_sales['商品编号'].astype(str).str.strip()
    df_sales['所在区域'] = df_sales['所在区域'].astype(str).str.strip()
    df_sales['所在省份'] = df_sales['所在省份'].astype(str).str.strip()
    df_sales['所在地市'] = df_sales['所在地市'].astype(str).str.strip()
    df_sales['订购数量'] = df_sales['订购数量'].astype(int)
    df_sales['订购单价'] = df_sales['订购单价'].astype(float)
    df_sales['金额'] = df_sales['金额'].astype(float)
    
    print(f"  销售订单表清洗后: {len(df_sales)} 条 (删除 {original_sales_len - len(df_sales)} 条)")
    
    # ===== 商品信息表清洗 =====
    original_prod_len = len(df_prod)
    
    # 1. 处理缺失值
    df_prod = df_prod[df_prod['商品编号'].notna()]
    df_prod['商品名称'] = df_prod['商品名称'].fillna('未知商品')
    df_prod['商品小类'] = df_prod['商品小类'].fillna('未知小类')
    df_prod['商品大类'] = df_prod['商品大类'].fillna('未知大类')
    # 先进行类型转换，再填充缺失值
    df_prod['销售单价'] = pd.to_numeric(df_prod['销售单价'], errors='coerce').fillna(0)
    df_prod['销售单价'] = df_prod['销售单价'].apply(lambda x: max(0, x))
    
    # 2. 去除重复商品
    df_prod = df_prod.drop_duplicates(subset=['商品编号'], keep='first')
    
    # 3. 类型转换
    df_prod['商品编号'] = df_prod['商品编号'].astype(str).str.strip()
    df_prod['商品名称'] = df_prod['商品名称'].astype(str).str.strip()
    df_prod['商品小类'] = df_prod['商品小类'].astype(str).str.strip()
    df_prod['商品大类'] = df_prod['商品大类'].astype(str).str.strip()
    df_prod['销售单价'] = df_prod['销售单价'].astype(float)
    
    print(f"  商品信息表清洗后: {len(df_prod)} 条 (删除 {original_prod_len - len(df_prod)} 条)")
    
    return df_sales, df_prod


# ============================================================
# 数据合成与特征工程
# ============================================================

def feature_engineering_tmall(df):
    """天猫订单数据特征工程"""
    print("\n特征工程 - tmall_order_report...")
    
    # 1. 时间特征
    df['下单日期'] = df['订单创建时间'].dt.date
    df['下单小时'] = df['订单创建时间'].dt.hour
    df['下单星期'] = df['订单创建时间'].dt.weekday  # 0=周一, 6=周日
    df['是否周末'] = df['下单星期'].apply(lambda x: 1 if x >= 5 else 0)
    df['下单月份'] = df['订单创建时间'].dt.month
    
    # 2. 支付行为特征
    df['付款延迟(小时)'] = df.apply(
        lambda row: (row['订单付款时间'] - row['订单创建时间']).total_seconds() / 3600 
        if pd.notna(row['订单付款时间']) and pd.notna(row['订单创建时间']) else None,
        axis=1
    )
    df['付款延迟(小时)'] = df['付款延迟(小时)'].fillna(-1)  # -1表示未付款
    
    df['是否付款'] = df['买家实际支付金额'].apply(lambda x: 1 if x > 0 else 0)
    df['是否退款'] = df['退款金额'].apply(lambda x: 1 if x > 0 else 0)
    
    # 3. 金额特征
    df['实际收入'] = df['买家实际支付金额'] - df['退款金额']
    df['退款比例'] = df.apply(
        lambda row: row['退款金额'] / row['买家实际支付金额'] if row['买家实际支付金额'] > 0 else 0,
        axis=1
    )
    
    # 4. 金额分段
    df['金额段'] = pd.cut(
        df['总金额'],
        bins=[0, 50, 100, 200, 500, 1000, float('inf')],
        labels=['0-50', '50-100', '100-200', '200-500', '500-1000', '1000+']
    )
    
    # 5. 时段分类
    df['下单时段'] = pd.cut(
        df['下单小时'],
        bins=[-1, 6, 12, 18, 24],
        labels=['凌晨(0-6)', '上午(6-12)', '下午(12-18)', '晚上(18-24)']
    )
    
    print(f"  新增特征: 下单日期、下单小时、下单星期、是否周末、付款延迟、是否付款、是否退款、实际收入、退款比例、金额段、下单时段")
    
    return df


def feature_engineering_beauty(df):
    """双十一美妆数据特征工程"""
    print("\n特征工程 - 双十一淘宝美妆数据...")
    
    # 1. GMV计算
    df['GMV'] = df['price'] * df['sale_count']
    
    # 2. 评论率
    df['评论率'] = df.apply(
        lambda row: row['comment_count'] / row['sale_count'] if row['sale_count'] > 0 else 0,
        axis=1
    )
    
    # 3. 价格段分类
    df['价格段'] = pd.cut(
        df['price'],
        bins=[0, 50, 100, 200, 500, 1000, float('inf')],
        labels=['0-50元', '50-100元', '100-200元', '200-500元', '500-1000元', '1000元以上']
    )
    
    # 4. 是否爆款（销量超过10000）
    df['是否爆款'] = df['sale_count'].apply(lambda x: 1 if x >= 10000 else 0)
    
    # 5. 商品标题关键词提取
    def extract_keywords(title):
        keywords = []
        keyword_map = {
            '面膜': '面膜', '面霜': '面霜', '精华': '精华', '洁面': '洁面',
            '爽肤水': '爽肤水', '眼霜': '眼霜', '防晒': '防晒', '口红': '口红',
            '粉底': '粉底', '眼影': '眼影', '睫毛': '睫毛膏', '蜜粉': '蜜粉',
            '乳液': '乳液', '卸妆': '卸妆', '隔离': '隔离霜', '喷雾': '喷雾',
        }
        for kw, label in keyword_map.items():
            if kw in str(title):
                keywords.append(label)
        return keywords[0] if keywords else '其他'
    
    df['商品品类'] = df['title'].apply(extract_keywords)
    
    # 6. 品牌级别分类（按均价）
    brand_avg_price = df.groupby('店名')['price'].mean()
    df['品牌均价'] = df['店名'].map(brand_avg_price)
    df['品牌级别'] = pd.cut(
        df['品牌均价'],
        bins=[0, 100, 200, 500, float('inf')],
        labels=['大众品牌', '中端品牌', '高端品牌', '奢侈品牌']
    )
    
    print(f"  新增特征: GMV、评论率、价格段、是否爆款、商品品类、品牌均价、品牌级别")
    
    return df


def feature_engineering_rihua(df_sales, df_prod):
    """日化销售数据特征工程"""
    print("\n特征工程 - 日化数据...")
    
    # 1. 合并商品信息
    df_merged = df_sales.merge(
        df_prod[['商品编号', '商品名称', '商品小类', '商品大类', '销售单价']],
        on='商品编号',
        how='left'
    )
    
    # 2. 时间特征
    df_merged['订单年份'] = df_merged['订单日期'].dt.year
    df_merged['订单月份'] = df_merged['订单日期'].dt.month
    df_merged['订单季度'] = df_merged['订单日期'].dt.quarter
    df_merged['订单星期'] = df_merged['订单日期'].dt.weekday
    df_merged['是否周末'] = df_merged['订单星期'].apply(lambda x: 1 if x >= 5 else 0)
    
    # 3. 金额特征
    df_merged['单价差异'] = df_merged['订购单价'] - df_merged['销售单价']
    df_merged['折扣率'] = df_merged.apply(
        lambda row: row['订购单价'] / row['销售单价'] if row['销售单价'] > 0 else 1,
        axis=1
    )
    
    # 4. 区域编码
    region_order = {'北区': 1, '南区': 2, '东区': 3, '西区': 4}
    df_merged['区域编码'] = df_merged['所在区域'].map(region_order).fillna(0)
    
    # 5. 客单价计算（按订单）
    order_amount = df_merged.groupby('订单编码')['金额'].sum().reset_index()
    order_amount.columns = ['订单编码', '订单总金额']
    df_merged = df_merged.merge(order_amount, on='订单编码', how='left')
    
    print(f"  合并后数据: {len(df_merged)} 条")
    print(f"  新增特征: 订单年份、订单月份、订单季度、是否周末、单价差异、折扣率、区域编码、订单总金额")
    
    return df_merged


def compute_rfm(df_sales):
    """计算客户RFM指标"""
    print("\n计算客户RFM指标...")
    
    # 取最近日期作为参照
    reference_date = df_sales['订单日期'].max() + pd.Timedelta(days=1)
    
    # 按客户聚合
    rfm = df_sales.groupby('客户编码').agg({
        '订单日期': lambda x: (reference_date - x.max()).days,  # Recency
        '订单编码': 'nunique',  # Frequency
        '金额': 'sum',  # Monetary
    }).reset_index()
    
    rfm.columns = ['客户编码', 'Recency', 'Frequency', 'Monetary']
    
    # RFM评分（4分制）
    rfm['R_Score'] = pd.qcut(rfm['Recency'], 4, labels=[4, 3, 2, 1]).astype(int)
    rfm['F_Score'] = pd.qcut(rfm['Frequency'].rank(method='first'), 4, labels=[1, 2, 3, 4]).astype(int)
    rfm['M_Score'] = pd.qcut(rfm['Monetary'].rank(method='first'), 4, labels=[1, 2, 3, 4]).astype(int)
    
    # RFM总分
    rfm['RFM总分'] = rfm['R_Score'] + rfm['F_Score'] + rfm['M_Score']
    
    # 客户分层
    def classify_customer(r, f, m):
        total = r + f + m
        if total >= 10:
            return '核心客户'
        elif total >= 7:
            return '重要客户'
        elif total >= 5:
            return '一般客户'
        else:
            return '流失风险'
    
    rfm['客户层级'] = rfm.apply(lambda x: classify_customer(x['R_Score'], x['F_Score'], x['M_Score']), axis=1)
    
    # 是否流失标签（用于预测模型）
    rfm['是否流失'] = rfm['客户层级'].apply(lambda x: 1 if x == '流失风险' else 0)
    
    print(f"  RFM客户数: {len(rfm)}")
    print(f"  客户层级分布:")
    print(rfm['客户层级'].value_counts().to_string())
    
    return rfm


def compute_monthly_sales(df_sales):
    """计算月度销售统计（用于销量预测）"""
    print("\n计算月度销售统计...")
    
    # 添加月份标识
    df_sales['月份'] = df_sales['订单日期'].dt.to_period('M')
    
    # 按月份聚合
    monthly = df_sales.groupby('月份').agg({
        '金额': 'sum',
        '订单编码': 'nunique',
        '客户编码': 'nunique',
        '订购数量': 'sum',
    }).reset_index()
    
    monthly.columns = ['月份', '销售额', '订单数', '客户数', '销量']
    monthly['月份'] = monthly['月份'].astype(str)
    
    # 环比增长
    monthly['销售额环比增长'] = monthly['销售额'].pct_change()
    monthly['销量环比增长'] = monthly['销量'].pct_change()
    
    # 同比增长（需要跨年数据）
    monthly['年份'] = monthly['月份'].apply(lambda x: int(x.split('-')[0]))
    monthly['月'] = monthly['月份'].apply(lambda x: int(x.split('-')[1]))
    
    print(f"  月度数据: {len(monthly)} 个月")
    
    return monthly


# ============================================================
# 数据导出函数
# ============================================================

def export_cleaned_data(df_tmall, df_beauty, df_rihua_merged, rfm_df, monthly_df, output_dir):
    """导出清洗后的数据"""
    print(f"\n导出清洗后数据到 {output_dir}...")
    
    # 导出为CSV
    df_tmall.to_csv(f"{output_dir}/tmall_cleaned.csv", index=False, encoding='utf-8-sig')
    df_beauty.to_csv(f"{output_dir}/beauty_cleaned.csv", index=False, encoding='utf-8-sig')
    df_rihua_merged.to_csv(f"{output_dir}/rihua_merged.csv", index=False, encoding='utf-8-sig')
    rfm_df.to_csv(f"{output_dir}/customer_rfm.csv", index=False, encoding='utf-8-sig')
    monthly_df.to_csv(f"{output_dir}/monthly_sales.csv", index=False, encoding='utf-8-sig')
    
    print("  导出完成:")
    print(f"    - tmall_cleaned.csv: {len(df_tmall)} 条")
    print(f"    - beauty_cleaned.csv: {len(df_beauty)} 条")
    print(f"    - rihua_merged.csv: {len(df_rihua_merged)} 条")
    print(f"    - customer_rfm.csv: {len(rfm_df)} 条")
    print(f"    - monthly_sales.csv: {len(monthly_df)} 条")


# ============================================================
# 主函数
# ============================================================

def main():
    """数据预处理主流程"""
    import os
    
    # 获取数据文件路径
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, 'cleaned_data')
    
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    
    print("=" * 60)
    print("日化美妆数据预处理")
    print("=" * 60)
    
    # ========== 1. 数据加载 ==========
    print("\n【步骤1】数据加载")
    print("-" * 40)
    
    df_tmall = load_tmall_data(os.path.join(base_dir, 'tmall_order_report.xlsx'))
    df_beauty = load_beauty_data(os.path.join(base_dir, '双十一淘宝美妆数据.xlsx'))
    df_sales, df_prod = load_rihua_data(os.path.join(base_dir, '日化.xlsx'))
    
    # ========== 2. 数据清洗 ==========
    print("\n【步骤2】数据清洗")
    print("-" * 40)
    
    df_tmall_clean = clean_tmall_data(df_tmall)
    df_beauty_clean = clean_beauty_data(df_beauty)
    df_sales_clean, df_prod_clean = clean_rihua_data(df_sales, df_prod)
    
    # ========== 3. 特征工程 ==========
    print("\n【步骤3】特征工程")
    print("-" * 40)
    
    df_tmall_features = feature_engineering_tmall(df_tmall_clean)
    df_beauty_features = feature_engineering_beauty(df_beauty_clean)
    df_rihua_merged = feature_engineering_rihua(df_sales_clean, df_prod_clean)
    
    # ========== 4. 数据合成 ==========
    print("\n【步骤4】数据合成")
    print("-" * 40)
    
    rfm_df = compute_rfm(df_sales_clean)
    monthly_df = compute_monthly_sales(df_sales_clean)
    
    # ========== 5. 数据导出 ==========
    print("\n【步骤5】数据导出")
    print("-" * 40)
    
    export_cleaned_data(
        df_tmall_features, 
        df_beauty_features, 
        df_rihua_merged, 
        rfm_df, 
        monthly_df,
        output_dir
    )
    
    # ========== 6. 数据质量报告 ==========
    print("\n【步骤6】数据质量报告")
    print("-" * 40)
    
    print("\n数据质量摘要:")
    print(f"  tmall数据:")
    print(f"    - 有效订单: {len(df_tmall_features)}")
    print(f"    - 付款率: {df_tmall_features['是否付款'].mean()*100:.1f}%")
    print(f"    - 退款率: {df_tmall_features['是否退款'].mean()*100:.1f}%")
    
    print(f"\n  双十一美妆数据:")
    print(f"    - 有效SKU: {len(df_beauty_features)}")
    print(f"    - 品牌数: {df_beauty_features['店名'].nunique()}")
    print(f"    - 总GMV: ¥{df_beauty_features['GMV'].sum()/1e8:.2f}亿")
    print(f"    - 爆款比例: {df_beauty_features['是否爆款'].mean()*100:.1f}%")
    
    print(f"\n  日化渠道数据:")
    print(f"    - 有效订单: {df_rihua_merged['订单编码'].nunique()}")
    print(f"    - 客户数: {rfm_df['客户编码'].nunique()}")
    print(f"    - 总销售额: ¥{df_rihua_merged['金额'].sum()/1e8:.2f}亿")
    print(f"    - 核心客户占比: {(rfm_df['客户层级']=='核心客户').mean()*100:.1f}%")
    print(f"    - 流失风险占比: {(rfm_df['客户层级']=='流失风险').mean()*100:.1f}%")
    
    print("\n" + "=" * 60)
    print("数据预处理完成！")
    print("=" * 60)
    
    return df_tmall_features, df_beauty_features, df_rihua_merged, rfm_df, monthly_df


if __name__ == "__main__":
    df_tmall, df_beauty, df_rihua, rfm, monthly = main()