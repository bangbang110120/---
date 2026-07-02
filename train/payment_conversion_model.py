# -*- coding: utf-8 -*-
"""
订单支付转化预测模型（弃单挽回预测）

预测目标(Y): 未支付订单最终是否会付款（1=会付，0=死单）
业务场景: 客服资源有限，确定哪些未付款订单值得去追

模型列表:
- 基础模型: XGBoost, RandomForest, LogisticRegression, GradientBoosting
- 高级模型: PyTorch DNN (支持GPU), LightGBM
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
import os
import joblib
import json
import warnings
warnings.filterwarnings('ignore')

# 尝试导入必要的库
try:
    import xgboost as xgb
    from xgboost import XGBClassifier
except ImportError:
    print("请先安装 xgboost: pip install xgboost")

try:
    import lightgbm as lgb
    from lightgbm import LGBMClassifier
    HAS_LIGHTGBM = True
except ImportError:
    print("未安装 lightgbm: pip install lightgbm")
    HAS_LIGHTGBM = False

try:
    import torch
    import torch.nn as nn
    import torch.optim as optim
    from torch.utils.data import DataLoader, TensorDataset
    HAS_TORCH = True
    DEVICE = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"PyTorch 设备: {DEVICE}")
    if torch.cuda.is_available():
        print(f"  GPU: {torch.cuda.get_device_name(0)}")
except ImportError:
    print("未安装 torch: pip install torch")
    HAS_TORCH = False
    DEVICE = None

try:
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    matplotlib.rcParams['axes.unicode_minus'] = False
except ImportError:
    print("请先安装 matplotlib: pip install matplotlib")

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    confusion_matrix, roc_auc_score, roc_curve
)


# ============================================================
# 配置路径
# ============================================================

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MODEL_DIR = os.path.join(BASE_DIR, 'models')
METRICS_DIR = os.path.join(BASE_DIR, 'model_metrics')
CHARTS_DIR = os.path.join(BASE_DIR, 'model_charts')

for dir_path in [MODEL_DIR, METRICS_DIR, CHARTS_DIR]:
    os.makedirs(dir_path, exist_ok=True)


# ============================================================
# PyTorch 深度神经网络
# ============================================================

class PaymentDNN(nn.Module):
    """深度神经网络分类器"""
    def __init__(self, input_dim, hidden_dims=[128, 64, 32], dropout_rate=0.3):
        super(PaymentDNN, self).__init__()
        
        layers = []
        prev_dim = input_dim
        
        for hidden_dim in hidden_dims:
            layers.append(nn.Linear(prev_dim, hidden_dim))
            layers.append(nn.BatchNorm1d(hidden_dim))
            layers.append(nn.ReLU())
            layers.append(nn.Dropout(dropout_rate))
            prev_dim = hidden_dim
        
        layers.append(nn.Linear(prev_dim, 1))
        layers.append(nn.Sigmoid())
        
        self.network = nn.Sequential(*layers)
    
    def forward(self, x):
        return self.network(x)


class PyTorchClassifier:
    """PyTorch 分类器包装类"""
    def __init__(self, input_dim, hidden_dims=[128, 64, 32], dropout_rate=0.3, 
                 learning_rate=0.001, epochs=100, batch_size=64, device=None):
        self.input_dim = input_dim
        self.hidden_dims = hidden_dims
        self.dropout_rate = dropout_rate
        self.learning_rate = learning_rate
        self.epochs = epochs
        self.batch_size = batch_size
        self.device = device if device else (DEVICE if HAS_TORCH else 'cpu')
        self.model = None
        self.scaler = StandardScaler()
    
    def fit(self, X, y):
        # 标准化
        X_scaled = self.scaler.fit_transform(X)
        
        # 转换为 Tensor
        X_tensor = torch.FloatTensor(X_scaled).to(self.device)
        y_tensor = torch.FloatTensor(y.values if isinstance(y, pd.Series) else y).reshape(-1, 1).to(self.device)
        
        # 创建 DataLoader
        dataset = TensorDataset(X_tensor, y_tensor)
        loader = DataLoader(dataset, batch_size=self.batch_size, shuffle=True)
        
        # 初始化模型
        self.model = PaymentDNN(self.input_dim, self.hidden_dims, self.dropout_rate).to(self.device)
        criterion = nn.BCELoss()
        optimizer = optim.Adam(self.model.parameters(), lr=self.learning_rate, weight_decay=1e-5)
        
        # 训练
        for epoch in range(self.epochs):
            self.model.train()
            total_loss = 0
            
            for batch_X, batch_y in loader:
                optimizer.zero_grad()
                outputs = self.model(batch_X)
                loss = criterion(outputs, batch_y)
                loss.backward()
                optimizer.step()
                total_loss += loss.item()
            
            if (epoch + 1) % 20 == 0:
                avg_loss = total_loss / len(loader)
                print(f"    Epoch {epoch+1}/{self.epochs}, Loss: {avg_loss:.4f}")
    
    def predict(self, X):
        X_scaled = self.scaler.transform(X)
        X_tensor = torch.FloatTensor(X_scaled).to(self.device)
        
        self.model.eval()
        with torch.no_grad():
            probs = self.model(X_tensor).cpu().numpy().flatten()
        
        return (probs > 0.5).astype(int)
    
    def predict_proba(self, X):
        X_scaled = self.scaler.transform(X)
        X_tensor = torch.FloatTensor(X_scaled).to(self.device)
        
        self.model.eval()
        with torch.no_grad():
            probs = self.model(X_tensor).cpu().numpy().flatten()
        
        return np.column_stack([1 - probs, probs])
    
    def save(self, filepath):
        """保存模型"""
        torch.save({
            'model_state_dict': self.model.state_dict(),
            'scaler': self.scaler,
            'input_dim': self.input_dim,
            'hidden_dims': self.hidden_dims,
            'dropout_rate': self.dropout_rate
        }, filepath)
    
    def load(self, filepath):
        """加载模型"""
        checkpoint = torch.load(filepath, map_location=self.device)
        self.scaler = checkpoint['scaler']
        self.model = PaymentDNN(
            checkpoint['input_dim'], 
            checkpoint['hidden_dims'], 
            checkpoint['dropout_rate']
        ).to(self.device)
        self.model.load_state_dict(checkpoint['model_state_dict'])


# ============================================================
# 数据加载
# ============================================================

def safe_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        s = str(v).replace('元','').replace('¥','').replace(',','').replace(' ','').strip()
        return float(s) if s else 0.0
    except:
        return 0.0


def load_tmall_data(filepath):
    print("正在加载 tmall_order_report.xlsx...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            '订单编号': row[0],
            '总金额': safe_float(row[1]),
            '买家实际支付金额': safe_float(row[2]),
            '收货地址': str(row[3]).strip() if row[3] else '',
            '订单创建时间': row[4],
            '订单付款时间': row[5],
            '退款金额': safe_float(row[6]),
        })
    df = pd.DataFrame(rows)
    wb.close()
    print(f"  加载完成: {len(df)} 条记录")
    return df


# ============================================================
# 特征工程
# ============================================================

def get_time_period(hour):
    if 0 <= hour < 6:
        return '凌晨'
    elif 6 <= hour < 12:
        return '上午'
    elif 12 <= hour < 14:
        return '午间'
    elif 14 <= hour < 18:
        return '下午'
    elif 18 <= hour < 22:
        return '晚间'
    else:
        return '深夜'


def extract_features(df):
    df['订单创建时间'] = pd.to_datetime(df['订单创建时间'], errors='coerce')
    df['订单付款时间'] = pd.to_datetime(df['订单付款时间'], errors='coerce')
    df = df[df['订单创建时间'].notna()].copy()
    df['是否付款'] = df['订单付款时间'].notna().astype(int)
    
    # 时间特征
    df['下单小时'] = df['订单创建时间'].dt.hour
    df['下单星期'] = df['订单创建时间'].dt.weekday
    df['是否周末'] = (df['下单星期'] >= 5).astype(int)
    df['时段'] = df['下单小时'].apply(get_time_period)
    
    # 金额特征
    df['金额区间'] = df['总金额'].apply(
        lambda x: 0 if x <= 50 else (1 if x <= 100 else (2 if x <= 200 else (3 if x <= 500 else (4 if x <= 1000 else 5))))
    )
    
    # 历史转化率特征
    conversion_stats = df.groupby(['金额区间', '时段']).agg({
        '是否付款': ['sum', 'count']
    }).reset_index()
    conversion_stats.columns = ['金额区间', '时段', '付款数', '订单数']
    conversion_stats['历史支付率'] = conversion_stats['付款数'] / conversion_stats['订单数']
    
    conversion_map = {}
    for _, row in conversion_stats.iterrows():
        conversion_map[(row['金额区间'], row['时段'])] = row['历史支付率']
    
    df['历史支付率'] = df.apply(lambda r: conversion_map.get((r['金额区间'], r['时段']), 0.5), axis=1)
    
    # 地域特征
    df['省份'] = df['收货地址'].apply(lambda x: x.split()[0] if isinstance(x, str) else x)
    prov_conversion = df.groupby('省份')['是否付款'].mean().to_dict()
    df['省份支付率'] = df['省份'].map(prov_conversion).fillna(0.5)
    
    le_province = LabelEncoder()
    df['省份编码'] = le_province.fit_transform(df['省份'].astype(str))
    
    return df, le_province, conversion_map, prov_conversion


# ============================================================
# 模型训练
# ============================================================

def train_models(X_train, y_train, X_test, y_test):
    """训练多个模型并对比"""
    results = {}
    
    # 1. 基础模型
    print("\n【基础模型训练】")
    
    basic_models = {
        'XGBoost': XGBClassifier(
            n_estimators=100, max_depth=6, learning_rate=0.1,
            use_label_encoder=False, eval_metric='logloss', random_state=42
        ),
        'RandomForest': RandomForestClassifier(
            n_estimators=100, max_depth=10, random_state=42, class_weight='balanced'
        ),
        'LogisticRegression': LogisticRegression(
            max_iter=1000, class_weight='balanced', random_state=42
        ),
        'GradientBoosting': GradientBoostingClassifier(
            n_estimators=100, max_depth=6, random_state=42
        ),
    }
    
    for name, model in basic_models.items():
        print(f"\n  训练 {name}...")
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        y_pred_proba = model.predict_proba(X_test)[:, 1]
        
        metrics = {
            'accuracy': accuracy_score(y_test, y_pred),
            'precision': precision_score(y_test, y_pred, zero_division=0),
            'recall': recall_score(y_test, y_pred, zero_division=0),
            'f1': f1_score(y_test, y_pred, zero_division=0),
            'auc': roc_auc_score(y_test, y_pred_proba)
        }
        
        results[name] = {
            'model': model,
            'metrics': metrics,
            'confusion_matrix': confusion_matrix(y_test, y_pred),
            'y_pred_proba': y_pred_proba,
            'is_pytorch': False
        }
        
        print(f"    AUC: {metrics['auc']:.4f}, F1: {metrics['f1']:.4f}")
    
    # 2. LightGBM（高级）
    if HAS_LIGHTGBM:
        print("\n【高级模型训练 - LightGBM】")
        print(f"\n  训练 LightGBM...")
        
        lgb_model = LGBMClassifier(
            n_estimators=200,
            max_depth=8,
            learning_rate=0.05,
            num_leaves=31,
            min_child_samples=20,
            subsample=0.8,
            colsample_bytree=0.8,
            random_state=42,
            verbose=-1
        )
        
        lgb_model.fit(X_train, y_train)
        y_pred = lgb_model.predict(X_test)
        y_pred_proba = lgb_model.predict_proba(X_test)[:, 1]
        
        metrics = {
            'accuracy': accuracy_score(y_test, y_pred),
            'precision': precision_score(y_test, y_pred, zero_division=0),
            'recall': recall_score(y_test, y_pred, zero_division=0),
            'f1': f1_score(y_test, y_pred, zero_division=0),
            'auc': roc_auc_score(y_test, y_pred_proba)
        }
        
        results['LightGBM'] = {
            'model': lgb_model,
            'metrics': metrics,
            'confusion_matrix': confusion_matrix(y_test, y_pred),
            'y_pred_proba': y_pred_proba,
            'is_pytorch': False
        }
        
        print(f"    AUC: {metrics['auc']:.4f}, F1: {metrics['f1']:.4f}")
    
    # 3. PyTorch DNN（高级）
    if HAS_TORCH:
        print("\n【高级模型训练 - PyTorch DNN】")
        print(f"  设备: {DEVICE}")
        print(f"\n  训练 DNN...")
        
        dnn_model = PyTorchClassifier(
            input_dim=X_train.shape[1],
            hidden_dims=[256, 128, 64, 32],
            dropout_rate=0.3,
            learning_rate=0.001,
            epochs=100,
            batch_size=64,
            device=DEVICE
        )
        
        dnn_model.fit(X_train, y_train)
        y_pred = dnn_model.predict(X_test)
        y_pred_proba = dnn_model.predict_proba(X_test)[:, 1]
        
        metrics = {
            'accuracy': accuracy_score(y_test, y_pred),
            'precision': precision_score(y_test, y_pred, zero_division=0),
            'recall': recall_score(y_test, y_pred, zero_division=0),
            'f1': f1_score(y_test, y_pred, zero_division=0),
            'auc': roc_auc_score(y_test, y_pred_proba)
        }
        
        results['DNN'] = {
            'model': dnn_model,
            'metrics': metrics,
            'confusion_matrix': confusion_matrix(y_test, y_pred),
            'y_pred_proba': y_pred_proba,
            'is_pytorch': True
        }
        
        print(f"    AUC: {metrics['auc']:.4f}, F1: {metrics['f1']:.4f}")
    
    return results


# ============================================================
# 可视化
# ============================================================

def plot_results(results, y_test, charts_dir):
    # ROC曲线对比
    plt.figure(figsize=(10, 8))
    for name, result in results.items():
        fpr, tpr, _ = roc_curve(y_test, result['y_pred_proba'])
        auc = result['metrics']['auc']
        plt.plot(fpr, tpr, label=f'{name} (AUC={auc:.3f})', linewidth=2)
    
    plt.plot([0, 1], [0, 1], 'k--', label='随机')
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('订单支付转化预测 ROC曲线对比')
    plt.legend(loc='lower right')
    plt.grid(True, alpha=0.3)
    plt.savefig(os.path.join(charts_dir, 'payment_conversion_roc.png'), dpi=150, bbox_inches='tight')
    plt.close()
    
    # 指标对比
    metrics_names = ['accuracy', 'precision', 'recall', 'f1', 'auc']
    fig, axes = plt.subplots(1, 5, figsize=(20, 4))
    model_names = list(results.keys())
    colors = ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6', '#1abc9c', '#e67e22']
    
    for i, metric_name in enumerate(metrics_names):
        values = [results[name]['metrics'].get(metric_name, 0) for name in model_names]
        bars = axes[i].bar(model_names, values, color=colors[:len(model_names)])
        axes[i].set_title(metric_name.upper())
        axes[i].set_ylim(0, 1)
        axes[i].tick_params(axis='x', rotation=45)
        for j, v in enumerate(values):
            axes[i].text(j, v + 0.02, f'{v:.3f}', ha='center', fontsize=8)
    
    plt.tight_layout()
    plt.savefig(os.path.join(charts_dir, 'payment_conversion_metrics.png'), dpi=150, bbox_inches='tight')
    plt.close()


# ============================================================
# 主函数
# ============================================================

def main():
    print("=" * 60)
    print("订单支付转化预测模型（弃单挽回预测）")
    print("=" * 60)
    
    # 1. 加载数据
    tmall_path = os.path.join(BASE_DIR, 'tmall_order_report.xlsx')
    df_tmall = load_tmall_data(tmall_path)
    
    # 2. 特征工程
    print("\n【特征工程】")
    df, le_province, conversion_map, prov_conversion = extract_features(df_tmall)
    
    features = ['总金额', '下单小时', '是否周末', '金额区间', 
                '历史支付率', '省份编码', '省份支付率']
    
    X = df[features].fillna(0).values
    y = df['是否付款'].values
    
    print(f"\n数据统计:")
    print(f"  总订单数: {len(df)}")
    print(f"  已付款: {y.sum()} ({y.sum()/len(y)*100:.1f}%)")
    print(f"  未付款: {len(y)-y.sum()} ({(len(y)-y.sum())/len(y)*100:.1f}%)")
    
    # 3. 数据划分
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42, stratify=y
    )
    
    print(f"\n数据划分:")
    print(f"  训练集: {len(X_train)} 条")
    print(f"  测试集: {len(X_test)} 条")
    
    # 4. 模型训练
    print("\n【模型训练】")
    results = train_models(X_train, y_train, X_test, y_test)
    
    # 5. 选择最佳模型
    print("\n【模型对比】")
    
    metrics_table = []
    for name, result in results.items():
        metrics_table.append({
            '模型': name,
            'AUC': result['metrics']['auc'],
            'F1': result['metrics']['f1'],
            'Precision': result['metrics']['precision'],
            'Recall': result['metrics']['recall']
        })
    
    metrics_df = pd.DataFrame(metrics_table).sort_values('AUC', ascending=False)
    print("\n" + metrics_df.to_string(index=False))
    
    best_model_name = metrics_df.iloc[0]['模型']
    best_result = results[best_model_name]
    
    print(f"\n最佳模型: {best_model_name} (AUC={best_result['metrics']['auc']:.4f})")
    
    # 6. 保存模型
    print("\n【保存模型】")
    
    # 保存最佳模型
    if best_result['is_pytorch']:
        best_result['model'].save(os.path.join(MODEL_DIR, 'payment_conversion_best.pt'))
        print(f"  最佳模型: payment_conversion_best.pt (PyTorch)")
    else:
        joblib.dump(best_result['model'], os.path.join(MODEL_DIR, 'payment_conversion_best.joblib'))
        print(f"  最佳模型: payment_conversion_best.joblib")
    
    # 保存所有模型
    for name, result in results.items():
        if result['is_pytorch']:
            result['model'].save(os.path.join(MODEL_DIR, f'payment_conversion_{name.lower()}.pt'))
        else:
            joblib.dump(result['model'], os.path.join(MODEL_DIR, f'payment_conversion_{name.lower()}.joblib'))
    
    # 保存编码器
    encoders = {
        'province_encoder': le_province,
        'conversion_map': conversion_map,
        'prov_conversion': prov_conversion,
        'features': features
    }
    joblib.dump(encoders, os.path.join(MODEL_DIR, 'payment_encoders.joblib'))
    
    # 7. 保存指标
    metrics_dict = {name: result['metrics'] for name, result in results.items()}
    
    with open(os.path.join(METRICS_DIR, 'payment_conversion_metrics.json'), 'w', encoding='utf-8') as f:
        json.dump(metrics_dict, f, indent=2, ensure_ascii=False)
    
    metrics_df.to_csv(os.path.join(METRICS_DIR, 'payment_conversion_metrics.csv'), 
                      encoding='utf-8-sig', index=False)
    
    # 8. 可视化
    print("\n【可视化】")
    plot_results(results, y_test, CHARTS_DIR)
    
    # 9. 特征重要性（树模型）
    tree_models = ['XGBoost', 'RandomForest', 'GradientBoosting', 'LightGBM']
    for name in tree_models:
        if name in results and hasattr(results[name]['model'], 'feature_importances_'):
            importance_df = pd.DataFrame({
                '特征': features,
                '重要性': results[name]['model'].feature_importances_
            }).sort_values('重要性', ascending=False)
            
            print(f"\n{name} 特征重要性:")
            for _, row in importance_df.iterrows():
                print(f"  {row['特征']}: {row['重要性']:.4f}")
    
    # 10. 总结
    print("\n" + "=" * 60)
    print("训练完成！")
    print("=" * 60)
    print(f"\n输出文件:")
    print(f"  模型: {MODEL_DIR}/payment_conversion_*.joblib / .pt")
    print(f"  指标: {METRICS_DIR}/payment_conversion_metrics.csv/json")
    print(f"  图表: {CHARTS_DIR}/payment_conversion_*.png")
    
    return results, encoders


if __name__ == "__main__":
    results, encoders = main()